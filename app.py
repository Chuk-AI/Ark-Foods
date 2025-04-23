from flask import (
    Flask,
    render_template,
    redirect,
    url_for,
    flash,
    request,
    session,
    jsonify,
    send_file,
    Blueprint

)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import and_, or_, func, text, case, true

from flask_login import (
    LoginManager,
    UserMixin,
    login_user,
    login_required,
    logout_user,
    current_user,
)
from werkzeug.security import generate_password_hash, check_password_hash
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField, SelectField
from wtforms.validators import InputRequired, Email, Length, EqualTo
from flask_bootstrap import Bootstrap
from datetime import timedelta, datetime
from flask_cors import CORS
import pytz
from pytz import timezone
import pandas as pd
from dateutil import parser
import gc  # garbage collection
import io
import xlsxwriter

# Flask-Admin Setup
from flask_admin import Admin
from flask_admin.contrib.sqla import ModelView

# For Scheduling
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
from apscheduler.triggers.cron import CronTrigger
import aiohttp
import asyncio
import os
import logging
import requests
import base64
import csv
from werkzeug.utils import secure_filename
import json
from dateutil.relativedelta import relativedelta


# Imports for Weather Forecasting
import time
import numpy as np
import copy
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
import logging

# IBM Environmental Intelligence Suite (EIS) related imports
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import ibmpairs.query as query
from ibmpairs.client import get_client
import logging
from functools import wraps
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity, get_jwt
from sqlalchemy.exc import SQLAlchemyError
from flask import send_from_directory
from notebook import get_best_start_dates, fetch_data_from_api
from fetch_shipping_point_data import fetch_shipping_point_data  # Replace with the correct module name
from flask_caching import Cache
import hashlib
from flask import request, make_response
from collections import defaultdict




# Configuration for Logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)


CSV_DIRECTORY = "data/"

# Initialize Flask app
app = Flask(__name__, static_folder= 'frontend/build', static_url_path="/")
app.config['JWT_SECRET_KEY'] = 'your_secret_key'  # Replace with a strong secret key
app.config['DEBUG'] = True
app.config['CACHE_NO_CACHE_ROUTES'] = [
    '/api/delete-alert-by-id',
    '/api/clear-alerts',
    '/api/alert-settings',
    '/api/alert-entries-fresh',
    '/api/break_even'
]
jwt = JWTManager(app)



@app.route("/api/normalize_price_data")
def normalize_price_data():
    try:
        records = PriceData.query.all()
        updated = 0

        for record in records:
            original_city = record.city_name
            original_commodity = record.commodity

            normalized_city = original_city.strip().lower().title()
            normalized_commodity = original_commodity.strip().lower().title()

            if original_city != normalized_city or original_commodity != normalized_commodity:
                record.city_name = normalized_city
                record.commodity = normalized_commodity
                updated += 1

        db.session.commit()
        return jsonify({"status": "success", "updated_records": updated})

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500




@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve(path):
    if path != "" and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    else:
        return send_from_directory(app.static_folder, "index.html")


# Configure Flask-Caching to use Redis
app.config['CACHE_TYPE'] = 'redis'
app.config['CACHE_REDIS_URL'] = 'redis://localhost:6379/0'  # Adjust if needed
# app.config['CACHE_REDIS_URL'] = os.environ.get("REDIS_URL", "redis://<MEMORISTORE_IP>:6379/0")

cache = Cache(app)



# Redis implementation for the caching of data coming from the db
def generate_cache_key():
    # Use the full path (which includes query parameters)
    key = request.full_path  # e.g., "/api/sales_dashboard?commodity=Jalapeno&source=USDA"
    # Optionally, you can hash it to ensure it's a consistent format
    return hashlib.md5(key.encode('utf-8')).hexdigest()


@app.before_request
def serve_from_cache():
    if request.method == 'GET':
        cache_key = generate_cache_key()
        cached_response = cache.get(cache_key)
        if cached_response:
            app.logger.info(f"Cache hit for key: {cache_key}")
            response = make_response(cached_response)
            response.headers["Content-Type"] = "application/json"
            return response
        else:
            app.logger.info(f"Cache miss for key: {cache_key}")

@app.after_request
def cache_response(response):
    if request.method == 'GET' and response.status_code == 200:
        cache_key = generate_cache_key()
        # time out 
        ttl = 86400  # Set TTL to 1 day (24 hours)
        # For more dynamic TTL, you can calculate it based on some parameters
        # e.g., setting a shorter TTL for endpoints that change often
        if 'historical_data' in request.full_path:
            ttl = 3600  # Shorter TTL of 1 hour for historical data
 

        cache.set(cache_key, response.get_data(), timeout=ttl)
        app.logger.info(f"Cached response for key: {cache_key}")
    return response




CORS(
    app,
    supports_credentials=True,
    resources={
        r"/api/*": {
            "origins": [
                "http://localhost:3000",  # Local frontend (for development)
                "https://arkfoods.klicksai.com"  # Replace with your production domain
            ],
            "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],  # Add OPTIONS
            "allow_headers": [
                "Content-Type", 
                "Authorization", 
                "Access-Control-Allow-Credentials"
            ]
        }
    },
)




@jwt.unauthorized_loader
def custom_unauthorized_response(err):
    return jsonify({"error": "Unauthorized. Please log in."}), 401

load_dotenv()

# setting up env variable for IBM API keys
# EIS_API_KEY = os.getenv("EIS_API_KEY")
# EIS_TENANT_ID = os.getenv("EIS_TENANT_ID")
# EIS_ORG_ID = os.getenv("EIS_ORG_ID")
# if not all([EIS_API_KEY, EIS_TENANT_ID, EIS_ORG_ID]):
#     logging.error(
#         "One or more IBM credentials are missing in the environment variables."
#     )
# print(f"API Key: {os.getenv('EIS_API_KEY')}")
# print(f"Tenant ID: {os.getenv('EIS_TENANT_ID')}")
# print(f"Endpoint: https://api.ibm.com/geospatial/run/na/core/v3/query")


# Enable CORS
CORS(app, supports_credentials=True, origins=["http://localhost:3000", "http://127.0.0.1"])


# Set the secret key for session handling


# app.config["SECRET_KEY"] = "your_secret_key_here"

app.config['JWT_BLACKLIST_ENABLED'] = True
app.config['JWT_BLACKLIST_TOKEN_CHECKS'] = ['access']
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=1)


blacklist = set()

@jwt.token_in_blocklist_loader
def check_if_token_in_blacklist(jwt_header, jwt_payload):
    return jwt_payload['jti'] in blacklist
# Set session cookie options
# app.config["SESSION_COOKIE_HTTPONLY"] = True  # Prevent access to cookies via JS
# app.config["SESSION_COOKIE_SAMESITE"] = "Lax"  # Adjust SameSite based on needs
# app.config["SESSION_COOKIE_SECURE"] = False   # Set to True in production with HTTPS


# Check if DATABASE_URL is present for Heroku's PostgreSQL, otherwise use SQLite for local development
DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL:
    # Replace the "postgres://" scheme with "postgresql://" for compatibility with SQLAlchemy
    app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL.replace(
        "postgres://", "postgresql://", 1
    )
else:
    # Use SQLite for local development
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(
        app.instance_path, "users.db"
    )

# Configure other Flask settings
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(
    days=1
)  # Session timeout after 1 day
# for uploading data
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {"csv"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Initialize extensions
db = SQLAlchemy(app)




# Initialize LoginManager for handling user sessions
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

# Initialize Bootstrap for front-end styling
Bootstrap(app)


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Data Base Models
class User(UserMixin, db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    email = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(256), nullable=False)
    role = db.Column(
        db.String(50), nullable=False
    )  # Role can be 'admin', 'sales', 'owner'
    approved = db.Column(
        db.Boolean, default=False
    )  # User needs to be approved to log in

    def is_admin(self):
        return self.role == "admin"

    def is_owner(self):
        return self.role == "owner"


class PriceData(db.Model):
    __tablename__ = 'price_data'  # Match the actual table name in the database

    id = db.Column(db.Integer, primary_key=True)
    city_name = db.Column(db.String(100), nullable=False)
    commodity = db.Column(db.String(100), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    day = db.Column(db.Integer, nullable=False)  # Day of the year
    price = db.Column(db.Float, nullable=False)
    source = db.Column(db.String(50), nullable=False)
    season = db.Column(db.String(20), nullable=False)


class BreakEvenEstimation(db.Model):
    __tablename__ = 'break_even_estimations'
    
    id = db.Column(db.Integer, primary_key=True)
    
    # Revenue Calculator Fields
    variety = db.Column(db.String(100), nullable=False)
    city = db.Column(db.String(100), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    forecast_date = db.Column(db.Date, nullable=False)
    yield_per_acre = db.Column(db.Float, nullable=False)
    
    # Cost Fields
    cost_per_acre = db.Column(db.Float, nullable=False)
    harvest_cost_per_box = db.Column(db.Float, nullable=False)
    cost_of_box = db.Column(db.Float, nullable=False)
    boxes_bonus_per_yield = db.Column(db.Float, nullable=False)
    
    # Analysis Date Range
    start_date_range = db.Column(db.Date, nullable=False)
    end_date_range = db.Column(db.Date, nullable=False)
    
    # Results
    forecasted_price = db.Column(db.Float, nullable=True)
    revenue_per_acre = db.Column(db.Float, nullable=True)
    revenue_after_costs = db.Column(db.Float, nullable=True)
    revenue_per_box = db.Column(db.Float, nullable=True)
    season = db.Column(db.String(20), nullable=True)
    
    # Timestamps
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def __repr__(self):
        return f'<BreakEvenEstimation {self.id} for {self.variety} in {self.city}>'
    
    def to_dict(self):
        """Convert instance to dictionary for API responses"""
        return {
            'id': self.id,
            'variety': self.variety,
            'city': self.city,
            'start_date': self.start_date.strftime('%Y-%m-%d') if self.start_date else None,
            'forecast_date': self.forecast_date.strftime('%Y-%m-%d') if self.forecast_date else None,
            'yield_per_acre': self.yield_per_acre,
            'cost_per_acre': self.cost_per_acre,
            'harvest_cost_per_box': self.harvest_cost_per_box,
            'cost_of_box': self.cost_of_box,
            'boxes_bonus_per_yield': self.boxes_bonus_per_yield,
            'start_date_range': self.start_date_range.strftime('%Y-%m-%d') if self.start_date_range else None,
            'end_date_range': self.end_date_range.strftime('%Y-%m-%d') if self.end_date_range else None,
            'forecasted_price': self.forecasted_price,
            'revenue_per_acre': self.revenue_per_acre,
            'revenue_after_costs': self.revenue_after_costs,
            'revenue_per_box': self.revenue_per_box,
            'season': self.season,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }

class ShippingPriceData(db.Model):

    id = db.Column(db.Integer, primary_key=True)
    commodity = db.Column(db.String, nullable=False)
    region_name = db.Column(db.String, nullable=False)
    year = db.Column(db.Integer, nullable=False)
    day = db.Column(db.Integer, nullable=False)
    source = db.Column(db.String, nullable=False)
    price = db.Column(db.Float)
    season = db.Column(db.String, nullable=True)  

    


class UShippingPriceData(db.Model):
    __tablename__ = 'usda_shipping_price_data'

    id = db.Column(db.Integer, primary_key=True)
    city_name = db.Column(db.String(100), nullable=False)
    commodity = db.Column(db.String(100), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    price = db.Column(db.Float, nullable=False)
    source = db.Column(db.String(50), nullable=False, default="USDA")
    season = db.Column(db.String(20), nullable=False)


class WeatherForecast(db.Model):
    __tablename__ = "weather_forecast"
    id = db.Column(db.Integer, primary_key=True)
    city_name = db.Column(db.String(100), nullable=False)
    latitude = db.Column(db.Float, nullable=False)
    longitude = db.Column(db.Float, nullable=False)
    forecast_date = db.Column(db.Date, nullable=False)  # Date of forecast
    variable = db.Column(db.String(10), nullable=False)  # PRECIP, TMIN, TMAX, TAVG
    forecasted_value = db.Column(
        db.Float, nullable=False
    )  # Forecasted value (e.g., temp or precipitation)
    ensemble_member = db.Column(db.Integer, nullable=False)  # Ensemble member
    source = db.Column(db.String(50), nullable=False)  # Source of the data, e.g., 'IBM'
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class ClimatologyData(db.Model):
    __tablename__ = "climatology_data"
    id = db.Column(db.Integer, primary_key=True)
    city_name = db.Column(db.String(255), nullable=False)
    latitude = db.Column(db.Float, nullable=False)
    longitude = db.Column(db.Float, nullable=False)
    forecast_date = db.Column(db.Date, nullable=False)  # Forecast date
    variable = db.Column(db.String(10), nullable=False)  # e.g., PRECIP, TAVG
    climatology_value = db.Column(db.Float, nullable=False)
    source = db.Column(db.String(50), nullable=False)  # Data source (e.g., ERA5)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class AlertSetting(db.Model):
    """Model for price alert settings."""
    
    __tablename__ = 'alert_settings'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)  # Add user_id
    city = db.Column(db.String(100), nullable=False)  
    commodity = db.Column(db.String(100), nullable=False)
    threshold = db.Column(db.Float, nullable=False, default=5.0)  # Default 5% threshold
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationship with User
    user = db.relationship('User', backref=db.backref('alert_settings', lazy=True))

    def __repr__(self):
        return f'<AlertSetting {self.id}: {self.commodity} @ {self.threshold}%>'
    
    def to_dict(self):
        """Convert instance to dictionary."""
        return {
            'id': self.id,
            'user_id': self.user_id,
            'city': self.city,
            'commodity': self.commodity,
            'threshold': self.threshold,
            'isActive': self.is_active,
            'createdAt': self.created_at.isoformat(),
            'updatedAt': self.updated_at.isoformat()
        }

class Notification(db.Model):
    """Model for notifications."""
    
    __tablename__ = 'notifications'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)  # Add user_id
    alert_setting_id = db.Column(db.Integer, db.ForeignKey('alert_settings.id'), nullable=True)
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text, nullable=False)
    read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relationships
    user = db.relationship('User', backref=db.backref('notifications', lazy=True))
    alert_setting = db.relationship('AlertSetting', backref=db.backref('notifications', lazy=True))
    
    def __repr__(self):
        return f'<Notification {self.id}: {self.title}>'
    
    def to_dict(self):
        """Convert instance to dictionary."""
        return {
            'id': self.id,
            'user_id': self.user_id,
            'title': self.title,
            'message': self.message,
            'read': self.read,
            'created_at': self.created_at.isoformat(),
            'alert_setting_id': self.alert_setting_id
        }

# USDA DATA IMPORT SETTINGS HERE!
INTERESTED_CITIES = [
    "BALTIMORE",
    "BOSTON",
    "CHICAGO",
    "COLUMBIA",
    "MIAMI",
    "NEW YORK",
    "PHILADELPHIA",
    "LOS ANGELES",
]

INTERESTED_COMMODITIES = [
    "Fresno",
    "Habanero",
    "Hungarian Wax",
    "Jalapeno",
    "Long Hot",
    "Shishito",
    "Anaheim",
    "Cubanelles",
    "Poblano",
    "Serrano",
]

INTERESTED_COMMODITIES_USDA = [
    "Fresno",
    "Habanero",
    "Hungarian Wax",
    "Jalapeno",
    "Long Hot",
    "Shishito",
    "Anaheim",
    "Cubanelle",
    "Poblano",
    "Serrano",
]



@app.route('/api/fetch-data', methods=['GET'])
def fetch_data():
    try:
        start_dates = get_best_start_dates()
        raw_data = fetch_data_from_api(start_dates)
        return jsonify(raw_data)  # Convert raw data to JSON response
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Function to fetch daily USDA data
def get_last_fetched_usda_date():
    # Step 1: Find the maximum year
    max_year = (
        db.session.query(func.max(PriceData.year))
        .filter(PriceData.source.in_(["USDA", "Historical"]))
        .scalar()
    )

    if max_year is not None:
        # Step 2: Find the maximum day within the maximum year
        max_day = (
            db.session.query(func.max(PriceData.day))
            .filter(
                PriceData.year == max_year, PriceData.source.in_(["USDA", "Historical"])
            )
            .scalar()
        )

        # Validate the day
        if max_day is not None:
            try:
                # Ensure that the day is valid for the year
                # For leap years, we can have a day value of 366
                date_check = datetime.strptime(f"{max_year}-{max_day}", "%Y-%j")
                last_fetched_date = pd.Timestamp(date_check)
                logging.info(
                    f"Last fetched USDA date: {last_fetched_date} (year: {max_year}, day: {max_day})"
                )
                return last_fetched_date
            except ValueError:
                logging.error(
                    f"Invalid date found: year={max_year}, day={max_day}. Falling back to default date."
                )
        else:
            logging.warning(
                f"No valid day found for year {max_year}. Using fallback date 2020-01-01."
            )
    else:
        logging.warning(
            "No valid USDA or Historical data found. Using fallback date 2020-01-01."
        )

    # If no valid date was found, return a default fallback date
    last_fetched_date = pd.Timestamp("2018-01-01")
    logging.info(f"Final fetched USDA date: {last_fetched_date}")
    return last_fetched_date


# Function to fetch USDA data for multiple days from the last fetched date
def fetch_usda_daily_data():
    with app.app_context():
        logging.info("USDA data fetch job triggered")
        base_endpoint = "https://marsapi.ams.usda.gov/services/v1.2/marketTypes/1030/cr?dsId=/1/1/&q=report_date="

        # Get the last fetched date
        start_dt = get_last_fetched_usda_date()
        logging.info(f"Last fetched USDA date: {start_dt}")
        end_dt = pd.Timestamp.today()

        # Prepare the authorization header
        api_key = "7XWk8PBs9+O3lWfreM+DtrsaM3OCkzOx"
        encoded_api_key = base64.b64encode(f"{api_key}:".encode()).decode()
        headers = {
            "Authorization": f"Basic {encoded_api_key}",
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json",
        }

        current_dt = start_dt
        while current_dt <= end_dt:
            current_date_formatted = current_dt.strftime("%m/%d/%Y")
            logging.info(f"Fetching USDA data for {current_date_formatted}")

            # Check if data for the current date already exists in the database
            existing_record = (
                db.session.query(PriceData)
                .filter(
                    PriceData.year == current_dt.year,
                    PriceData.day == current_dt.timetuple().tm_yday,
                    PriceData.source == "USDA",
                )
                .first()
            )

            if existing_record:
                logging.info(
                    f"Data for {current_date_formatted} already exists. Skipping."
                )
            else:
                endpoint = base_endpoint + current_date_formatted
                response = requests.get(endpoint, headers=headers)

                logging.info(f"API Response Status Code: {response.status_code}")

                if response.status_code == 200:
                    json_data = response.json()
                    if json_data.get("results") and isinstance(
                        json_data["results"], list
                    ):
                        logging.info(f"Valid data found for {current_date_formatted}")
                        process_usda_data(json_data["results"])
                    else:
                        logging.warning(
                            f"No valid results found in the API response for {current_date_formatted}"
                        )
                else:
                    logging.error(
                        f"Error fetching USDA data for {current_date_formatted}: {response.status_code} - {response.text}"
                    )
            current_dt += timedelta(days=1)
            json_data = None
            gc.collect()




# Process and store USDA data in the database, filtered by interested commodities and cities
def process_usda_data(data):
    for report in data:
        commodity_name = report.get("commodity", "")
        # Remove 'Peppers, ' prefix before comparing
        if commodity_name.startswith("Peppers, "):
            commodity_name = commodity_name.replace("Peppers, ", "").strip()

        # Filter out only the commodities of interest
        if commodity_name in INTERESTED_COMMODITIES_USDA:
            report_date = report.get("report_date", "")
            year = (
                datetime.strptime(report_date, "%m/%d/%Y").year if report_date else ""
            )
            day_of_year = (
                datetime.strptime(report_date, "%m/%d/%Y").timetuple().tm_yday
                if report_date
                else ""
            )

            # Check if price is missing, skip the entry if it is
            price = report.get("low_price")
            if price is None:
                logging.warning(
                    f"Skipping entry due to missing price for {commodity_name} on {report_date} in {report.get('location', '')}"
                )
                continue  # Skip this entry if price is missing

            # Extract the city name before the comma, convert to uppercase
            city_name_full = report.get("location", "")
            city_name = city_name_full.split(",")[
                0
            ].upper()  # Get part before comma and convert to uppercase

            # Check if the extracted city is in the list of interested cities
            if city_name not in INTERESTED_CITIES:
                logging.warning(
                    f"Skipping entry for city {city_name_full} as it is not in the list of interested cities"
                )
                continue  # Skip cities that are not in the interested list

            # Prepare the price_data entry
            price_data = PriceData(
                city_name=city_name,  # Now contains only the city name in uppercase
                commodity=commodity_name,  # Now contains only the commodity name without prefix
                year=year,
                day=day_of_year,
                price=price,  # Only insert if the price is valid
                source="USDA",
                season=determine_season(report_date),
            )

            db.session.add(price_data)

    db.session.commit()
    logging.info(f"USDA data saved for valid entries")


# Determine season based on month
def determine_season(report_date):
    try:
        # Handle date format like '10/02/2024' (USDA reports)
        month = datetime.strptime(report_date, "%m/%d/%Y").month
    except ValueError:
        # Handle format like '2024-10-02' if other data sources use this
        month = datetime.strptime(report_date, "%Y-%m-%d").month

    if month in [3, 4, 5]:
        return "Spring"
    elif month in [6, 7, 8]:
        return "Summer"
    elif month in [9, 10, 11]:
        return "Autumn"
    else:
        return "Winter"


def determine_season_for_dashboard(forecast_date):
    # forecast_date is expected to be in the format '%Y-%m-%d'
    month = forecast_date.month

    if month in [3, 4, 5]:
        return "Spring"
    elif month in [6, 7, 8]:
        return "Summer"
    elif month in [9, 10, 11]:
        return "Autumn"
    else:
        return "Winter"


def determine_season_from_year_day(year, day_of_year):
    """
    This function converts a given year and day-of-year into a valid date and determines the season.
    """
    try:
        # Convert the year and day of the year to a valid date
        date_from_day = datetime.strptime(f"{year}-{day_of_year}", "%Y-%j")
        month = date_from_day.month

        # Determine the season based on the month
        if month in [3, 4, 5]:
            return "Spring"
        elif month in [6, 7, 8]:
            return "Summer"
        elif month in [9, 10, 11]:
            return "Autumn"
        else:
            return "Winter"

    except ValueError:
        # If there is an error, return a default season or handle accordingly
        return "Unknown"


# PRODUCE IQ DATA FETCHING IS HERE!!
# This is for pricedata
# Function to fetch the last fetched date
def get_last_fetched_date():
    # Step 1: Find the maximum year
    max_year = (
        db.session.query(func.max(PriceData.year))
        .filter(PriceData.source.in_(["ProduceIQ", "Historical"]))
        .scalar()
    )

    if max_year is not None:
        # Step 2: Find the maximum day within the maximum year
        max_day = (
            db.session.query(func.max(PriceData.day))
            .filter(
                PriceData.year == max_year,
                PriceData.source.in_(["ProduceIQ", "Historical"]),
            )
            .scalar()
        )

        # Validate the day
        if max_day is not None:
            try:
                # Ensure that the day is valid for the year
                # For leap years, we can have a day value of 366
                date_check = datetime.strptime(f"{max_year}-{max_day}", "%Y-%j")
                last_fetched_date = pd.Timestamp(date_check)
                logging.info(
                    f"Last fetched ProduceIQ date: {last_fetched_date} (year: {max_year}, day: {max_day})"
                )
                return last_fetched_date
            except ValueError:
                logging.error(
                    f"Invalid date found: year={max_year}, day={max_day}. Falling back to default date."
                )
        else:
            logging.warning(
                f"No valid day found for year {max_year}. Using fallback date 2020-01-01."
            )
    else:
        logging.warning(
            "No valid ProduceIQ or Historical data found. Using fallback date 2020-01-01."
        )

    # If no valid date was found, return a default fallback date
    last_fetched_date = pd.Timestamp("2018-01-01")
    logging.info(f"Final fetched ProduceIQ date: {last_fetched_date}")
    return last_fetched_date


# Fetching data from Produce IQ
def fetch_daily_data():
    with app.app_context():
        base_url = "https://api.produceiq.com/index/v2/trends/"
        headers = {"Api-Subscription-Key": "5aa11f87fed04300b05addd031c56ffa"}

        # Define the 10 specific commodities to filter by
        wanted_commodities = [
            "Anaheim",
            "Cubanelles",
            "Fresno",
            "Habanero",
            "Hungarian Wax",
            "Jalapeno",
            "Long Hot",
            "Poblano",
            "Serrano",
            "Shishito",
        ]

        # Standardize the list to lowercase for comparison
        wanted_commodities = [commodity.lower() for commodity in wanted_commodities]

        # Mapping standardized names back to desired format
        standardized_name = {
            "anaheim": "Anaheim",
            "cubanelles": "Cubanelles",
            "fresno": "Fresno",
            "habanero": "Habanero",
            "hungarian wax": "Hungarian Wax",
            "jalapeno": "Jalapeno",
            "long hot": "Long Hot",
            "poblano": "Poblano",
            "serrano": "Serrano",
            "shishito": "Shishito",
        }

        # Get the last fetched date
        start_dt = get_last_fetched_date()
        end_dt = pd.Timestamp.today()  # Fetch data up to today

        # Loop through each day one by one from the last fetched date to today
        current_dt = start_dt
        while current_dt <= end_dt:
            # Check if data for the current date already exists in the database
            existing_data = (
                db.session.query(PriceData)
                .filter(
                    PriceData.year == current_dt.year,
                    PriceData.day == current_dt.day_of_year,
                    PriceData.source == "ProduceIQ",
                )
                .first()
            )

            if existing_data:
                logging.info(
                    f"Data for {current_dt.strftime('%Y-%m-%d')} already exists. Skipping."
                )
                current_dt += pd.Timedelta(days=1)
                continue

            params = {
                "commodityId": 18,  # Adjust this for different commodities
                "from": current_dt.strftime("%Y-%m-%d"),
                "to": current_dt.strftime(
                    "%Y-%m-%d"
                ),  # Fetch data for one day at a time
            }

            # Fetch data for the current day
            response = requests.get(
                f"{base_url}terminal-market-trends",
                headers=headers,
                params=params,
                verify=False,
            )

            # Log the full response for debugging
            logging.info(f"API Response for {current_dt.strftime('%Y-%m-%d')}")

            if response.status_code == 200:
                data = response.json().get("subset", [])
                logging.info(f"Fetched data for {current_dt.strftime('%Y-%m-%d')}")
            else:
                logging.error(
                    f"Failed to fetch data for {current_dt.strftime('%Y-%m-%d')}. Status code: {response.status_code}"
                )
                logging.error(
                    f"API Error Response: {response.text}"
                )  # Log the error response
                # Move to the next day even if this request fails
                current_dt += pd.Timedelta(days=1)
                continue

            # Process and save the data
            if not data:
                logging.error(
                    f"No data found in the response for {current_dt.strftime('%Y-%m-%d')}."
                )
                # Move to the next day even if no data is found
                current_dt += pd.Timedelta(days=1)
                continue

            for item in data:
                # Standardize variety name from the API
                variety_name = item.get("varietyName", "").strip().lower()

                # Compare in standardized format
                if variety_name in wanted_commodities:
                    # Map back to the desired format
                    variety_name = standardized_name.get(variety_name, variety_name)

                    city_name = item.get("terminalMarketCityName", "").strip().lower().title()
                    year = item.get("isoYear")
                    day_of_year = item.get("day")
                    price = item.get("price")
                    source = "ProduceIQ"


                    # Calculate the season based on the month
                    month = (
                        pd.Timestamp(year=year, day=1, month=1).day_of_year // 30 + 1
                    )
                    if month in [3, 4, 5]:
                        season = "Spring"
                    elif month in [6, 7, 8]:
                        season = "Summer"
                    elif month in [9, 10, 11]:
                        season = "Autumn"
                    else:
                        season = "Winter"

                    # Save to the database
                    price_data = PriceData(
                        city_name=city_name,
                        commodity=variety_name,
                        year=year,
                        day=day_of_year,
                        price=price,
                        source=source,
                        season=season,
                    )

                    db.session.add(price_data)
                    logging.info(
                        f"Added {variety_name} for {current_dt.strftime('%Y-%m-%d')} in {city_name}"
                    )

            # Commit the data for the current day
            db.session.commit()
            gc.collect()
            logging.info(
                f'Data for {current_dt.strftime("%Y-%m-%d")} saved to the database.'
            )

            # Move to the next day
            current_dt += pd.Timedelta(days=1)
            data = None  # Release JSON data memory
            gc.collect()  # Explicit garbage collection

        logging.info(
            f"Data fetching completed from {start_dt.strftime('%Y-%m-%d')} to {end_dt.strftime('%Y-%m-%d')}."
        )





import threading

def fetch_shipping_data():
    with app.app_context():  # Ensure Flask app context is active
        try:
            fetch_shipping_point_data()  # Long-running task
            print("Data fetch completed successfully.")
        except Exception as e:
            print(f"Error: {str(e)}")

@app.route('/fetch-shipping', methods=['GET'])
def fetch_shipping():
    thread = threading.Thread(target=fetch_shipping_data)
    thread.start()
    return jsonify({"status": "success", "message": "Task started in background"}), 202



@app.route("/api/test_ibm_live", methods=["GET"])
def test_ibm_live():
    # Example parameters – adjust as needed
    lat = 26.4187
    lon = -81.4173
    # Build a sample query JSON payload for a variable, e.g., TAVG
    iso_8601 = "%Y-%m-%dT%H:%M:%SZ"
    query_json = {
        "layers": [
            {
                "type": "raster",
                "id": 50685,  # ID for TAVG
                "temporal": {
                    "intervals": [
                        {
                            "start": (datetime.utcnow() - timedelta(minutes=1)).strftime(iso_8601),
                            "end": (datetime.utcnow() + timedelta(minutes=1)).strftime(iso_8601)
                        }
                    ]
                },
                "dimensions": [
                    {"name": "forecast", "value": "01"},
                    {"name": "horizon", "value": 0}
                ],
            }
        ],
        "spatial": {"type": "point", "coordinates": [lon, lat]},
        "temporal": {"intervals": [{"snapshot": "1982-01-01T00:00:00Z"}]},
        "outputType": "json"
    }

    try:
        # This sends the query live to IBM’s PAIRS API
        # # df = query.submit(query_json).point_data_as_dataframe()
        df = query.submit(query_json, client=eis_client).point_data_as_dataframe()

        # Return the data directly as JSON
        return jsonify(df.to_dict(orient="records"))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import ibmpairs.query as query
import logging


# def fetch_and_store_weather_forecast(
#     start_forecast_date,
#     end_forecast_horizon_date,
#     start_forecast_horizon_date,
#     start_ensembles,
#     end_ensembles,
#     store_in_excel,
# ):

#     # IBM API Configuration
#     eis_client = get_client(
#         api_key="PHXfXSpwiWwQ9NoUg9IYhhaf24cXmmMwZa0zPoW23hktX8",
#         tenant_id="7370463f-df01-4aa0-b204-d6d15ff71f85",
#         org_id=EIS_ORG_ID,
#         legacy=False,
#     )

#     # Forecast parameters
#     layers_TWC = {"PRECIP": 50686, "TAVG": 50685}
#     iso_8601 = "%Y-%m-%dT%H:%M:%SZ"

#     # List of cities with their latitude and longitude
#     cities = {
#         "Immokalee Fl": {"lat": "26.4187", "lon": "-81.4173"},
#         "Palm Beach County, fl": {"lat": "26.7153", "lon": "-80.0534"},
#         "Vineland NJ": {"lat": "39.4802", "lon": "-75.0138"},
#         "Sodus, Michigan": {"lat": "42.0086", "lon": "-86.3614"},
#         "Sinaloa": {"lat": "25.1721", "lon": "-107.4795"},
#         "Sonora": {"lat": "29.2972", "lon": "-110.3309"},
#         "Ensenada": {"lat": "31.86613056", "lon": "-116.59971944"},
#         "Baja Mx": {"lat": "28.0444", "lon": "-115.2062"},
#         "Culican": {"lat": "24.8091", "lon": "-107.3940"},
#         "Hendersonville, NC": {"lat": "35.3187", "lon": "-82.4610"},
#         "Cameron SC": {"lat": "33.5568", "lon": "-80.7151"},
#         "Adel, Ga": {"lat": "31.13633333", "lon": "-83.42216389"},
#         "Lake Park Ga": {"lat": "30.6844", "lon": "-83.1849"},
#         "Bowling Green Fl": {"lat": "27.6386", "lon": "-81.8265"},
#     }

#     # Ensure 'start_forecast_horizon_date' and 'end_forecast_horizon_date' are datetime objects
#     if isinstance(start_forecast_horizon_date, str):
#         start_forecast_horizon_date = datetime.strptime(
#             start_forecast_horizon_date, "%Y-%m-%d"
#         )
#     if isinstance(end_forecast_horizon_date, str):
#         end_forecast_horizon_date = datetime.strptime(
#             end_forecast_horizon_date, "%Y-%m-%d"
#         )

#     # Ensure 'start_forecast_date' is a datetime object
#     if isinstance(start_forecast_date, str):
#         start_forecast_date = datetime.strptime(start_forecast_date, "%Y-%m-%d")

#     # Generate valid dates and horizons
#     valid_dates_horizons = []
#     count = 0
#     date = start_forecast_horizon_date
#     while date <= end_forecast_horizon_date:
#         valid_date = date
#         horizon = (valid_date - start_forecast_date).days
#         valid_dates_horizons.append((valid_date, horizon))
#         count += 1
#         date += timedelta(days=1)
#     logging.info(f"Generated {len(valid_dates_horizons)} forecast horizons.")

#     # Ensemble members in smaller batches
#     ensemble_members_batches = [
#         [str(x).zfill(2) for x in range(i, min(i + 1, end_ensembles + 1))]
#         for i in range(start_ensembles, end_ensembles + 1)
#     ]

#     # Loop over each city, variable, and ensemble members in batches
#     for city, coordinates in cities.items():
#         lat = float(coordinates["lat"])
#         lon = float(coordinates["lon"])
#         logging.info(f"Starting data query for city: {city} (lat: {lat}, lon: {lon})")
#         temperature_adjustment = fetch_elevation_data(lat, lon)

#         for VARIABLE in layers_TWC.keys():
#             logging.info(f"Starting data query for variable: {VARIABLE}")
#             for ensemble_members in ensemble_members_batches:
#                 logging.info(f"Querying ensemble members: {ensemble_members}")

#                 query_layers = []
#                 for valid_date, horizon in valid_dates_horizons:
#                     for ens in ensemble_members:
#                         query_layers.append(
#                             {
#                                 "type": "raster",
#                                 "id": layers_TWC[VARIABLE],
#                                 "temporal": {
#                                     "intervals": [
#                                         {
#                                             "start": (
#                                                 valid_date - timedelta(seconds=60)
#                                             ).strftime(iso_8601),
#                                             "end": (
#                                                 valid_date + timedelta(seconds=60)
#                                             ).strftime(iso_8601),
#                                         }
#                                     ]
#                                 },
#                                 "dimensions": [
#                                     {"name": "forecast", "value": ens},
#                                     {"name": "horizon", "value": horizon},
#                                 ],
#                             }
#                         )

#                 # Construct query JSON payload
#                 query_json = {
#                     "layers": query_layers,
#                     "spatial": {"type": "point", "coordinates": [lat, lon]},
#                     "temporal": {"intervals": [{"snapshot": "1982-01-01"}]},
#                     "outputType": "json",
#                 }
#                 logging.info(
#                     f"Constructed query JSON for ensemble members {ensemble_members}."
#                 )

#                 # Submit the query and process the results
#                 try:
#                     logging.info("Submitting query to IBM PAIRS API.")
#                     # df = query.submit(query_json).point_data_as_dataframe()
#                     df = query.submit(query_json, client=eis_client).point_data_as_dataframe()


#                     if df.empty:
#                         logging.warning(
#                             f"No data returned for ensemble members {ensemble_members}"
#                         )
#                         continue

#                     logging.info(f"Data retrieved, processing {len(df)} records.")
#                     if VARIABLE == "TAVG":
#                         df["value"] = df["value"].astype(float) + temperature_adjustment
#                     if store_in_excel == 1:
#                         process_and_store_in_excel(df, city, lat, lon, VARIABLE)
#                     else:
#                         process_and_store_data(df, city, lat, lon, VARIABLE)
#                     gc.collect()
#                     df = None
#                     gc.collect()
#                 except Exception as e:
#                     logging.error(f"Error during query submission for {VARIABLE}: {e}")
#                     continue

#     logging.info("Weather forecast data query completed successfully.")
#     return "Weather forecast data query completed."



#  Weather forecast data fetching insertion and pruning
def fetch_and_store_weather_forecast(
    start_forecast_date,
    end_forecast_horizon_date,
    start_forecast_horizon_date,
    start_ensembles,
    end_ensembles,
    store_in_excel,
):
    # IBM API Configuration
    eis_client = get_client(
        api_key="PHXfXSpwiWwQ9NoUg9IYhhaf24cXmmMwZa0zPoW23hktX8",
        tenant_id="7370463f-df01-4aa0-b204-d6d15ff71f85",
        org_id=EIS_ORG_ID,
        legacy=False,
    )

    # Forecast parameters
    layers_TWC = {"PRECIP": 50686, "TAVG": 50685}
    iso_8601 = "%Y-%m-%dT%H:%M:%SZ"

    # List of cities with their latitude and longitude
    cities = {
        "Immokalee Fl": {"lat": "26.4187", "lon": "-81.4173"},
        "Palm Beach County, fl": {"lat": "26.7153", "lon": "-80.0534"},
        "Vineland NJ": {"lat": "39.4802", "lon": "-75.0138"},
        "Sodus, Michigan": {"lat": "42.0086", "lon": "-86.3614"},
        "Sinaloa": {"lat": "25.1721", "lon": "-107.4795"},
        "Sonora": {"lat": "29.2972", "lon": "-110.3309"},
        "Ensenada": {"lat": "31.86613056", "lon": "-116.59971944"},
        "Baja Mx": {"lat": "28.0444", "lon": "-115.2062"},
        "Culican": {"lat": "24.8091", "lon": "-107.3940"},
        "Hendersonville, NC": {"lat": "35.3187", "lon": "-82.4610"},
        "Cameron SC": {"lat": "33.5568", "lon": "-80.7151"},
        "Adel, Ga": {"lat": "31.13633333", "lon": "-83.42216389"},
        "Lake Park Ga": {"lat": "30.6844", "lon": "-83.1849"},
        "Bowling Green Fl": {"lat": "27.6386", "lon": "-81.8265"},
    }

    # Ensure date arguments are datetime objects
    if isinstance(start_forecast_date, str):
        start_forecast_date = datetime.strptime(start_forecast_date, "%Y-%m-%d")
    if isinstance(start_forecast_horizon_date, str):
        start_forecast_horizon_date = datetime.strptime(
            start_forecast_horizon_date, "%Y-%m-%d"
        )
    if isinstance(end_forecast_horizon_date, str):
        end_forecast_horizon_date = datetime.strptime(
            end_forecast_horizon_date, "%Y-%m-%d"
        )

    # Generate valid dates/horizons
    valid_dates_horizons = []
    date_ptr = start_forecast_horizon_date
    while date_ptr <= end_forecast_horizon_date:
        horizon = (date_ptr - start_forecast_date).days
        valid_dates_horizons.append((date_ptr, horizon))
        date_ptr += timedelta(days=1)

    logging.info(f"Generated {len(valid_dates_horizons)} forecast horizons.")

    # Build smaller batches of ensemble members
    ensemble_members_batches = [
        [str(x).zfill(2) for x in range(i, min(i + 1, end_ensembles + 1))]
        for i in range(start_ensembles, end_ensembles + 1)
    ]

    # Loop over each city, variable, and ensemble members in batches
    for city, coordinates in cities.items():
        lat = float(coordinates["lat"])
        lon = float(coordinates["lon"])
        logging.info(f"Starting data query for city: {city} (lat: {lat}, lon: {lon})")

        # Elevation-based temperature adjustment (if TAVG)
        temperature_adjustment = fetch_elevation_data(lat, lon)

        for VARIABLE in layers_TWC.keys():
            logging.info(f"Starting data query for variable: {VARIABLE}")

            for ensemble_members in ensemble_members_batches:
                logging.info(f"Querying ensemble members: {ensemble_members}")

                # Build up the query JSON
                query_layers = []
                for valid_date, horizon in valid_dates_horizons:
                    query_layers.append(
                        {
                            "type": "raster",
                            "id": layers_TWC[VARIABLE],
                            "temporal": {
                                "intervals": [
                                    {
                                        "start": (
                                            valid_date - timedelta(seconds=60)
                                        ).strftime(iso_8601),
                                        "end": (
                                            valid_date + timedelta(seconds=60)
                                        ).strftime(iso_8601),
                                    }
                                ]
                            },
                            "dimensions": [
                                {"name": "forecast", "value": ens}
                                for ens in ensemble_members
                            ]
                            + [{"name": "horizon", "value": horizon}],
                        }
                    )

                query_json = {
                    "layers": query_layers,
                    "spatial": {"type": "point", "coordinates": [lat, lon]},
                    "temporal": {"intervals": [{"snapshot": "1982-01-01"}]},
                    "outputType": "json",
                }

                try:
                    logging.info("Submitting query to IBM PAIRS API.")
                    df = query.submit(query_json, client=eis_client).point_data_as_dataframe()

                    if df.empty:
                        logging.warning(
                            f"No data returned for ensemble {ensemble_members} / {VARIABLE}"
                        )
                        continue

                    logging.info(f"Data retrieved, processing {len(df)} records.")

                    # If TAVG, apply temperature adjustment
                    if VARIABLE == "TAVG":
                        df["value"] = df["value"].astype(float) + temperature_adjustment

                    # ------------------------------------------------------------------
                    # (A) If user wants to store in Excel instead of DB:
                    # ------------------------------------------------------------------
                    if store_in_excel == 1:
                        process_and_store_in_excel(df, city, lat, lon, VARIABLE)
                        # No DB insertion, so no pruning needed in that scenario.
                    else:
                        # ------------------------------------------------------------------
                        # (B) Otherwise, insert into DB and prune old rows 1:1
                        # ------------------------------------------------------------------
                        new_rows = []
                        for _, row in df.iterrows():
                            try:
                                date_val = int(row["timestamp"])
                                value_val = float(row["value"])
                                # property might look like "forecast:01; horizon:3"
                                # so parse out ensemble:
                                ens_str = row.get("property", "forecast:0").split(";")[0]
                                ens_member = int(ens_str.split(":")[1])

                                forecast_date = datetime.fromtimestamp(
                                    date_val / 1000, tz=pytz.utc
                                ).date()

                                wf = WeatherForecast(
                                    city_name=city,
                                    latitude=lat,
                                    longitude=lon,
                                    forecast_date=forecast_date,
                                    variable=VARIABLE,
                                    forecasted_value=value_val,
                                    ensemble_member=ens_member,
                                    source="IBM",
                                )
                                new_rows.append(wf)
                            except Exception as row_e:
                                logging.error(f"Row parse error: {row_e}")
                                continue

                        # Insert all new rows
                        db.session.add_all(new_rows)
                        db.session.flush()  # so they get IDs

                        n_inserted = len(new_rows)
                        if n_inserted > 0:
                            # Prune that many oldest rows
                            subq = (
                                db.session.query(WeatherForecast.id)
                                .order_by(WeatherForecast.id.asc())
                                .limit(n_inserted)
                                .subquery()
                            )
                            db.session.query(WeatherForecast).filter(
                                WeatherForecast.id.in_(subq)
                            ).delete(synchronize_session=False)

                        # Commit once after insert + prune
                        db.session.commit()

                    # Clean up memory usage
                    gc.collect()
                    df = None
                    gc.collect()

                except Exception as e:
                    logging.error(f"Error during query submission for {VARIABLE}: {e}")
                    continue

    logging.info("Weather forecast data query completed successfully.")
    return "Weather forecast data query completed."



# @app.route("/api/test_fake_weather_forecast")
# def run_fake_weather_test():
#     test_fake_weather_forecast_insertion_and_prune()
#     return "Fake weather forecast test complete!"






# def test_fake_weather_forecast_insertion_and_prune():
#     """
#     Inserts a few fake WeatherForecast rows and prunes the same number of oldest rows.
#     """

#     # 1) Create a fake DataFrame with columns we need.
#     #    We'll store 'timestamp' in milliseconds since epoch
#     #    and 'value' as the forecasted_value.
#     #    Optionally you can add a column for 'ensemble_member' if you want multiple members.
#     df = pd.DataFrame({
#         "timestamp": [1680000000000, 1680000001000, 1680000002000],
#         "value": [10.5, 12.2, 15.0],
#         "ensemble_member": [1, 1, 1]   # or vary them if you want multiple ensemble members
#     })

#     # 2) Decide on a city_name, lat, lon, variable, and source
#     city_name = "Fake City"
#     lat = 26.4187
#     lon = -81.4173
#     variable = "PRECIP"      # or "TAVG", etc.
#     source = "IBM"           # or "TestSource"

#     # 3) Build WeatherForecast rows from the DataFrame
#     new_rows = []
#     for _, row in df.iterrows():
#         try:
#             timestamp_val = int(row["timestamp"])
#             forecast_val = float(row["value"])
#             ensemble = int(row["ensemble_member"])

#             # Convert timestamp to a Python date
#             forecast_date = datetime.utcfromtimestamp(timestamp_val / 1000).date()

#             # Build the WeatherForecast object
#             wdata = WeatherForecast(
#                 city_name=city_name,
#                 latitude=lat,
#                 longitude=lon,
#                 forecast_date=forecast_date,
#                 variable=variable,
#                 forecasted_value=forecast_val,
#                 ensemble_member=ensemble,
#                 source=source,
#             )
#             new_rows.append(wdata)
#         except Exception as row_e:
#             print(f"Error parsing row for {variable}: {row_e}")
#             continue

#     # 4) Insert all new rows
#     db.session.add_all(new_rows)
#     db.session.flush()  # so they get assigned IDs

#     # 5) Prune the same number of oldest rows by ascending ID
#     n_inserted = len(new_rows)
#     if n_inserted > 0:
#         subq = (
#             db.session.query(WeatherForecast.id)
#             .order_by(WeatherForecast.id.asc())
#             .limit(n_inserted)
#             .subquery()
#         )
#         db.session.query(WeatherForecast).filter(
#             WeatherForecast.id.in_(subq)
#         ).delete(synchronize_session=False)

#     # 6) Commit
#     db.session.commit()
#     gc.collect()

#     print(f"Inserted {n_inserted} fake WeatherForecast rows and pruned {n_inserted} oldest rows.")


def process_and_store_data(df, city, lat, lon, VARIABLE):
    # Process and store data in the database
    for _, row in df.iterrows():
        try:
            date = int(row["timestamp"])
            value = float(row["value"])
            ens = int(row.get("property", "forecast:0").split(";")[0].split(":")[1])
            logging.debug(
                f"Processing record - Date: {date}, Value: {value}, Forecast: {ens}"
            )

            # Convert timestamp to datetime
            forecast_date = datetime.fromtimestamp(date / 1000, tz=pytz.utc).date()

            # Store data in the WeatherForecast database
            weather_forecast = WeatherForecast(
                city_name=city,
                latitude=lat,
                longitude=lon,
                forecast_date=forecast_date,
                variable=VARIABLE,
                forecasted_value=value,
                ensemble_member=ens,
                source="IBM",
            )
            db.session.add(weather_forecast)
        except Exception as e:
            logging.error(f"Error processing record for {VARIABLE}: {e}")
            logging.error(f"Record data: {row}")
            continue

    db.session.commit()
    logging.info(f"Data points for city {city} stored successfully.")


def process_and_store_in_excel(df, city, lat, lon, VARIABLE):
    # Process and store data in an Excel file
    import os
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Create directory if it doesn't exist
    data_dir = "data"
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)

    # Define Excel file path for each variable
    file_path = os.path.join(
        data_dir, f"{city.replace(',', '').replace(' ', '_')}_{VARIABLE}.xlsx"
    )

    # Load or create workbook
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
    else:
        workbook = Workbook()
        if workbook.active:
            workbook.remove(workbook.active)

    # Create or get sheet for the data
    sheet_name = f"{VARIABLE}_data"
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=sheet_name)
        # Add headers
        sheet.append(
            [
                "city_name",
                "latitude",
                "longitude",
                "forecast_date",
                "variable",
                "forecasted_value",
                "ensemble_member",
                "source",
            ]
        )
    else:
        sheet = workbook[sheet_name]

    # Append data to the sheet
    for _, row in df.iterrows():
        try:
            date = int(row["timestamp"])
            value = float(row["value"])
            ens = int(row.get("property", "forecast:0").split(";")[0].split(":")[1])

            # Convert timestamp to datetime
            forecast_date = datetime.fromtimestamp(date / 1000, tz=pytz.utc).date()

            # Append data in the format of the database
            sheet.append([city, lat, lon, forecast_date, VARIABLE, value, ens, "IBM"])
        except Exception as e:
            logging.error(f"Error processing record for {VARIABLE}: {e}")
            logging.error(f"Record data: {row}")
            continue

    # Save workbook periodically to minimize data loss risk
    workbook.save(file_path)
    logging.info(f"Data points for city {city} stored in Excel file {file_path}.")


def fetch_elevation_data(lat, lon):
    layers_ELEVATION = {"twc_elevation": 51219, "srtm_elevation": 49506}
    elevation = {}

    for VARIABLE in ["twc_elevation", "srtm_elevation"]:
        query_json = {
            "layers": [{"type": "raster", "id": layers_ELEVATION[VARIABLE]}],
            "spatial": {"type": "point", "coordinates": [lat, lon]},
            "temporal": {"intervals": [{"snapshot": "2020-01-01T00:00:00Z"}]},
        }
        try:
            logging.info(f"Submitting elevation data query for {VARIABLE}.")
            # df = query.submit(query_json).point_data_as_dataframe()
            df = query.submit(query_json, client=eis_client).point_data_as_dataframe()

            if len(df) > 0:
                elevation[VARIABLE] = float(df.iloc[0]["value"])
        except Exception as e:
            logging.error(f"Error retrieving elevation data for {VARIABLE}: {e}")

    if "twc_elevation" in elevation and "srtm_elevation" in elevation:
        # Compute lapse-rate correction of Temperature
        elevation_diff = elevation["srtm_elevation"] - elevation["twc_elevation"]
        temperature_adjustment = elevation_diff * (-0.0098)
        logging.info(
            f"Temperature adjustment based on elevation difference: {temperature_adjustment}"
        )
    else:
        temperature_adjustment = 0
        logging.warning(
            f"Could not calculate temperature adjustment due to missing elevation data."
        )

    return temperature_adjustment






# for climatology Data
# def fetch_and_store_climatology_data(start_climo_date, end_climo_date):
#     # IBM API Configuration
#     eis_client = get_client(
#         api_key=EIS_API_KEY, tenant_id=EIS_TENANT_ID, org_id=EIS_ORG_ID, legacy=False
#     )

#     # Climatology parameters
#     layers_ERA5 = {"PRECIP": 51198, "TAVG": 51199}
#     iso_8601 = "%Y-%m-%dT%H:%M:%SZ"

#     # List of cities with their latitude and longitude
#     cities = {
#         "Ensenada": {"lat": "31.86613056", "lon": "-116.59971944"},
#         "Baja Mx": {"lat": "28.0444", "lon": "-115.2062"},
#         "Culican": {"lat": "24.8091", "lon": "-107.3940"},
#         "Hendersonville, NC": {"lat": "35.3187", "lon": "-82.4610"},
#         "Cameron SC": {"lat": "33.5568", "lon": "-80.7151"},
#         "Adel, Ga": {"lat": "31.13633333", "lon": "-83.42216389"},
#         "Lake Park Ga": {"lat": "30.6844", "lon": "-83.1849"},
#         "Bowling Green Fl": {"lat": "27.6386", "lon": "-81.8265"},
#     }

#     # Loop over each city and variable
#     for city, coordinates in cities.items():
#         lat = float(coordinates["lat"])
#         lon = float(coordinates["lon"])
#         logging.info(
#             f"Starting climatology data query for city: {city} (lat: {lat}, lon: {lon})"
#         )

#         # Fetch elevation data for temperature adjustment
#         temperature_adjustment = fetch_elevation_data(lat, lon)

#         for VARIABLE in layers_ERA5.keys():
#             logging.info(f"Starting climatology data query for variable: {VARIABLE}")

#             # Construct query JSON payload
#             query_json = {
#                 "layers": [{"type": "raster", "id": layers_ERA5[VARIABLE]}],
#                 "spatial": {"type": "point", "coordinates": [lat, lon]},
#                 "temporal": {
#                     "intervals": [
#                         {
#                             "start": start_climo_date.strftime(iso_8601),
#                             "end": end_climo_date.strftime(iso_8601),
#                         }
#                     ]
#                 },
#                 "outputType": "json",
#             }
#             logging.info(f"Constructed query JSON for climatology data.")

#             # Submit the query and process the results
#             try:
#                 logging.info("Submitting query to IBM PAIRS API for climatology data.")
#                 # df = query.submit(query_json).point_data_as_dataframe()
#                 df = query.submit(query_json, client=eis_client).point_data_as_dataframe()

#                 if df.empty:
#                     logging.warning(
#                         f"No data returned for climatology data for {VARIABLE}"
#                     )
#                     continue

#                 # Apply temperature adjustment if the variable is TAVG
#                 if VARIABLE == "TAVG":
#                     df["value"] = df["value"].astype(float) + temperature_adjustment

#                 logging.info(
#                     f"Climatology data retrieved, processing {len(df)} records."
#                 )
#                 process_and_store_climatology_data(df, city, lat, lon, VARIABLE)
#                 gc.collect()
#                 df = None
#                 gc.collect()
#             except Exception as e:
#                 logging.error(
#                     f"Error during climatology data query for {VARIABLE}: {e}"
#                 )
#                 continue

#     logging.info("Climatology data query completed successfully.")
#     return "Climatology data query completed."



# Insertion and pruning

def fetch_and_store_climatology_data(start_climo_date, end_climo_date):
    # IBM API Configuration
    eis_client = get_client(
        api_key=EIS_API_KEY, tenant_id=EIS_TENANT_ID, org_id=EIS_ORG_ID, legacy=False
    )

    # Climatology parameters
    layers_ERA5 = {"PRECIP": 51198, "TAVG": 51199}
    iso_8601 = "%Y-%m-%dT%H:%M:%SZ"

    # List of cities with their latitude and longitude
    cities = {
        "Ensenada": {"lat": "31.86613056", "lon": "-116.59971944"},
        "Baja Mx": {"lat": "28.0444", "lon": "-115.2062"},
        "Culican": {"lat": "24.8091", "lon": "-107.3940"},
        "Hendersonville, NC": {"lat": "35.3187", "lon": "-82.4610"},
        "Cameron SC": {"lat": "33.5568", "lon": "-80.7151"},
        "Adel, Ga": {"lat": "31.13633333", "lon": "-83.42216389"},
        "Lake Park Ga": {"lat": "30.6844", "lon": "-83.1849"},
        "Bowling Green Fl": {"lat": "27.6386", "lon": "-81.8265"},
    }

    # Loop over each city and variable
    for city, coordinates in cities.items():
        lat = float(coordinates["lat"])
        lon = float(coordinates["lon"])
        logging.info(
            f"Starting climatology data query for city: {city} (lat: {lat}, lon: {lon})"
        )

        # Fetch elevation data for temperature adjustment (if TAVG)
        temperature_adjustment = fetch_elevation_data(lat, lon)

        for VARIABLE in layers_ERA5.keys():
            logging.info(f"Starting climatology data query for variable: {VARIABLE}")

            # Build the query JSON payload
            query_json = {
                "layers": [{"type": "raster", "id": layers_ERA5[VARIABLE]}],
                "spatial": {"type": "point", "coordinates": [lat, lon]},
                "temporal": {
                    "intervals": [
                        {
                            "start": start_climo_date.strftime(iso_8601),
                            "end": end_climo_date.strftime(iso_8601),
                        }
                    ]
                },
                "outputType": "json",
            }
            logging.info("Constructed query JSON for climatology data.")

            # Submit the query and process the results
            try:
                logging.info("Submitting query to IBM PAIRS API for climatology data.")
                df = query.submit(query_json, client=eis_client).point_data_as_dataframe()

                if df.empty:
                    logging.warning(f"No data returned for climatology data for {VARIABLE}")
                    continue

                # Apply temperature adjustment if the variable is TAVG
                if VARIABLE == "TAVG":
                    df["value"] = df["value"].astype(float) + temperature_adjustment

                logging.info(f"Climatology data retrieved, processing {len(df)} records.")

                # ------------------------------------------------------------------
                # Insert new rows and prune the oldest rows 1:1
                # ------------------------------------------------------------------
                new_rows = []
                for _, row in df.iterrows():
                    try:
                        timestamp_val = int(row["timestamp"])
                        value_val = float(row["value"])

                        # Convert timestamp (ms) to a Python date
                        cdate = datetime.utcfromtimestamp(timestamp_val / 1000).date()

                        # Build the ClimatologyData object
                        cdata = ClimatologyData(
                            city_name=city,
                            latitude=lat,
                            longitude=lon,
                            forecast_date=cdate,
                            variable=VARIABLE,
                            climatology_value=value_val,
                            source="IBM",
                        )
                        new_rows.append(cdata)
                    except Exception as row_e:
                        logging.error(f"Error parsing row for {VARIABLE}: {row_e}")
                        continue

                # Insert all new rows in bulk
                db.session.add_all(new_rows)
                db.session.flush()  # so they get assigned IDs

                # Prune the same number of oldest rows
                n_inserted = len(new_rows)
                if n_inserted > 0:
                    # Identify the oldest rows by ascending ID (or created_at if you prefer)
                    subq = (
                        db.session.query(ClimatologyData.id)
                        .order_by(ClimatologyData.id.asc())
                        .limit(n_inserted)
                        .subquery()
                    )
                    db.session.query(ClimatologyData).filter(
                        ClimatologyData.id.in_(subq)
                    ).delete(synchronize_session=False)

                # Commit once after insert + prune
                db.session.commit()

                # Cleanup
                gc.collect()
                df = None
                gc.collect()

            except Exception as e:
                logging.error(f"Error during climatology data query for {VARIABLE}: {e}")
                continue

    logging.info("Climatology data query completed successfully.")
    return "Climatology data query completed."





def process_and_store_climatology_data(df, city, lat, lon, VARIABLE):
    # Process and store data in the database
    for _, row in df.iterrows():
        try:
            date = int(row["timestamp"])
            value = float(row["value"])
            logging.debug(
                f"Processing climatology record - Date: {date}, Value: {value}"
            )

            # Convert timestamp to datetime
            forecast_date = datetime.utcfromtimestamp(date / 1000).date()

            # Store data in the Climatology database
            climatology_data = ClimatologyData(
                city_name=city,
                latitude=lat,
                longitude=lon,
                forecast_date=forecast_date,
                variable=VARIABLE,
                climatology_value=value,
                source="IBM",
            )
            db.session.add(climatology_data)
        except Exception as e:
            logging.error(f"Error processing climatology record for {VARIABLE}: {e}")
            logging.debug(f"Record data: {row}")
            continue

    db.session.commit()
    logging.info(f"Climatology data points for city {city} stored successfully.")


# Scheduler for API calls to USDA and Produce IQ
# Running the Scheduler
def schedule_jobs():
    # Define the Los Angeles timezone
    la_timezone = timezone("US/Pacific")

    scheduler = BackgroundScheduler(timezone=la_timezone)

    pk_timezone = timezone("Asia/Karachi")

    # Initialize the scheduler
    scheduler = BackgroundScheduler(timezone=pk_timezone)

    # Schedule ProduceIQ job to run every 2 hours but start at 12:25 AM to avoid overlap with weather forecast
    scheduler.add_job(
        func=fetch_daily_data,
        trigger=IntervalTrigger(
            hours=2, start_date="2024-10-01 00:25:00", timezone=la_timezone
        ),  # Runs every 2 hours, starting at 12:25 AM
        id="produce_iq_job",
    )

    # Schedule USDA job to run every 2 hours but 5 minutes after ProduceIQ job, i.e., at 12:30 AM
    scheduler.add_job(
        func=fetch_usda_daily_data,
        trigger=IntervalTrigger(
            hours=2, start_date="2024-10-01 00:30:00", timezone=la_timezone
        ),  # Runs every 2 hours, starting at 12:30 AM
        id="usda_job",
    )

    scheduler.start()
    logging.info(
        "Scheduler has started with adjusted timing for ProduceIQ, USDA, and Weather Forecast jobs."
    )


# Clean up sessions when the app context ends
@app.teardown_appcontext
def shutdown_session(exception=None):
    db.session.remove()


admin = Admin(app, name="Admin Panel")
admin.add_view(ModelView(User, db.session))
admin.add_view(ModelView(PriceData, db.session))  # Add PriceData to Admin panel


# User loader for Flask-Login
@login_manager.user_loader
def load_user(user_id):

    return User.query.get(int(user_id))


# Forms for login and registration
class RegisterForm(FlaskForm):
    username = StringField(
        "Username", validators=[InputRequired(), Length(min=4, max=150)]
    )
    email = StringField(
        "Email",
        validators=[InputRequired(), Email(message="Invalid email"), Length(max=150)],
    )
    password = PasswordField("Password", validators=[InputRequired(), Length(min=8)])
    confirm_password = PasswordField(
        "Confirm Password",
        validators=[
            InputRequired(),
            EqualTo("password", message="Passwords must match"),
        ],
    )
    role = SelectField(
        "Role", choices=[("sales", "Sales"), ("owner", "Owner"), ("admin", "Admin")]
    )
    submit = SubmitField("Register")


class LoginForm(FlaskForm):
    email = StringField(
        "Email", validators=[InputRequired(), Email(message="Invalid email")]
    )
    password = PasswordField("Password", validators=[InputRequired(), Length(min=8)])
    submit = SubmitField("Login")







# Home page
@app.route("/api/")
# @jwt_required()
def index():
    print("test1")
    if current_user.is_admin():
        return redirect(url_for("admin_dashboard"))
    elif current_user.is_owner():
        return redirect(url_for("owner_dashboard"))
    else:
        return redirect(url_for("sales_dashboard"))



# Current user route
@app.route("/api/current_user", methods=["GET"])
def get_current_user():
    print("Current User:", current_user)  # Debugging
    print("Is Authenticated:", current_user.is_authenticated)  # Check if true
    if current_user.is_authenticated:
        return jsonify({
            "isAuthenticated": True,
            "isAdmin": current_user.is_admin(),
            "isOwner": current_user.is_owner(),
            "username": current_user.username
        })
    return jsonify({
        "isAuthenticated": False,
        "isAdmin": False,
        "isOwner": False,
        "username": None
    })



# Dashboards
@app.route("/api/admin_dashboard", methods=["GET"])
# # @jwt_required()
def admin_dashboard():
    try:
        # Ensure current_user is available and authenticated
        if not current_user.is_authenticated:
            return jsonify({"error": "User not authenticated"}), 401

        # Return admin-specific data
        return jsonify({
            "message": "Welcome to the Admin Dashboard!",
            "username": current_user.username,
            "role": current_user.role,
        }), 200
    except Exception as e:
        # Log the exception
        app.logger.error(f"Error in /admin_dashboard: {str(e)}")
        return jsonify({"error": "Internal Server Error"}), 500

# @app.route("/admin_dashboard")
# @login_required
# def admin_dashboard():
#     return render_template("admin_dashboard.html")


@app.route("/api/owner_dashboard")
# # @jwt_required()
def owner_dashboard():
    return render_template("owner_dashboard.html")


@app.route("/api/weather_dashboard")
# # @jwt_required()
def weather_dashboard():
    return render_template("weather_dashboard.html")





@app.route("/api/sales_dashboard", methods=["GET"])
# @jwt_required()
def sales_dashboard_api():
    current_user = get_jwt_identity()

    # Role validation (optional)
    if current_user.get("role") not in ["admin", "sales"]:
        return jsonify({"error": "Access denied"}), 403

    try:
        selected_commodity = request.args.get("commodity", "Jalapeno")
        selected_source = request.args.get("source", "Historical")

        # Get the most recent year and day
        latest_year = db.session.query(db.func.max(PriceData.year)).scalar()
        latest_day = (
            db.session.query(db.func.max(PriceData.day))
            .filter_by(year=latest_year)
            .scalar()
        )

        # Query the best sell market
        best_market = (
            db.session.query(
                PriceData.city_name,
                db.func.max(PriceData.price).label("max_price"),
                PriceData.year,
                PriceData.day,
            )
            .filter(
                PriceData.commodity == selected_commodity,
                PriceData.source == selected_source,
                PriceData.year == latest_year,
                PriceData.day == latest_day,
            )
            .group_by(PriceData.city_name, PriceData.year, PriceData.day)
            .order_by(db.desc("max_price"))
            .all()
        )

        # Format the result
        formatted_best_market = [
            {
                "city_name": item.city_name,
                "max_price": item.max_price,
                "year": item.year,
                "day": item.day,
                "formatted_date": (datetime(item.year, 1, 1) + timedelta(days=item.day - 1)).strftime("%Y-%m-%d"),
            }
            for item in best_market
        ] if best_market else []

        return jsonify({
            "best_market": formatted_best_market,
            "selected_commodity": selected_commodity,
            "selected_source": selected_source,
        })

    except SQLAlchemyError as e:
        return jsonify({"error": "Database query failed", "details": str(e)}), 500



@app.route("/api/test_cache", methods=["GET"])
def test_cache():
    return jsonify({"message": "Hello from cache test"})



# Registration route
@app.route("/api/register", methods=["POST"])
def register():
    try:
        # Parse JSON data from the request
        data = request.json
        username = data.get("username")
        email = data.get("email")
        password = data.get("password")
        confirm_password = data.get("confirmPassword")
        role = data.get("role")

        # Validate the input data
        if not username or not email or not password or not confirm_password or not role:
            return jsonify({"error": "All fields are required"}), 400

        if password != confirm_password:
            return jsonify({"error": "Passwords do not match"}), 400

        # Check if the email already exists
        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            return jsonify({"error": "Email is already registered"}), 400

        # Hash the password
        hashed_password = generate_password_hash(password)

        # Create a new user instance
        new_user = User(
            username=username,
            email=email,
            password=hashed_password,
            role=role,
            approved=False,  # Default to not approved
        )
        logging.info(f"Received registration data: {data}")


        # Add the new user to the database
        db.session.add(new_user)
        db.session.commit()

        return jsonify({"message": "Registration successful!"}), 201

    except Exception as e:
        return jsonify({"error": str(e)}), 500




@app.route("/api/protected", methods=["GET"])
# @jwt_required()
def protected():
    current_user = get_jwt_identity()
    return jsonify({"logged_in_as": current_user}), 200


from datetime import timedelta

@app.route("/api/login", methods=["POST"])
def login():
    try:
        # Parse JSON data from the request
        data = request.json
        email = data.get("email")
        password = data.get("password")

        # Check if email and password are provided
        if not email or not password:
            return jsonify({"error": "Email and password are required"}), 400

        # Fetch the user from the database
        user = User.query.filter_by(email=email).first()

        # Validate the user and password
        if user and check_password_hash(user.password, password):
            if not user.approved:
                return jsonify({"error": "Your account is not approved yet. Please wait for approval."}), 403

            # Create a JWT access token with an expiration time
        
            access_token = create_access_token(
    identity=str(user.id),  # Use user.id or a unique string identifier
    additional_claims={"username": user.username, "role": user.role},
    expires_delta=timedelta(hours=5)
)

            return jsonify({
                "message": "Login successful!",
                "token": access_token,
                "role": user.role,
                "username": user.username,
                "user_id": user.id,
            }), 200
        else:
            return jsonify({"error": "Invalid email or password"}), 401
    except Exception as e:
        return jsonify({"error": str(e)}), 500



# Logout route

@app.route("/api/logout", methods=["POST"])
# @jwt_required()
def logout():
    # Add the token's JTI (unique identifier) to the blacklist
    jti = get_jwt()["jti"]
    blacklist.add(jti)
    return jsonify({"message": "Successfully logged out"}), 200

# Route for approving users


@app.route("/api/users", methods=["GET"])
# @jwt_required()
def approve_users():
    try:
        current_user_id = get_jwt_identity()  # Fetch the identity (user ID)
        claims = get_jwt()  # Fetch additional claims

        # Check role
        if claims['role'] not in ['admin', 'owner']:
            return jsonify({"error": "Access denied"}), 403

        # Fetch unapproved users
        unapproved_users = User.query.filter_by(approved=False).all()
        users = [{"id": user.id, "username": user.username, "email": user.email, "role": user.role} for user in unapproved_users]

        return jsonify({"users": users}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route("/api/approve_user/<int:user_id>", methods=["POST"])
# @jwt_required()
def approve_user(user_id):
    try:
        current_user = get_jwt_identity()  # Fetch identity from the token
        claims = get_jwt()  # Fetch additional claims

        # Check if the current user is an owner
        if claims['role'] != 'owner':
            return jsonify({"error": "Access denied"}), 403

        # Fetch the user to be approved
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        user.approved = True  # Approve the user
        db.session.commit()

        return jsonify({"message": f"User {user.username} approved successfully."}), 200
    except Exception as e:
        print(f"Error approving user: {e}")  # Debug log
        return jsonify({"error": "An error occurred while approving the user."}), 500




# FrontEND API internal
@app.route("/api/best_sell_market", methods=["GET"])
# @jwt_required()
def api_best_sell_market():
    selected_commodity = request.args.get("commodity", "Jalapeno")
    selected_source = request.args.get("source", "USDA")  # Default to USDA
    last7Days = request.args.get("last7Days", "false").lower() == "true"

    # Handle special case: If the commodity is "Cubanelles" and the source is "USDA", search for "Cubanelle"
    if selected_commodity == "Cubanelles" and selected_source == "USDA":
        selected_commodity = "Cubanelle"

    # Define the static cities
    cities = [
        "Baltimore",
        "Boston",
        "Chicago",
        "Columbia",
        "Miami",
        "New York",
        "Philadelphia",
        "Los Angeles",
    ]

    # Get the current US time and calculate the date 7 days ago
    us_time = datetime.now(timezone("US/Eastern"))
    seven_days_ago = us_time - timedelta(days=7)
    print(f"Current US time: {us_time}, 7 Days Ago: {seven_days_ago}")

    best_market_data = []

    for city in cities:

        query = (
            db.session.query(
                func.min(PriceData.price).label("min_price"),
                func.max(PriceData.price).label("max_price"),
                PriceData.year,
                PriceData.day,
            )
            .filter(
                func.upper(PriceData.city_name)
                == city.upper(),  # Convert to uppercase for case-insensitive comparison
                PriceData.commodity == selected_commodity,
                PriceData.source.in_(
                    ["Historical", selected_source]
                ),  # Either Historical or selected source
            )
            .group_by(PriceData.year, PriceData.day)
            .order_by(PriceData.year.desc(), PriceData.day.desc())
        )

        if last7Days:
            # Filter the query further if 'Only Consider Last 7 Days' is checked
            query = query.filter(
                PriceData.year >= seven_days_ago.year,
                PriceData.day >= seven_days_ago.timetuple().tm_yday,
            )

        # Execute the query and get the latest prices
        latest_prices = query.first()

        # If prices are found, format the range and return the prices
        if latest_prices:
            actual_date = datetime.strptime(
                f"{latest_prices.year}-{latest_prices.day}", "%Y-%j"
            ).strftime("%Y-%m-%d")
            price_range = (
                f"${latest_prices.min_price:.2f} - ${latest_prices.max_price:.2f}"
                if latest_prices.min_price != latest_prices.max_price
                else f"${latest_prices.max_price:.2f}"
            )
            best_market_data.append(
                {
                    "city_name": city,
                    "price_range": price_range,
                    "max_price": latest_prices.max_price,  # This will be used for sorting
                    "date": actual_date,
                }
            )

        else:
            # If no data is found for this city, return 'N/A'
            best_market_data.append(
                {
                    "city_name": city,
                    "price_range": "-",
                    "max_price": 0,  # Sorting placeholder
                    "date": "-",
                }
            )

    # Sort the cities by max price (highest to lowest), ensuring '-' prices go to the bottom
    best_market_data = sorted(
        best_market_data,
        key=lambda x: (
            x["max_price"] != "-",
            float(x["max_price"]) if x["max_price"] != "-" else float("-inf"),
        ),
        reverse=True,
    )

    return jsonify({"best_market": best_market_data})




from datetime import datetime, timedelta
from flask import Flask, request, jsonify
from sqlalchemy import func, or_, and_, true, select, over
from pytz import timezone

# Assume your app, SQLAlchemy (db), and PriceData model are already configured

@app.route("/api/most_recent_prices", methods=["GET"])
def api_most_recent_prices():
    # Define cities and commodities
    cities = [
        "Baltimore", "Boston", "Chicago", "Columbia",
        "Miami", "New York", "Philadelphia", "Los Angeles",
    ]
    commodities = [
        "Anaheim", "Cubanelles", "Fresno", "Habanero", "Hungarian Wax",
        "Jalapeno", "Long Hot", "Poblano", "Serrano", "Shishito",
    ]

    selected_source = request.args.get("source", "USDA")
    if selected_source == "Both":
        valid_sources = ["USDA", "ProduceIQ"]
    else:
        valid_sources = [selected_source]

    # Map commodities if needed
    commodity_map = {
        c: "Cubanelle" if (c == "Cubanelles" and selected_source == "USDA") else c
        for c in commodities
    }
    reverse_commodity_map = {v: k for k, v in commodity_map.items()}
    adjusted_commodities = list(commodity_map.values())

    # Create date filters for the past 7 days (US/Pacific timezone)
    tz = timezone("US/Pacific")
    seven_days_ago = datetime.now(tz) - timedelta(days=7)
    date_filters = [
        (d.year, d.timetuple().tm_yday)
        for d in [seven_days_ago + timedelta(days=i) for i in range(7)]
    ]

    city_lower_map = {city.lower(): city for city in cities}

    # Build filter conditions for cities and dates
    if cities:
        city_filter_conditions = or_(
            *[PriceData.city_name.ilike(city) for city in cities]
        )
    else:
        city_filter_conditions = true()

    if date_filters:
        date_filter_conditions = or_(
            *[and_(PriceData.year == year, PriceData.day == day) for year, day in date_filters]
        )
    else:
        date_filter_conditions = true()

    # --- First CTE: Aggregate average prices ---
    price_subquery_cte = (
        db.session.query(
            PriceData.commodity,
            PriceData.city_name,
            PriceData.year,
            PriceData.day,
            func.avg(func.nullif(PriceData.price, 0)).label("avg_price")
        )
        .filter(
            PriceData.commodity.in_(adjusted_commodities),
            city_filter_conditions,
            date_filter_conditions,
            PriceData.source.in_(valid_sources),
            PriceData.price != 0
        )
        .group_by(
            PriceData.commodity,
            PriceData.city_name,
            PriceData.year,
            PriceData.day,
        )
    ).cte("price_subquery")

    # --- Second CTE: Apply window function to rank prices ---
    window = over(
        func.row_number(),
        partition_by=[price_subquery_cte.c.commodity, price_subquery_cte.c.city_name],
        order_by=[price_subquery_cte.c.year.desc(), price_subquery_cte.c.day.desc()]
    )

    ranked_prices_cte = (
        select(
            price_subquery_cte.c.commodity,
            price_subquery_cte.c.city_name,
            price_subquery_cte.c.avg_price,
            window.label('rn')
        )
    ).cte("ranked_prices")

    # --- Final Query ---
    # Explicitly select from the ranked_prices_cte. Since that CTE
    # was built on price_subquery_cte, both will be included.
    final_query = (
        select(
            ranked_prices_cte.c.commodity,
            ranked_prices_cte.c.city_name,
            ranked_prices_cte.c.avg_price
        )
        .select_from(ranked_prices_cte)
        .where(ranked_prices_cte.c.rn == 1)
    )

    results = db.session.execute(final_query).all()

    # Build the dictionary for recent prices
    recent_prices = {
        commodity: {city: "-" for city in cities}
        for commodity in commodities
    }

    for row in results:
        original_commodity = reverse_commodity_map.get(row.commodity, row.commodity)
        normalized_city = row.city_name.lower()
        original_city = city_lower_map.get(normalized_city)
        if original_city and original_commodity in recent_prices:
            recent_prices[original_commodity][original_city] = float(row.avg_price) if row.avg_price is not None else "-"
    return jsonify({"prices": recent_prices})


@app.route("/api/sales_seasonal_prices", methods=["GET"])
def get_sales_seasonal_prices():
    # Fetch request parameters
    commodities_str = request.args.get("commodities")
    cities_str = request.args.get("cities")
    start_date_str = request.args.get("start_date")
    end_date_str = request.args.get("end_date")

    # Log received parameters
    app.logger.info(f"Received request with commodities: {commodities_str}, cities: {cities_str}, start_date: {start_date_str}, end_date: {end_date_str}")

    # Convert commodities and cities to lists if provided, else use all
    if commodities_str:
        commodities = commodities_str.split(",")
        commodities = [
            "Cubanelle" if commodity == "Cubanelles" else commodity
            for commodity in commodities
        ]
    else:
        commodities = [
            row[0] for row in db.session.query(PriceData.commodity).distinct().all()
        ]

    if cities_str:
        cities = cities_str.split(",")
       
    else:
        cities = [
            row[0] for row in db.session.query(PriceData.city_name).distinct().all()
        ]

    # Log converted commodities and cities
    app.logger.info(f"Commodities: {commodities}")
    app.logger.info(f"Cities: {cities}")

    # Convert start_date and end_date to datetime objects if provided
    start_date = None
    end_date = None
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD."}), 400

        if start_date > end_date:
            return jsonify({"error": "Start date cannot be after end date."}), 400

    # Log the dates
    app.logger.info(f"Start Date: {start_date}, End Date: {end_date}")

    # Prepare the query
    query = db.session.query(PriceData.season, PriceData.price).filter(
        PriceData.commodity.in_(commodities),
        func.lower(PriceData.city_name).in_([city.lower() for city in cities]),
        PriceData.source.in_(["USDA", "Historical"]),
    )

    # Apply date filters if both start_date and end_date are provided
    if start_date and end_date:
        start_year = start_date.year
        start_day = start_date.timetuple().tm_yday
        end_year = end_date.year
        end_day = end_date.timetuple().tm_yday

        # Log the filter conditions for date range
        app.logger.info(f"Query Filters: Commodities: {commodities}, Cities: {cities}, Start Year: {start_year}, End Year: {end_year}, Start Day: {start_day}, End Day: {end_day}")

        if start_year == end_year:
            query = query.filter(
                PriceData.year == start_year,
                PriceData.day >= start_day,
                PriceData.day <= end_day,
            )
        else:
            query = query.filter(
                or_(
                    and_(PriceData.year == start_year, PriceData.day >= start_day),
                    and_(PriceData.year == end_year, PriceData.day <= end_day),
                    and_(PriceData.year > start_year, PriceData.year < end_year),
                )
            )

    # Retrieve data - Ensure this happens in the proper order
    try:
        data = query.all()
        app.logger.info(f"Retrieved data: {data}")
    except Exception as e:
        app.logger.error(f"Error during query execution: {str(e)}")
        return jsonify({"error": "Error retrieving data from the database."}), 500

    # Calculate average prices per season
    seasonal_prices = {}
    season_price_data = {}

    for season, price in data:
        if season not in season_price_data:
            season_price_data[season] = []
        season_price_data[season].append(price)

    # Log the seasonal price data
    app.logger.info(f"Seasonal price data: {season_price_data}")

    # Calculate average price for each season
    for season in ["Spring", "Summer", "Autumn", "Winter"]:
        prices = season_price_data.get(season, [])
        if prices:
            average_price = sum(prices) / len(prices)
            seasonal_prices[season] = round(average_price, 2)
        else:
            seasonal_prices[season] = 0.0

    # Log final seasonal prices
    app.logger.info(f"Final seasonal prices: {seasonal_prices}")

    return jsonify(seasonal_prices)



@app.route("/api/historical_data", methods=["GET"])
def historical_data():
    try:
        # Fetch parameters from the frontend
        commodities = request.args.get("commodities", "").split(",")
        cities = request.args.get("cities", "").split(",")
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        source = request.args.get("source")
        avg_commodities = (
            request.args.get("averageCommodities", "false").lower() == "true"
        )
        avg_cities = request.args.get("averageCities", "false").lower() == "true"

        app.logger.info(f"Source: {source}")

        # Standardize Cubanelles
        standardized_commodities = [
            "Cubanelle" if commodity.lower().startswith("cubanelle") else commodity
            for commodity in commodities
        ]

        # Convert string dates to datetime objects
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")

        # Calculate day of the year for both start and end dates
        start_day = start_dt.timetuple().tm_yday
        end_day = end_dt.timetuple().tm_yday

        # Debug: Log date range
        app.logger.info(f"Start Day: {start_day}, End Day: {end_day}")

        # Query the database with proper date filtering
        query = PriceData.query.filter(
            func.upper(PriceData.commodity).in_([c.upper() for c in standardized_commodities]),
            func.upper(PriceData.city_name).in_([city.upper() for city in cities]),
            PriceData.source == source,
        )

        # Fix the date filtering logic
        if start_dt.year == end_dt.year:
            # If the start and end dates are in the same year
            query = query.filter(
                PriceData.year == start_dt.year,
                PriceData.day >= start_day,  # Include start_day
                PriceData.day <= end_day     # Include end_day
            )
        else:
            # If the start and end dates span multiple years
            query = query.filter(
                or_(
                    and_(PriceData.year == start_dt.year, PriceData.day >= start_day),
                    and_(PriceData.year == end_dt.year, PriceData.day <= end_day),
                    and_(PriceData.year > start_dt.year, PriceData.year < end_dt.year),
                )
            )

        data = query.all()
        app.logger.info(f"Query returned {len(data)} records")

        if not data:
            return jsonify({"labels": [], "datasets": []}), 200

        # Process and group data by date and commodity/city
        price_series = {}
        all_dates = set()

        for entry in data:
            entry_date = datetime(entry.year, 1, 1) + timedelta(days=entry.day - 1)
            date_str = entry_date.strftime("%Y-%m-%d")
            all_dates.add(date_str)

            display_commodity = (
                "Cubanelles"
                if entry.commodity.lower().startswith("cubanelle")
                else entry.commodity.strip().title()
            )
            display_city = entry.city_name.strip().lower().title()

            # Group data based on averaging preferences
            if avg_commodities and avg_cities:
                series_key = "Average Price"
            elif avg_commodities:
                series_key = display_city
            elif avg_cities:
                series_key = display_commodity
            else:
                series_key = f"{display_commodity} - {display_city}"

            if series_key not in price_series:
                price_series[series_key] = {}

            if date_str not in price_series[series_key]:
                price_series[series_key][date_str] = {"sum": 0, "count": 0}

            price_series[series_key][date_str]["sum"] += entry.price
            price_series[series_key][date_str]["count"] += 1

        # Sort the dates
        sorted_dates = sorted(list(all_dates))
        colors = ["#FF6384", "#36A2EB", "#FFCE56", "#4BC0C0", "#9966FF", "#FF9F40"]
        datasets = []

        # Create datasets for each series
        for idx, (series_name, date_data) in enumerate(price_series.items()):
            series_data = []
            for date in sorted_dates:
                if date in date_data:
                    avg_price = date_data[date]["sum"] / date_data[date]["count"]
                    series_data.append(round(avg_price, 2))
                else:
                    series_data.append(None)

            datasets.append(
                {
                    "label": series_name,
                    "data": series_data,
                    "borderColor": colors[idx % len(colors)],
                    "backgroundColor": colors[idx % len(colors)],
                }
            )

        # Prepare the final result
        result = {"labels": sorted_dates, "datasets": datasets}

        # Return the result as JSON
        return jsonify(result)

    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500


#  route for the shipping point price
@app.route("/api/shipping_point_price", methods=["GET"])
# @jwt_required()
def shipping_point_price():
    try:
        # Fetch parameters from the frontend
        commodities = request.args.get("commodities", "").split(",")
        regions = request.args.get("regions", "").split(",")
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        source = request.args.get("source")
        avg_commodities = (
            request.args.get("averageCommodities", "false").lower() == "true"
        )
        avg_regions = request.args.get("averageRegions", "false").lower() == "true"


        # Convert dates
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")

        # Calculate day of year for both dates
        start_day = start_dt.timetuple().tm_yday
        end_day = end_dt.timetuple().tm_yday

        # Debug: Log date range
        app.logger.info(f"Start Day: {start_day}, End Day: {end_day}")

        # Query the database with proper date filtering
        query = ShippingPriceData.query.filter(
            ShippingPriceData.commodity.in_(commodities),
            func.upper(ShippingPriceData.region_name).in_(
                [region.upper() for region in regions]
            ),
            ShippingPriceData.source == source,
        )

        # Add year-specific conditions
        if start_dt.year == end_dt.year:
            query = query.filter(
                ShippingPriceData.year == start_dt.year,
                ShippingPriceData.day.between(start_day, end_day),
            )
        else:
            query = query.filter(
                or_(
                    and_(ShippingPriceData.year == start_dt.year, ShippingPriceData.day >= start_day),
                    and_(ShippingPriceData.year == end_dt.year, ShippingPriceData.day <= end_day),
                    and_(ShippingPriceData.year > start_dt.year, ShippingPriceData.year < end_dt.year),
                )
            )

        query = query.order_by(ShippingPriceData.year.asc(), ShippingPriceData.day.asc())
        data = query.all()

        # Debug: Log raw query results
        app.logger.info(f"Query Results: {len(data)} records fetched.")

        if not data:
            return jsonify({"labels": [], "datasets": []}), 200

        # Process data with dates we actually have
        price_series = {}
        all_dates = set()

        # Group data based on averaging preferences
        for entry in data:
            entry_date = datetime(entry.year, 1, 1) + timedelta(days=entry.day - 1)

            if entry_date < start_dt or entry_date > end_dt:
                continue

            date_str = entry_date.strftime("%Y-%m-%d")
            all_dates.add(date_str)

            if avg_commodities and avg_regions:
                series_key = "Average Price"
            elif avg_commodities:
                series_key = entry.region_name
            elif avg_regions:
                series_key = entry.commodity
            else:
                series_key = f"{entry.commodity} - {entry.region_name}"

            if series_key not in price_series:
                price_series[series_key] = {}

            if date_str not in price_series[series_key]:
                price_series[series_key][date_str] = {"sum": 0, "count": 0}

            price_series[series_key][date_str]["sum"] += entry.price
            price_series[series_key][date_str]["count"] += 1

        # Debug: Log processed series
        app.logger.info(f"Price Series: {price_series}")

        # Sort dates
        sorted_dates = sorted(list(all_dates))

        # Create datasets for each series
        colors = ["#FF6384", "#36A2EB", "#FFCE56", "#4BC0C0", "#9966FF", "#FF9F40"]
        datasets = []

        for idx, (series_name, date_data) in enumerate(price_series.items()):
            series_data = []
            for date in sorted_dates:
                if date in date_data:
                    avg_price = date_data[date]["sum"] / date_data[date]["count"]
                    series_data.append(round(avg_price, 2))
                else:
                    series_data.append(None)

            datasets.append(
                {
                    "label": series_name,
                    "data": series_data,
                    "borderColor": colors[idx % len(colors)],
                    "backgroundColor": colors[idx % len(colors)],
                }
            )

        # Cleanup
        del data, price_series, all_dates
        gc.collect()

        # Debug: Log datasets
        app.logger.info(f"Datasets: {datasets}")

        return jsonify({"labels": sorted_dates, "datasets": datasets})

    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500




@app.route("/api/download_historical_data", methods=["GET"])
def download_historical_data():
    # Fetch the parameters from the frontend
    commodities = request.args.get("commodities").split(",")
    cities = request.args.get("cities").split(",")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    source = request.args.get("source")

    # Handle special case for "Cubanelles"
    commodities = [
        "Cubanelle" if commodity == "Cubanelles" and source == "USDA" else commodity
        for commodity in commodities
    ]

    # Convert start_date and end_date into year and day of the year
    start_year = datetime.strptime(start_date, "%Y-%m-%d").year
    start_day = datetime.strptime(start_date, "%Y-%m-%d").timetuple().tm_yday
    end_year = datetime.strptime(end_date, "%Y-%m-%d").year
    end_day = datetime.strptime(end_date, "%Y-%m-%d").timetuple().tm_yday

    # Query the PriceData table based on the filters
    query = PriceData.query.filter(
        PriceData.commodity.in_(commodities),
        PriceData.city_name.in_(cities),
        PriceData.source == source,
        or_(
            (PriceData.year == start_year) & (PriceData.day >= start_day),
            (PriceData.year == end_year) & (PriceData.day <= end_day),
            (PriceData.year > start_year) & (PriceData.year < end_year),
        ),
    ).order_by(PriceData.year.asc(), PriceData.day.asc())

    # Fetch the data
    data = query.all()

    # Create an Excel file in memory
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    worksheet = workbook.add_worksheet()

    # Write the headers to the Excel file
    headers = ["Date", "City", "Commodity", "Price"]
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header)

    # Write the data rows to the Excel file
    for row_num, entry in enumerate(data, start=1):
        # Reconstruct the date from year and day
        actual_date = datetime.strptime(f"{entry.year}-{entry.day}", "%Y-%j").strftime(
            "%Y-%m-%d"
        )
        worksheet.write(row_num, 0, actual_date)
        worksheet.write(row_num, 1, entry.city_name)
        worksheet.write(row_num, 2, entry.commodity)
        worksheet.write(row_num, 3, entry.price)

    workbook.close()
    output.seek(0)

    # Send the Excel file as a response
    return send_file(
        output,
        as_attachment=True,
        download_name="historical_data.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# The API endpoint to serve the stored weather data for a specific location
# helper functions for weather forecasting visualizations


# Helper function to aggregate climatology data across all years for each day of the year
def get_daily_climatology(lat, lon):
    climatology = ClimatologyData.query.filter_by(latitude=lat, longitude=lon).all()

    # Aggregate climatology by day of the year
    daily_climatology = {}
    for c in climatology:
        day_of_year = c.forecast_date.strftime("%m-%d")
        if day_of_year not in daily_climatology:
            daily_climatology[day_of_year] = []
        daily_climatology[day_of_year].append(c.climatology_value)

    # Average climatology data for each day of the year
    avg_daily_climatology = {
        day: np.mean(values) for day, values in daily_climatology.items()
    }
    return avg_daily_climatology


def get_daily_precip_climatology(lat, lon):
    climatology = ClimatologyData.query.filter_by(
        latitude=lat, longitude=lon, variable="PRECIP"
    ).all()

    # Aggregate climatology by day of the year (month and day)
    daily_climatology = {}
    for c in climatology:
        day_of_year = c.forecast_date.strftime("%m-%d")
        if day_of_year not in daily_climatology:
            daily_climatology[day_of_year] = []
        daily_climatology[day_of_year].append(c.climatology_value)

    # Calculate the average for each day of the year
    avg_daily_climatology = {
        day: np.mean(values) for day, values in daily_climatology.items()
    }

    return avg_daily_climatology


# Fetch and aggregate forecast data with min, max, and standard deviation by ensembles
def fetch_forecast_data(lat, lon, start_date, end_date):
    forecasts = (
        WeatherForecast.query.filter_by(latitude=lat, longitude=lon)
        .filter(
            WeatherForecast.forecast_date >= start_date,
            WeatherForecast.forecast_date <= end_date,
        )
        .all()
    )

    # Group forecast data by date and variable
    forecast_data = {"TAVG": {}, "PRECIP": {}}
    for f in forecasts:
        date_key = f.forecast_date.strftime("%Y-%m-%d")
        if date_key not in forecast_data[f.variable]:
            forecast_data[f.variable][date_key] = []
        forecast_data[f.variable][date_key].append(f.forecasted_value)

    # Calculate avg, min, max, and standard deviation for each day
    result = {
        "TAVG": {"dates": [], "avg": [], "min": [], "max": [], "std_dev": []},
        "PRECIP": {"dates": [], "avg": [], "min": [], "max": [], "std_dev": []},
    }

    for variable in forecast_data:
        for date, values in forecast_data[variable].items():
            avg = np.mean(values)
            min_val = np.min(values)
            max_val = np.max(values)
            std_dev = np.std(values)

            result[variable]["dates"].append(date)
            result[variable]["avg"].append(avg)
            result[variable]["min"].append(min_val)
            result[variable]["max"].append(max_val)
            result[variable]["std_dev"].append(std_dev)

    return result


# Calculate the accumulated climatology for the date range (blue line) and ensemble totals (gray bars)
def calculate_accumulated_precipitation(lat, lon, start_date, end_date):
    # Get daily climatology for precipitation
    daily_precip_climatology = get_daily_precip_climatology(lat, lon)

    # Calculate accumulated climatology precipitation
    accumulated_climo_precip = 0
    date = start_date
    while date <= end_date:
        day_of_year = date.strftime("%m-%d")
        if day_of_year in daily_precip_climatology:
            accumulated_climo_precip += daily_precip_climatology[day_of_year]
        date += timedelta(days=1)

    # Query forecasts
    forecasts = (
        WeatherForecast.query.filter_by(latitude=lat, longitude=lon, variable="PRECIP")
        .filter(
            WeatherForecast.forecast_date >= start_date,
            WeatherForecast.forecast_date <= end_date,
        )
        .all()
    )

    if not forecasts:
        print("Warning: No forecast data found for the given range.")
        return {
            "accumulated_climo_precip": accumulated_climo_precip,
            "ensemble_totals": [],
            "probability_below_climo": None,
        }

    # Initialize ensemble totals
    ensemble_totals = {}
    for f in forecasts:
        if f.ensemble_member not in ensemble_totals:
            ensemble_totals[f.ensemble_member] = 0
        ensemble_totals[f.ensemble_member] += f.forecasted_value

    if len(ensemble_totals) == 0:
        print("Warning: No ensemble data available.")
        return {
            "accumulated_climo_precip": accumulated_climo_precip,
            "ensemble_totals": [],
            "probability_below_climo": None,
        }

    # Calculate probability
    ensembles_below_climo = sum(
        1 for total in ensemble_totals.values() if total < accumulated_climo_precip
    )
    probability_below_climo = (ensembles_below_climo / len(ensemble_totals)) * 100

    return {
        "accumulated_climo_precip": accumulated_climo_precip,
        "ensemble_totals": list(ensemble_totals.values()),
        "probability_below_climo": probability_below_climo,
    }

# def calculate_accumulated_precipitation(lat, lon, start_date, end_date):
#     # Get daily climatology for precipitation (average per historical record per day of the year)
#     daily_precip_climatology = get_daily_precip_climatology(lat, lon)

#     # Calculate the accumulated climatology precipitation for the date range
#     accumulated_climo_precip = 0
#     date = start_date
#     while date <= end_date:
#         day_of_year = date.strftime("%m-%d")
#         if day_of_year in daily_precip_climatology:
#             accumulated_climo_precip += daily_precip_climatology[day_of_year]
#             print(accumulated_climo_precip)
#         date += timedelta(days=1)

#     # Calculate accumulated precipitation for each ensemble across the entire date range
#     forecasts = (
#         WeatherForecast.query.filter_by(latitude=lat, longitude=lon, variable="PRECIP")
#         .filter(
#             WeatherForecast.forecast_date >= start_date,
#             WeatherForecast.forecast_date <= end_date,
#         )
#         .all()
#     )

#     # Initialize ensemble totals as a dictionary
#     ensemble_totals = {}
#     for f in forecasts:
#         if f.ensemble_member not in ensemble_totals:
#             ensemble_totals[f.ensemble_member] = 0
#         # Accumulate values for each ensemble member over the entire date range
#         ensemble_totals[f.ensemble_member] += f.forecasted_value

#     # Calculate the probability (percentage) of ensembles predicting lower than climatology
#     ensembles_below_climo = sum(
#         1 for total in ensemble_totals.values() if total < accumulated_climo_precip
#     )
#     probability_below_climo = (ensembles_below_climo / len(ensemble_totals)) * 100

#     return {
#         "accumulated_climo_precip": accumulated_climo_precip,
#         "ensemble_totals": list(ensemble_totals.values()),
#         "probability_below_climo": probability_below_climo,
#     }


# Helper function to aggregate temperature climatology data across all years for each day of the year
def get_daily_tavg_climatology_v2(lat, lon):
    climatology = ClimatologyData.query.filter_by(
        latitude=lat, longitude=lon, variable="TAVG"
    ).all()

    # Aggregate climatology by day of the year (ignoring year)
    daily_climatology = {}
    for c in climatology:
        day_of_year = c.forecast_date.strftime("%m-%d")
        if day_of_year not in daily_climatology:
            daily_climatology[day_of_year] = []
        daily_climatology[day_of_year].append(c.climatology_value)

    # Average climatology data for each day of the year
    avg_daily_climatology = {
        day: np.mean(values) for day, values in daily_climatology.items()
    }
    return avg_daily_climatology


# Helper function to aggregate precipitation climatology data across all years for each day of the year
def get_daily_precip_climatology_v2(lat, lon):
    climatology = ClimatologyData.query.filter_by(
        latitude=lat, longitude=lon, variable="PRECIP"
    ).all()

    # Aggregate climatology by day of the year (ignoring year)
    daily_climatology = {}
    for c in climatology:
        day_of_year = c.forecast_date.strftime("%m-%d")
        if day_of_year not in daily_climatology:
            daily_climatology[day_of_year] = []
        daily_climatology[day_of_year].append(c.climatology_value)

    # Calculate the average for each day of the year
    avg_daily_climatology = {
        day: np.mean(values) for day, values in daily_climatology.items()
    }

    return avg_daily_climatology




# # Main API route: Serve aggregated forecast and climatology data
# @app.route("/api/weather_forecasts", methods=["GET"])
# def api_weather_forecasts():
#     lat = float(request.args.get("lat"))
#     lon = float(request.args.get("lon"))
#     start = request.args.get("start")
#     end = request.args.get("end")

#     # Set default date range if start or end dates are not provided
#     if not start:
#         start_date = datetime.now()
#     else:
#         start_date = datetime.strptime(start, "%Y-%m-%d")

#     if not end:
#         end_date = start_date + timedelta(
#             days=30
#         )  # Default to 30 days from start date if end date is not provided
#     else:
#         end_date = datetime.strptime(end, "%Y-%m-%d")

#     try:
#         # Fetch forecast data with min, max, std_dev for TAVG and PRECIP
#         forecast_data = fetch_forecast_data(lat, lon, start_date, end_date)

#         # Fetch climatology data for temperature and precipitation
#         daily_tavg_climatology = get_daily_tavg_climatology_v2(lat, lon)
#         daily_precip_climatology = get_daily_precip_climatology_v2(lat, lon)

#         # Combine TAVG and PRECIP into daily_climatology
#         daily_climatology = {
#             "TAVG": daily_tavg_climatology,
#             "PRECIP": daily_precip_climatology,
#         }

#         # Calculate accumulated precipitation and ensemble analysis for the selected date range
#         accumulation_data = calculate_accumulated_precipitation(
#             lat, lon, start_date, end_date
#         )

#         # Format response JSON
#         return jsonify(
#             {
#                 "forecast_data": forecast_data,
#                 "daily_climatology": daily_climatology,
#                 "accumulation_data": accumulation_data,
#             }
#         )
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500


@app.route("/api/weather_forecasts", methods=["GET"])
def api_weather_forecasts():
    try:
        lat = float(request.args.get("lat"))
        lon = float(request.args.get("lon"))
        start = request.args.get("start")
        end = request.args.get("end")

        print(f"Request Params - lat: {lat}, lon: {lon}, start: {start}, end: {end}")

        # Set default date range if start or end dates are not provided
        if not start:
            start_date = datetime.now()
        else:
            start_date = datetime.strptime(start, "%Y-%m-%d")

        if not end:
            end_date = start_date + timedelta(days=30)
        else:
            end_date = datetime.strptime(end, "%Y-%m-%d")

        print(f"Date Range - start_date: {start_date}, end_date: {end_date}")

        # Fetch data
        forecast_data = fetch_forecast_data(lat, lon, start_date, end_date)
        print("Forecast Data:", forecast_data)

        daily_tavg_climatology = get_daily_tavg_climatology_v2(lat, lon)
        daily_precip_climatology = get_daily_precip_climatology_v2(lat, lon)

        daily_climatology = {
            "TAVG": daily_tavg_climatology,
            "PRECIP": daily_precip_climatology,
        }
        print("Daily Climatology:", daily_climatology)

        accumulation_data = calculate_accumulated_precipitation(
            lat, lon, start_date, end_date
        )
        print("Accumulation Data:", accumulation_data)

        return jsonify(
            {
                "forecast_data": forecast_data,
                "daily_climatology": daily_climatology,
                "accumulation_data": accumulation_data,
            }
        )
    except Exception as e:
        print("Error in /api/weather_forecasts:", str(e))
        return jsonify({"error": str(e)}), 500



# City list for dropdowns
@app.route("/api/cities", methods=["GET"])
def get_cities():
    cities = [
        {"name": "Sinaloa", "lat": 25.1721, "lon": -107.4795},
        {"name": "Sonora", "lat": 29.2972, "lon": -110.3309},
        {"name": "Ensenada", "lat": 31.86613056, "lon": -116.59971944},
        {"name": "Baja Mx", "lat": 28.0444, "lon": -115.2062},
        {"name": "Culican", "lat": 24.8091, "lon": -107.3940},
        {"name": "Hendersonville, NC", "lat": 35.3187, "lon": -82.4610},
        {"name": "Cameron SC", "lat": 33.5568, "lon": -80.7151},
        {"name": "Adel, Ga", "lat": 31.13633333, "lon": -83.42216389},
        {"name": "Lake Park Ga", "lat": 30.6844, "lon": -83.1849},
        {"name": "Bowling Green Fl", "lat": 27.6386, "lon": -81.8265},
        {"name": "Immokalee Fl", "lat": 26.4187, "lon": -81.4173},
        {"name": "Palm Beach County, FL", "lat": 26.7153, "lon": -80.0534},
        {"name": "Vineland NJ", "lat": 39.4802, "lon": -75.0138},
        {"name": "Sodus, Michigan", "lat": 42.0086, "lon": -86.3614},
    ]
    return jsonify(cities)


# Forecast Calculator
# def calculate_forecasted_price(variety, start_date, forecast_date):
#     # Step 1: Determine the season of the forecast date
#     season = determine_season_for_dashboard(forecast_date)

#     # Step 2: Convert start date to year and day of the year
#     start_year = start_date.year
#     start_day = start_date.timetuple().tm_yday

#     # Step 3: Query the PriceData for the given variety and season, starting from the start_date
#     historical_data = (
#         db.session.query(PriceData.price)
#         .filter(
#             PriceData.commodity == variety,          # Match the commodity (variety)
#             PriceData.season == season,                # Match the season
#             PriceData.year >= start_year,              # Consider data from the start year onward
#             PriceData.day >= start_day,                # Ensure data is after the start date in the year
#             PriceData.source == "ProduceIQ",
#         )
#         .all()
#     )

#     # Step 4: Calculate the average price from the historical data
#     if historical_data:
#         total_price = sum(entry.price for entry in historical_data)
#         average_price = total_price / len(historical_data)
#     else:
#         average_price = 0.0

#     return average_price




# @app.route("/api/calculate_forecast", methods=["POST"])
# def calculate_forecast():
#     data = request.json

#     # Required form fields
#     variety = data.get("variety")
#     start_date_str = data.get("start_date")
#     forecast_date_str = data.get("forecast_date")
#     yield_per_acre = data.get("yield_per_acre")

#     # NEW cost fields
#     cost_per_acre = data.get("cost_per_acre", 0)
#     harvest_cost_per_box = data.get("harvest_cost_per_box", 0)
#     cost_of_box = data.get("cost_of_box", 0)
#     boxes_bonus_per_yield = data.get("boxes_bonus_per_yield", 0)

#     # Validate the main required fields
#     if not all([variety, start_date_str, forecast_date_str, yield_per_acre]):
#         return jsonify({"error": "All form fields are required"}), 400

#     try:
#         # Convert to correct data types
#         start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
#         forecast_date = datetime.strptime(forecast_date_str, "%Y-%m-%d")
#         yield_per_acre = float(yield_per_acre)

#         # Convert the cost fields to float; default to 0 if missing
#         cost_per_acre = float(cost_per_acre) if cost_per_acre else 0
#         harvest_cost_per_box = float(harvest_cost_per_box) if harvest_cost_per_box else 0
#         cost_of_box = float(cost_of_box) if cost_of_box else 0
#         boxes_bonus_per_yield = float(boxes_bonus_per_yield) if boxes_bonus_per_yield else 0

#     except ValueError:
#         return jsonify({"error": "Invalid date or numeric format."}), 400

#     # 1) Calculate the forecasted price (already have your helper function):
#     forecasted_price = calculate_forecasted_price(variety, start_date, forecast_date)
#     revenue_per_acre = forecasted_price * yield_per_acre

#     # 2) Determine the season
#     season = determine_season_for_dashboard(forecast_date)

#     # 3) Calculate total costs:
#     # cost_per_acre + (harvest_cost_per_box * yield_per_acre) + (cost_of_box * yield_per_acre) + (boxes_bonus_per_yield)
#     total_costs = cost_per_acre  + (harvest_cost_per_box * yield_per_acre)  + (cost_of_box * yield_per_acre)  + (boxes_bonus_per_yield)

#     # 4) Subtract total costs from revenue
#     revenue_after_costs = revenue_per_acre - total_costs


#         # Add the debug print here:
#     print("DEBUG: forecasted_price =", forecasted_price,
#           " revenue_per_acre_after_costings =", revenue_after_costs)


#     return jsonify({
#         "forecasted_price": round(forecasted_price, 2),
#         "revenue_per_acre": round(revenue_per_acre, 2),
#         "revenue_per_acre_after_costings": round(revenue_after_costs, 2),
#         "season": season,
#     })


def calculate_forecasted_price(variety, start_date, forecast_date, city=None):
    # Step 1: Determine the season of the forecast date
    season = determine_season_for_dashboard(forecast_date)

    # Step 2: Convert start date to year and day of the year
    start_year = start_date.year
    start_day = start_date.timetuple().tm_yday

    # Step 3: Start building the query
    query = (
        db.session.query(PriceData.price)
        .filter(
            PriceData.commodity == variety,          # Match the commodity (variety)
            PriceData.season == season,              # Match the season
            PriceData.year >= start_year,            # Consider data from the start year onward
            PriceData.day >= start_day,              # Ensure data is after the start date in the year
            PriceData.source == "ProduceIQ",
        )
    )
    
    # Add city filter if provided
    if city and city != "All cities":
        query = query.filter(PriceData.city_name == city)
    
    # Execute the query
    historical_data = query.all()

    # Step 4: Calculate the average price from the historical data
    if historical_data:
        total_price = sum(entry.price for entry in historical_data)
        average_price = total_price / len(historical_data)
    else:
        average_price = 0.0

    return average_price



@app.route("/api/calculate_forecast", methods=["POST"])
def calculate_forecast():
    data = request.json

    # Required form fields
    variety = data.get("variety")
    start_date_str = data.get("start_date")
    forecast_date_str = data.get("forecast_date")
    yield_per_acre = data.get("yield_per_acre")
    
    # Get the city parameter (optional)
    city = data.get("city", "All cities")

    # NEW cost fields
    cost_per_acre = data.get("cost_per_acre", 0)
    harvest_cost_per_box = data.get("harvest_cost_per_box", 0)
    cost_of_box = data.get("cost_of_box", 0)
    boxes_bonus_per_yield = data.get("boxes_bonus_per_yield", 0)

    # Validate the main required fields
    if not all([variety, start_date_str, forecast_date_str, yield_per_acre]):
        return jsonify({"error": "All form fields are required"}), 400

    try:
        # Convert to correct data types
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        forecast_date = datetime.strptime(forecast_date_str, "%Y-%m-%d")
        yield_per_acre = float(yield_per_acre)

        # Convert the cost fields to float; default to 0 if missing
        cost_per_acre = float(cost_per_acre) if cost_per_acre else 0
        harvest_cost_per_box = float(harvest_cost_per_box) if harvest_cost_per_box else 0
        cost_of_box = float(cost_of_box) if cost_of_box else 0
        boxes_bonus_per_yield = float(boxes_bonus_per_yield) if boxes_bonus_per_yield else 0

    except ValueError:
        return jsonify({"error": "Invalid date or numeric format."}), 400

    # 1) Calculate the forecasted price (now with city parameter):
    forecasted_price = calculate_forecasted_price(variety, start_date, forecast_date, city)
    revenue_per_acre = forecasted_price * yield_per_acre

    # Rest of function remains the same...

    # 2) Determine the season
    season = determine_season_for_dashboard(forecast_date)

    # 3) Calculate total costs:
    # cost_per_acre + (harvest_cost_per_box * yield_per_acre) + (cost_of_box * yield_per_acre) + (boxes_bonus_per_yield)
    total_costs = cost_per_acre  + (harvest_cost_per_box * yield_per_acre)  + (cost_of_box * yield_per_acre)  + (boxes_bonus_per_yield)
    # 4) Subtract total costs from revenue
    revenue_after_costs = revenue_per_acre - total_costs


        # Add the debug print here:
    print("DEBUG: forecasted_price =", forecasted_price,
          " revenue_per_acre_after_costings =", revenue_after_costs)


    return jsonify({
        "forecasted_price": round(forecasted_price, 2),
        "revenue_per_acre": round(revenue_per_acre, 2),
        "revenue_per_acre_after_costings": round(revenue_after_costs, 2),
        "season": season,
    })


# Endpoint to get price averages with normalized commodity names
@app.route('/api/price_averages', methods=['GET'])
def get_price_averages():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        source = request.args.get('source', 'both').lower()
        city = request.args.get('city', 'All cities')
        
        if not start_date or not end_date:
            return jsonify({'error': 'Both start and end dates are required'}), 400
        
        # Convert dates to year/day-of-year
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        start_year, start_day = start_date_obj.year, start_date_obj.timetuple().tm_yday
        end_year, end_day = end_date_obj.year, end_date_obj.timetuple().tm_yday
        
        # Mapping USDA city names to normalized names
        usda_city_mapping = {
            "BALTIMORE": "Baltimore",
            "BOSTON": "Boston",
            "CHICAGO": "Chicago",
            "COLUMBIA": "Columbia",
            "LOS ANGELES": "Los Angeles",
            "MIAMI": "Miami",
            "NEW YORK": "New York",
            "PHILADELPHIA": "Philadelphia",
        }
        
        # Normalize city names
        normalized_city = case(
            *[(PriceData.city_name == key, value) for key, value in usda_city_mapping.items()],
            else_=PriceData.city_name
        ).label("normalized_city")
        
        # Normalize commodity name - ensure consistent naming between sources
        # Handle special case where Cubanelle in USDA matches Cubanelles in ProduceIQ
        normalized_commodity = func.lower(func.trim(PriceData.commodity)).label("normalized_commodity")
        
        # Build basic query filtering by date range
        query = db.session.query(
            normalized_city,
            normalized_commodity,
            PriceData.source,
            func.avg(PriceData.price).label('avg_price')
        ).filter(
            or_(
                and_(PriceData.year == start_year, PriceData.day >= start_day),
                and_(PriceData.year == end_year, PriceData.day <= end_day),
                and_(PriceData.year > start_year, PriceData.year < end_year)
            )
        )
        
        # Apply city filter if not "All cities"
        if city.lower() != 'all cities':
            query = query.filter(normalized_city == city)
        
        # Apply source filter if not "both"
        if source == 'usda':
            query = query.filter(PriceData.source == "USDA")
        elif source == 'produceiq':
            query = query.filter(PriceData.source == "ProduceIQ")
        
        # Group by city, commodity, and source
        query = query.group_by(normalized_city, normalized_commodity, PriceData.source)
        query = query.order_by(normalized_commodity)
        
        results = query.all()
        
        # Convert query results to dictionary format for easier manipulation
        raw_data = []
        for row in results:
            # Skip if no data (shouldn't happen with avg, but just in case)
            if row.avg_price is None:
                continue
                
            raw_data.append({
                'city': row.normalized_city,
                'commodity': row.normalized_commodity,
                'source': row.source,
                'avg_price': round(row.avg_price, 2)
            })
        
        # Format the results based on filters
        price_averages = []
        
        # Special handling for the "cubanelle/cubanelles" case
        cubanelle_map = {
            "cubanelle": "cubanelles",
            "cubanelles": "cubanelles"
        }
        
        # Standardize commodity names
        for item in raw_data:
            if item['commodity'] in cubanelle_map:
                item['commodity'] = cubanelle_map[item['commodity']]
        
        # For "both" sources, we need to combine USDA and ProduceIQ
        if source == 'both':
            # Create a dictionary to combine by commodity (and city if relevant)
            combined_data = {}
            
            if city.lower() == 'all cities':
                # Combine by commodity across all cities and both sources
                for item in raw_data:
                    commodity = item['commodity']
                    if commodity not in combined_data:
                        combined_data[commodity] = {
                            'commodity': commodity,
                            'source': 'U/P',
                            'sum_price': item['avg_price'],
                            'count': 1
                        }
                    else:
                        combined_data[commodity]['sum_price'] += item['avg_price']
                        combined_data[commodity]['count'] += 1
                
                # Calculate averages and create the final output
                for commodity, data in combined_data.items():
                    price_averages.append({
                        'commodity': commodity,
                        'source': 'U/P',
                        'avg_price': round(data['sum_price'] / data['count'], 2)
                    })
            else:
                # Combine by commodity and city
                for item in raw_data:
                    key = (item['commodity'], item['city'])
                    if key not in combined_data:
                        combined_data[key] = {
                            'commodity': item['commodity'],
                            'city': item['city'],
                            'source': 'Both',
                            'sum_price': item['avg_price'],
                            'count': 1
                        }
                    else:
                        combined_data[key]['sum_price'] += item['avg_price']
                        combined_data[key]['count'] += 1
                
                # Calculate averages
                for data in combined_data.values():
                    price_averages.append({
                        'commodity': data['commodity'],
                        'city_name': data['city'],
                        'source': 'U/P',
                        'avg_price': round(data['sum_price'] / data['count'], 2)
                    })
        else:
            # For single source queries
            if city.lower() == 'all cities':
                # Combine by commodity across all cities
                commodity_data = {}
                for item in raw_data:
                    commodity = item['commodity']
                    if commodity not in commodity_data:
                        commodity_data[commodity] = {
                            'commodity': commodity,
                            'source': item['source'],
                            'sum_price': item['avg_price'],
                            'count': 1
                        }
                    else:
                        commodity_data[commodity]['sum_price'] += item['avg_price']
                        commodity_data[commodity]['count'] += 1
                
                # Calculate averages
                for commodity, data in commodity_data.items():
                    price_averages.append({
                        'commodity': commodity,
                        'source': data['source'],
                        'avg_price': round(data['sum_price'] / data['count'], 2)
                    })
            else:
                # Direct mapping for specific city and source
                for item in raw_data:
                    price_averages.append({
                        'commodity': item['commodity'],
                        'city_name': item['city'],
                        'source': item['source'],
                        'avg_price': item['avg_price']
                    })
        
        # Sort results by commodity for consistency
        price_averages.sort(key=lambda x: x['commodity'])
        
        return jsonify({'price_averages': price_averages}), 200
    
    except Exception as e:
        app.logger.error(f'Error fetching price averages: {str(e)}')
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500

# API FOR Forecast visual
@app.route("/api/seasonal_prices", methods=["GET"])
def get_seasonal_prices():
    variety = request.args.get("variety")
    city = request.args.get("city", "All cities")
    start_date_str = request.args.get("start_date")
    forecast_date_str = request.args.get("forecast_date")  # Added forecast date parameter
    
    # Check if required parameters are missing
    if not variety:
        return jsonify({"error": "Missing variety"}), 400
        
    # Parse dates if provided
    start_date = None
    forecast_date = None
    start_year = None
    start_day = None
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            start_year = start_date.year
            start_day = start_date.timetuple().tm_yday
        except ValueError:
            return jsonify({"error": "Invalid start date format"}), 400
            
    if forecast_date_str:
        try:
            forecast_date = datetime.strptime(forecast_date_str, "%Y-%m-%d")
        except ValueError:
            return jsonify({"error": "Invalid forecast date format"}), 400

    # Initialize seasonal_prices dictionary
    seasonal_prices = {"Spring": 0, "Summer": 0, "Autumn": 0, "Winter": 0}
    
    # Loop through each season and calculate the average price
    for season in seasonal_prices.keys():
        # Create a query similar to calculate_forecasted_price
        query = (
            db.session.query(PriceData.price)
            .filter(
                PriceData.commodity == variety,  # Match the commodity (variety)
                PriceData.season == season,      # Match the season
                PriceData.source == "ProduceIQ",
            )
        )
        
        # Add date filters if start date is provided
        if start_date:
            query = query.filter(
                PriceData.year >= start_year,     # Consider data from the start year onward
                PriceData.day >= start_day,       # Ensure data is after the start date in the year
            )
        
        # Add city filter if provided and not "All cities"
        if city and city != "All cities":
            query = query.filter(PriceData.city_name == city)
            
        # Execute the query
        historical_data = query.all()

        # Calculate the average price from the historical data
        if historical_data:
            total_price = sum([entry.price for entry in historical_data])
            average_price = total_price / len(historical_data)
            seasonal_prices[season] = round(average_price, 2)
        else:
            seasonal_prices[season] = 0.0
            
        # Log for debugging
        print(f"Season {season} for {variety}, City: {city}, Start: {start_date_str}, Forecast: {forecast_date_str} - Found {len(historical_data)} records, Avg: {seasonal_prices[season]}")

    return jsonify(seasonal_prices)



@app.route("/api/forecast_line_data", methods=["GET"])
def get_forecast_line_data():
    """
    Build season‑by‑season price forecasts.

    • If averageCities=true   → ignore the city filter and aggregate across every city
    • Otherwise               → forecast for each (commodity, city) pair exactly as before
    """
    try:
        from collections import defaultdict

        # ────────── 1. Parse & validate input ──────────
        commodities = [c.strip() for c in request.args.get("commodities", "").split(",") if c.strip()]
        cities_raw  = [c.strip() for c in request.args.get("cities", "").split(",") if c.strip()]
        avg_cities  = request.args.get("averageCities", "false").lower() == "true"
        forecast_years = int(request.args.get("forecastYears", "1"))

        if not commodities:
            return jsonify({"error": "Missing commodities"}), 400
        if not avg_cities and not cities_raw:
            return jsonify({"error": "Missing cities"}), 400

        # remove sentinel "ALL" if it slipped through
        cities = [c for c in cities_raw if c.upper() != "ALL"]

        # ────────── 2. Season helpers ──────────
        seasons = ["Winter", "Spring", "Summer", "Autumn"]
        now     = datetime.now()
        current_year = now.year
        month   = now.month
        current_season = (
            "Spring" if 3 <= month <= 5 else
            "Summer" if 6 <= month <= 8 else
            "Autumn" if 9 <= month <= 11 else
            "Winter"
        )

        # Build season labels: current season + forecast_years * 4
        season_labels = []
        idx0 = seasons.index(current_season)
        for yr in range(current_year, current_year + forecast_years + 1):
            for i in range(4):
                season_labels.append(f"{seasons[(idx0 + i) % 4]} {yr}")
        season_labels = season_labels[: 1 + 4 * forecast_years]

        result = {"labels": season_labels, "datasets": []}

        # ────────── 3. One DB query for all needed history ──────────
        filters = [
            PriceData.commodity.in_(commodities),
            PriceData.year >= current_year - 5,
            PriceData.source == "ProduceIQ",
        ]
        if not avg_cities:
            filters.append(PriceData.city_name.in_(cities))

        rows = (
            db.session.query(
                PriceData.commodity,
                PriceData.city_name,
                PriceData.year,
                PriceData.season,
                PriceData.price,
            )
            .filter(*filters)
            .all()
        )

        # If we’re averaging across cities, remember which cities actually came back
        if avg_cities:
            all_cities_in_db = sorted({r.city_name for r in rows})

        # group: data[key][season][year] = [prices…]
        data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        for r in rows:
            key = f"{r.commodity}_{r.city_name}"
            data[key][r.season][r.year].append(r.price)

        # ────────── 4. Seasonal averages & trend factors ──────────
        season_stats = {}     # season_stats[key][season] = {avg_price, trend}
        seasons_list = seasons  # alias for clarity
        for key, season_dict in data.items():
            season_stats[key] = {}
            for season in seasons_list:
                yearly_avgs, years_list = [], []
                for yr, prices in season_dict[season].items():
                    if prices:
                        yearly_avgs.append(sum(prices) / len(prices))
                        years_list.append(yr)

                overall_avg = sum(yearly_avgs) / len(yearly_avgs) if yearly_avgs else 0
                # default trend = +2 %/yr
                trend = 0.02
                n = len(yearly_avgs)
                if n >= 2:
                    sx, sy = sum(years_list), sum(yearly_avgs)
                    sxx = sum(x * x for x in years_list)
                    sxy = sum(x * y for x, y in zip(years_list, yearly_avgs))
                    denom = n * sxx - sx * sx
                    if denom:
                        slope = (n * sxy - sx * sy) / denom
                        trend = slope / (sy / n) if sy else 0.02

                season_stats[key][season] = {"avg": overall_avg, "trend": trend}

        # ────────── 5. Build datasets ──────────
        COLORS = ["#FF6384", "#36A2EB", "#FFCE56", "#4BC0C0", "#9966FF", "#FF9F40"]
        ds_idx = 0

        if avg_cities:
            # average across *all* cities present in rows
            for commodity in commodities:
                prices = []
                for label in season_labels:
                    season, yr = label.split()
                    yr = int(yr)
                    yrs_out = yr - current_year

                    total, cnt = 0, 0
                    for city in all_cities_in_db:
                        key = f"{commodity}_{city}"
                        st = season_stats.get(key, {}).get(season)
                        if st and st["avg"] > 0:
                            total += st["avg"] * (1 + st["trend"] * yrs_out)
                            cnt += 1
                    prices.append(round(total / cnt, 2) if cnt else 0)

                result["datasets"].append(
                    {
                        "label": f"{commodity} – Avg Across Cities",
                        "data": prices,
                        "borderColor": COLORS[ds_idx % len(COLORS)],
                        "backgroundColor": COLORS[ds_idx % len(COLORS)],
                        "borderDash": [5, 5],
                    }
                )
                ds_idx += 1
        else:
            # one dataset per (commodity, city)
            for commodity in commodities:
                for city in cities:
                    key = f"{commodity}_{city}"
                    prices = []
                    for label in season_labels:
                        season, yr = label.split()
                        yr = int(yr)
                        yrs_out = yr - current_year
                        st = season_stats.get(key, {}).get(season)
                        if st and st["avg"] > 0:
                            prices.append(round(st["avg"] * (1 + st["trend"] * yrs_out), 2))
                        else:
                            prices.append(0)

                    result["datasets"].append(
                        {
                            "label": f"{commodity} – {city} Forecast",
                            "data": prices,
                            "borderColor": COLORS[ds_idx % len(COLORS)],
                            "backgroundColor": COLORS[ds_idx % len(COLORS)],
                            "borderDash": [5, 5],
                        }
                    )
                    ds_idx += 1

        return jsonify(result)

    except Exception as exc:
        app.logger.error(f"forecast_line_data error: {exc}")
        return jsonify({"error": str(exc)}), 500





@app.route("/api/volatility_data", methods=["GET"])
def get_volatility_data():
    try:
        # Parse request parameters
        commodity = request.args.get("commodity", "")  # Single string for commodity
        cities = request.args.get("cities", "").split(",")
        time_frame = request.args.get("timeFrame", "1m")  # Time frame parameter
        
        # Parse date range parameters
        start_date_str = request.args.get("startDate", "")
        end_date_str = request.args.get("endDate", "")
        
        # Check if commodity or cities are missing or empty
        if not commodity or not cities or cities[0] == '':
            return jsonify({"error": "Missing commodity or cities"}), 400
            
        # Parse start and end dates
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            except ValueError:
                return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
        else:
            # Default to current year if dates aren't provided
            end_date = datetime.now()
            start_date = datetime(end_date.year, 1, 1)  # January 1st of current year
            
        # Convert time frame to number of days
        days_to_fetch = 30  # Default to 1m
        if time_frame == '1d':
            days_to_fetch = 1
        elif time_frame == '3d':
            days_to_fetch = 3
        elif time_frame == '7d':
            days_to_fetch = 7
        elif time_frame == '14d':
            days_to_fetch = 14
        # 1m is default 30 days
        
        # Define result structure
        result = {
            "labels": [],
            "datasets": []
        }
        
        # Calculate the number of intervals based on date range and time frame
        # This determines how many candlesticks we'll have
        date_delta = (end_date - start_date).days + 1
        
        # For short time frames with few intervals, we'll show detailed labels
        if time_frame in ['1d', '3d', '7d']:
            # Generate daily labels based on date range
            if date_delta <= 90:  # If less than 3 months, show daily dates
                current_date = start_date
                labels = []
                while current_date <= end_date:
                    labels.append(current_date.strftime("%b %d"))  # Format like "Apr 18"
                    current_date += timedelta(days=1)
                result["labels"] = labels
            else:
                # If more than 90 days, group into the time frame intervals
                # Calculate how many intervals are required to cover the date range
                num_intervals = max(1, date_delta // days_to_fetch)
                current_date = start_date
                labels = []
                
                for i in range(num_intervals):
                    interval_end = min(current_date + timedelta(days=days_to_fetch - 1), end_date)
                    label = f"{current_date.strftime('%b %d')} - {interval_end.strftime('%b %d')}"
                    labels.append(label)
                    current_date += timedelta(days=days_to_fetch)
                    if current_date > end_date:
                        break
                        
                result["labels"] = labels

        elif time_frame == '14d':
            # For 14-day intervals, create bi-weekly labels
            # Calculate how many 14-day periods fit in the date range
            num_intervals = max(1, date_delta // 14)
            current_date = start_date
            labels = []
            
            for i in range(num_intervals):
                interval_end = min(current_date + timedelta(days=13), end_date)  # 14 days including start date
                label = f"{current_date.strftime('%b %d')} - {interval_end.strftime('%b %d')}"
                labels.append(label)
                current_date += timedelta(days=14)
                if current_date > end_date:
                    break
                    
            result["labels"] = labels
                
        elif time_frame == '1m':
            # For monthly view, group by months
            months_data = {}
            current_date = start_date
            labels = []
            
            # Generate month labels
            while current_date <= end_date:
                month_key = current_date.strftime("%Y-%m")
                month_label = current_date.strftime("%B %Y")
                
                if month_key not in months_data:
                    months_data[month_key] = {"label": month_label}
                    labels.append(month_label)
                    
                current_date += timedelta(days=1)
                # If we've moved to a new month, update current_date to first day of that month
                if current_date.day == 1:
                    pass  # Already at the first day of a month
                
            result["labels"] = labels
        
        # Fetch price range data (min/max/open/close) for the selected commodity, across cities
        price_range_data = calculate_price_range_for_timeframe(
            commodity, cities, start_date, end_date, time_frame, days_to_fetch
        )
            
        # Colors for the dataset
        colors = ["#FF6384", "#36A2EB", "#FFCE56", "#4BC0C0", "#9966FF", "#FF9F40"]
        color_index = 0
        
        # Create dataset for this commodity's minimum prices
        min_dataset = {
            "label": f"{commodity} - Min Price",
            "data": price_range_data.get("min", [0] * len(result["labels"])),
            "borderColor": colors[color_index % len(colors)],
            "backgroundColor": "transparent",
            "borderWidth": 2,
            "borderDash": [],
            "pointRadius": 3,
            "fill": 'false',
            "type": "line"
        }
        
        # Create dataset for this commodity's maximum prices
        max_dataset = {
            "label": f"{commodity} - Max Price",
            "data": price_range_data.get("max", [0] * len(result["labels"])),
            "borderColor": colors[color_index % len(colors)],
            "backgroundColor": colors[color_index % len(colors)] + "33",  # Add transparency
            "borderWidth": 2,
            "borderDash": [],
            "pointRadius": 3,
            "fill": "-1",  # Fill to previous dataset (min price)
            "type": "line"
        }
        
        # Create datasets for opening and closing prices (needed for candlestick)
        open_dataset = {
            "label": f"{commodity} - Open Price",
            "data": price_range_data.get("open", [0] * len(result["labels"])),
            "hidden": True,  # Hide from regular charts
            "borderColor": "transparent",
            "backgroundColor": "transparent",
            "pointRadius": 0,
            "fill": 'false'
        }
        
        close_dataset = {
            "label": f"{commodity} - Close Price",
            "data": price_range_data.get("close", [0] * len(result["labels"])),
            "hidden": True,  # Hide from regular charts
            "borderColor": "transparent",
            "backgroundColor": "transparent",
            "pointRadius": 0,
            "fill": 'false'
        }
        
        # Add datasets to result
        result["datasets"].append(min_dataset)
        result["datasets"].append(max_dataset)
        result["datasets"].append(open_dataset)
        result["datasets"].append(close_dataset)
        
        # Get the latest price data
        latest_price_data = get_latest_price(commodity, cities)
        
        # Add it to the result dictionary
        result["latest_price"] = latest_price_data
        
        return jsonify(result)
        
    except Exception as e:
        app.logger.error(f"Error in price range data: {str(e)}")
        return jsonify({"error": str(e)}), 500

def calculate_price_range_for_timeframe(commodity, cities, start_date, end_date, time_frame, days_to_fetch):
    """
    Calculate price ranges (min, max, open, close) for the selected commodity,
    aggregated across all selected cities, using the same averaging approach as historical_data.
    """
    try:
        from collections import defaultdict
        
        result = {
            "min": [],
            "max": [],
            "open": [],  # First price in the period
            "close": []  # Last price in the period
        }
        
        # Convert dates to days of year for filtering
        start_year = start_date.year
        end_year = end_date.year
        
        start_day_of_year = start_date.timetuple().tm_yday
        end_day_of_year = end_date.timetuple().tm_yday
        
        # Query conditions to filter by date range
        query_conditions = [
            PriceData.commodity == commodity,
            PriceData.city_name.in_(cities),
            PriceData.source == "ProduceIQ"
        ]
        
        # Handle date range query differently depending on if it spans multiple years
        if start_year == end_year:
            # Single year query
            query_conditions.extend([
                PriceData.year == start_year,
                PriceData.day >= start_day_of_year,
                PriceData.day <= end_day_of_year
            ])
            
            all_data = db.session.query(
                PriceData.day,
                PriceData.price,
                PriceData.year,
                PriceData.city_name
            ).filter(*query_conditions).all()
            
        else:
            # Multi-year query (need to handle each year separately)
            # First year: from start_day to end of year
            first_year_conditions = query_conditions.copy()
            first_year_conditions.extend([
                PriceData.year == start_year,
                PriceData.day >= start_day_of_year
            ])
            
            # Last year: from beginning of year to end_day
            last_year_conditions = query_conditions.copy()
            last_year_conditions.extend([
                PriceData.year == end_year,
                PriceData.day <= end_day_of_year
            ])
            
            # Middle years (if any): entire years
            middle_years = list(range(start_year + 1, end_year))
            middle_year_data = []
            
            if middle_years:
                middle_year_conditions = query_conditions.copy()
                middle_year_conditions.append(PriceData.year.in_(middle_years))
                middle_year_data = db.session.query(
                    PriceData.day,
                    PriceData.price,
                    PriceData.year,
                    PriceData.city_name
                ).filter(*middle_year_conditions).all()
            
            # Query for first and last year data
            first_year_data = db.session.query(
                PriceData.day,
                PriceData.price,
                PriceData.year,
                PriceData.city_name
            ).filter(*first_year_conditions).all()
            
            last_year_data = db.session.query(
                PriceData.day,
                PriceData.price,
                PriceData.year,
                PriceData.city_name
            ).filter(*last_year_conditions).all()
            
            # Combine all the data
            all_data = first_year_data + middle_year_data + last_year_data
        
        # First, average data by day and city (just like in historical_data)
        daily_city_data = {}
        
        for row in all_data:
            # Create a date object for this data point
            row_date = datetime(row.year, 1, 1) + timedelta(days=row.day - 1)
            date_key = row_date.strftime("%Y-%m-%d")
            city_key = row.city_name
            
            # Create compound key for date+city
            compound_key = f"{date_key}_{city_key}"
            
            # Initialize the structure for this date+city if it doesn't exist
            if compound_key not in daily_city_data:
                daily_city_data[compound_key] = {
                    "sum": 0, 
                    "count": 0, 
                    "date": row_date,
                    "city": city_key
                }
            
            # Sum up prices and count entries for this date+city
            daily_city_data[compound_key]["sum"] += row.price
            daily_city_data[compound_key]["count"] += 1
        
        # Then average across cities for each day
        daily_avg_data = {}
        
        for key, data in daily_city_data.items():
            date_str = data["date"].strftime("%Y-%m-%d")
            
            if date_str not in daily_avg_data:
                daily_avg_data[date_str] = {"sum": 0, "count": 0, "date": data["date"]}
            
            # Add the city average to the daily total
            city_avg = data["sum"] / data["count"]
            daily_avg_data[date_str]["sum"] += city_avg
            daily_avg_data[date_str]["count"] += 1
        
        # Convert daily averages to date-price pairs
        daily_data = []
        for date_str, data in daily_avg_data.items():
            if data["count"] > 0:  # Ensure we have data for this day
                avg_price = data["sum"] / data["count"]
                daily_data.append((data["date"], avg_price))
        
        # Sort by date
        daily_data.sort(key=lambda x: x[0])
        
        # Process based on time frame
        if time_frame == '1m':
            # Group by month
            data_by_month = defaultdict(list)
            data_by_month_with_dates = defaultdict(list)
            
            for date, price in daily_data:
                month_key = date.strftime("%Y-%m")
                data_by_month[month_key].append(price)
                data_by_month_with_dates[month_key].append((date, price))
            
            # Sort months chronologically
            sorted_months = sorted(data_by_month.keys())
            
            # Calculate min, max, open, and close for each month
            for month in sorted_months:
                prices = data_by_month[month]
                date_price_pairs = data_by_month_with_dates[month]
                
                if prices:
                    # Sort by date to determine open (first) and close (last)
                    date_price_pairs.sort(key=lambda x: x[0])
                    open_price = date_price_pairs[0][1]  # First price in month
                    close_price = date_price_pairs[-1][1]  # Last price in month
                    
                    result["min"].append(round(min(prices), 2))
                    result["max"].append(round(max(prices), 2))
                    result["open"].append(round(open_price, 2))
                    result["close"].append(round(close_price, 2))
                else:
                    result["min"].append(0)
                    result["max"].append(0)
                    result["open"].append(0)
                    result["close"].append(0)
        else:
            # For other time frames, group by the specified interval
            # Convert all data to actual dates for easier grouping
            
            # Group data into intervals based on time_frame
            interval_data = defaultdict(list)
            
            current_interval_start = start_date
            interval_index = 0
            
            # Store data by interval with date info for sorting
            interval_data_with_dates = defaultdict(list)
            
            for date, price in daily_data:
                # Check if this data point belongs to current interval or we need to move to next interval
                while date > current_interval_start + timedelta(days=days_to_fetch - 1):
                    # Move to next interval
                    current_interval_start += timedelta(days=days_to_fetch)
                    interval_index += 1
                    
                    # If we've moved past the end date, break
                    if current_interval_start > end_date:
                        break
                        
                # If date is within current interval, add the price
                if current_interval_start <= date <= current_interval_start + timedelta(days=days_to_fetch - 1):
                    interval_data[interval_index].append(price)
                    interval_data_with_dates[interval_index].append((date, price))
            
            # Calculate min, max, open, and close for each interval
            for i in range(len(interval_data)):
                prices = interval_data.get(i, [])
                date_price_pairs = interval_data_with_dates.get(i, [])
                
                if prices:
                    # Sort by date to determine open (first) and close (last)
                    date_price_pairs.sort(key=lambda x: x[0])
                    open_price = date_price_pairs[0][1]  # First price in interval
                    close_price = date_price_pairs[-1][1]  # Last price in interval
                    
                    result["min"].append(round(min(prices), 2))
                    result["max"].append(round(max(prices), 2))
                    result["open"].append(round(open_price, 2))
                    result["close"].append(round(close_price, 2))
                else:
                    result["min"].append(0)
                    result["max"].append(0)
                    result["open"].append(0)
                    result["close"].append(0)
        
        return result

    except Exception as e:
        app.logger.error(f"Error calculating price ranges: {str(e)}")
        return {"min": [], "max": [], "open": [], "close": []}

def get_latest_price(commodity, cities):
    """
    Get the absolute most recent price for the specified commodity and cities.
    """
    try:
        # Simple query to get the latest price
        latest_price_query = db.session.query(
            PriceData.price,
            PriceData.year, 
            PriceData.day
        ).filter(
            PriceData.commodity == commodity,
            PriceData.city_name.in_(cities),
            PriceData.source == "ProduceIQ"
        ).order_by(
            PriceData.year.desc(),
            PriceData.day.desc()
        ).first()
        
        if not latest_price_query:
            return {
                "price": 0,
                "date": None
            }
        
        # Calculate the date
        latest_date = datetime(latest_price_query.year, 1, 1) + timedelta(days=latest_price_query.day - 1)
        
        # Format the result
        result = {
            "price": round(latest_price_query.price, 2),
            "date": latest_date.strftime("%b %d, %Y")
        }
        
        return result
        
    except Exception as e:
        app.logger.error(f"Error getting latest price: {str(e)}")
        return {
            "price": 0,
            "date": None
        }


# Helper function to get the first and last day of a month
def get_month_day_range(month, year):
    """
    Returns the first and last day of the month as days of the year.
    """
    # First day of the month
    first_day = datetime(year, month, 1)
    first_day_of_year = first_day.timetuple().tm_yday
    
    # Last day of the month (first day of next month - 1 day)
    if month == 12:
        last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1) - timedelta(days=1)
        
    last_day_of_year = last_day.timetuple().tm_yday
    
    return first_day_of_year, last_day_of_year

@app.route("/api/harvest_planning", methods=["POST"])
def harvest_planning():
    try:
        payload = request.json
        if not payload or "varieties" not in payload:
            return jsonify({"error": "Missing varieties data"}), 400

        result = []
        # ──────── shared forecast logic ──────── #
        def compute_forecast(commodity, cities, avg_cities, forecast_years=1):
            # 1) season setup
            seasons = ["Winter", "Spring", "Summer", "Autumn"]
            now = datetime.now()
            current_year = now.year
            month = now.month
            current_season = (
                "Spring" if 3 <= month <= 5 else
                "Summer" if 6 <= month <= 8 else
                "Autumn" if 9 <= month <= 11 else
                "Winter"
            )
            # build labels
            idx0 = seasons.index(current_season)
            season_labels = [
                f"{seasons[(idx0 + i) % 4]} {yr}"
                for yr in range(current_year, current_year + forecast_years + 1)
                for i in range(4)
            ][ : 1 + 4 * forecast_years ]

            # 2) pull your PriceData rows
            filters = [
                PriceData.commodity == commodity,
                PriceData.year >= current_year - 5,
                PriceData.source == "ProduceIQ",
            ]
            if not avg_cities:
                filters.append(PriceData.city_name.in_(cities))

            rows = (
                db.session.query(
                    PriceData.season,
                    PriceData.year,
                    PriceData.price
                )
                .filter(*filters)
                .all()
            )

            # 3) aggregate into season_stats
            data = defaultdict(lambda: defaultdict(list))
            for season, year, price in rows:
                data[season][year].append(price)

            season_stats = {}
            for season in seasons:
                yearly_avgs = [
                    sum(prs) / len(prs)
                    for yr, prs in data[season].items()
                    if prs
                ]
                overall_avg = sum(yearly_avgs) / len(yearly_avgs) if yearly_avgs else 0
                # simple trend = +2% per year (or compute via regression if you like)
                trend = 0.02
                season_stats[season] = {"avg": overall_avg, "trend": trend}

            # 4) build a single series of prices
            prices = []
            for label in season_labels:
                season, yr_s = label.split()
                yr = int(yr_s)
                yrs_out = yr - current_year
                st = season_stats.get(season)
                if st and st["avg"] > 0:
                    prices.append(round(st["avg"] * (1 + st["trend"] * yrs_out), 2))
                else:
                    prices.append(0)

            return {"labels": season_labels, "data": prices}

        # ──────── main loop ──────── #
        for variety in payload["varieties"]:
            # validate
            for f in ("name", "plantingDate", "harvestingDate", "growingDays"):
                if f not in variety:
                    return jsonify({"error": f"Missing field {f}"}), 400

            # parse dates & calculate harvest date
            plant = datetime.fromisoformat(variety["plantingDate"].replace("Z","+00:00"))
            harvest = datetime.fromisoformat(variety["harvestingDate"].replace("Z","+00:00"))
            grow_days = int(variety["growingDays"])
            calc_harvest = plant + timedelta(days=grow_days)

            commodity = variety["name"]
            market = variety.get("market","").strip()
            # handle “Select All” as avg_cities=True
            if market.lower() in ("select all","all","national",""):
                avg_cities = True
                cities = []
            else:
                avg_cities = False
                cities = [market]

            # get one‐year forecast
            fc = compute_forecast(commodity, cities, avg_cities, forecast_years=1)

            # find max price & which label it is
            highest_price = 0
            best_season = best_year = None
            for i, price in enumerate(fc["data"]):
                if price > highest_price:
                    highest_price = price
                    season, yr_s = fc["labels"][i].split()
                    best_season, best_year = season, int(yr_s)

            # pick a midpoint date for that season
            best_date = None
            if best_season and best_year:
                month_map = {"Winter":1, "Spring":4, "Summer":7, "Autumn":10}
                d = month_map[best_season]
                best_date = datetime(best_year, d, 15)
                # ensure it's after harvest
                if best_date < harvest:
                    best_date = datetime(best_year+1, d, 15)

            result.append({
                "name": commodity,
                "plantingDate": plant.isoformat(),
                "harvestingDate": harvest.isoformat(),
                "calculatedHarvestDate": calc_harvest.isoformat(),
                "growingDays": grow_days,
                "bestSellingTime": {
                    "season": best_season,
                    "year": best_year,
                    "price": highest_price,
                    "date": best_date.isoformat() if best_date else None
                }
            })

        return jsonify({"varieties": result})

    except Exception as e:
        app.logger.exception("Error in harvest_planning")
        return jsonify({"error": str(e)}), 500






# Existing calculate_forecast helper functions can be reused
from sqlalchemy import desc

@app.route("/api/break_even/save", methods=["POST"])
def save_break_even_estimation():
    """
    Save a break-even estimation to the database
    Uses the same data from calculate_forecast and adds additional fields
    """
    data = request.json
    
    try:
        # Extract all fields from the request
        variety = data.get("variety")
        city = data.get("city")  # Added for forecast line data
        start_date_str = data.get("start_date")
        forecast_date_str = data.get("forecast_date")
        yield_per_acre = data.get("yield_per_acre")
        cost_per_acre = data.get("cost_per_acre", 0)
        harvest_cost_per_box = data.get("harvest_cost_per_box", 0)
        cost_of_box = data.get("cost_of_box", 0)
        boxes_bonus_per_yield = data.get("boxes_bonus_per_yield", 0)
        start_date_range_str = data.get("start_date_range")
        end_date_range_str = data.get("end_date_range")
        
        # Validate required fields
        if not all([
            variety, city, start_date_str, forecast_date_str, yield_per_acre, 
            start_date_range_str, end_date_range_str
        ]):
            return jsonify({"error": "Missing required fields"}), 400
        
        # Convert string dates to datetime objects
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        forecast_date = datetime.strptime(forecast_date_str, "%Y-%m-%d")
        start_date_range = datetime.strptime(start_date_range_str, "%Y-%m-%d")
        end_date_range = datetime.strptime(end_date_range_str, "%Y-%m-%d")
        
        # Convert numeric fields to float
        yield_per_acre = float(yield_per_acre)
        cost_per_acre = float(cost_per_acre)
        harvest_cost_per_box = float(harvest_cost_per_box)
        cost_of_box = float(cost_of_box)
        boxes_bonus_per_yield = float(boxes_bonus_per_yield)
        
        # Calculate forecasted price using the same function as the forecast component
        forecasted_price = calculate_forecasted_price(variety, start_date, forecast_date, city)
        revenue_per_acre = forecasted_price * yield_per_acre
        
        # Calculate total costs
        total_costs = cost_per_acre + (harvest_cost_per_box * yield_per_acre) + (cost_of_box * yield_per_acre) + boxes_bonus_per_yield
        
        # Calculate revenue after costs
        revenue_after_costs = revenue_per_acre - total_costs
        
        # Determine season
        season = determine_season_for_dashboard(forecast_date)
        
        # Calculate revenue per box
        revenue_per_box = revenue_after_costs / yield_per_acre if yield_per_acre > 0 else 0
        
        # Create new BreakEvenEstimation instance
        estimation = BreakEvenEstimation(
            variety=variety,
            city=city,
            start_date=start_date,
            forecast_date=forecast_date,
            yield_per_acre=yield_per_acre,
            cost_per_acre=cost_per_acre,
            harvest_cost_per_box=harvest_cost_per_box,
            cost_of_box=cost_of_box,
            boxes_bonus_per_yield=boxes_bonus_per_yield,
            start_date_range=start_date_range,
            end_date_range=end_date_range,
            forecasted_price=forecasted_price,
            revenue_per_acre=revenue_per_acre,
            revenue_after_costs=revenue_after_costs,
            revenue_per_box=revenue_per_box,
            season=season
        )
        
        db.session.add(estimation)
        db.session.commit()
        
        return jsonify({
            "id": estimation.id,
            "message": "Break-even estimation saved successfully"
        }), 201
        
    except ValueError as e:
        return jsonify({"error": f"Invalid data: {str(e)}"}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Failed to save estimation: {str(e)}"}), 500
    

@app.route("/api/break_even", methods=["GET"])
def get_break_even_estimations():
    """Get all break-even estimations with guaranteed fresh data."""
    try:
        # Always bypass cache for this route
        print(f"Fetching fresh break-even estimations for route: {request.path}")
        
        # Query all estimations, ordered by creation date (newest first)
        estimations = BreakEvenEstimation.query.order_by(desc(BreakEvenEstimation.created_at)).all()
        
        # Convert to dictionaries for JSON response
        result = [estimation.to_dict() for estimation in estimations]
        
        # Completely remove any cached data
        cache.clear()  # This clears the entire cache
        
        # Create response with aggressive no-cache headers
        response = jsonify(result)
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0, s-maxage=0, proxy-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        response.headers['X-Accel-Expires'] = '0'  # Nginx cache control
        
        return response, 200
    
    except Exception as e:
        print(f"Error fetching break-even estimations: {str(e)}")
        
        response = jsonify({
            "error": "Failed to retrieve break-even estimations",
            "details": str(e)
        })
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0, s-maxage=0, proxy-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        response.headers['X-Accel-Expires'] = '0'
        
        return response, 500


@app.route("/api/break_even/<int:estimation_id>", methods=["GET"])
def get_break_even_estimation(estimation_id):
    """Get a specific break-even estimation"""
    try:
        # Query the estimation by ID
        estimation = BreakEvenEstimation.query.filter_by(id=estimation_id).first()
        
        if not estimation:
            return jsonify({"error": "Estimation not found"}), 404
        
        return jsonify(estimation.to_dict()), 200
    except Exception as e:
        return jsonify({"error": f"Failed to retrieve estimation: {str(e)}"}), 500

@app.route("/api/break_even/<int:estimation_id>", methods=["DELETE"])
def delete_break_even_estimation(estimation_id):
    """Delete a break-even estimation"""
    try:
        # Query the estimation by ID
        estimation = BreakEvenEstimation.query.filter_by(id=estimation_id).first()
        
        if not estimation:
            return jsonify({"error": "Estimation not found"}), 404
        
        # Delete the estimation
        db.session.delete(estimation)
        db.session.commit()
        
        return jsonify({"message": "Estimation deleted successfully"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Failed to delete estimation: {str(e)}"}), 500

@app.route("/api/break_even/<int:estimation_id>", methods=["PUT"])
def update_break_even_estimation(estimation_id):
    """Update an existing break-even estimation"""
    data = request.json
    
    try:
        # Query the estimation by ID
        estimation = BreakEvenEstimation.query.filter_by(id=estimation_id).first()
        
        if not estimation:
            return jsonify({"error": "Estimation not found"}), 404
        
        # Update fields if provided
        if "variety" in data:
            estimation.variety = data["variety"]
        if "city" in data:
            estimation.city = data["city"]
        if "start_date" in data:
            estimation.start_date = datetime.strptime(data["start_date"], "%Y-%m-%d")
        if "forecast_date" in data:
            estimation.forecast_date = datetime.strptime(data["forecast_date"], "%Y-%m-%d")
        if "yield_per_acre" in data:
            estimation.yield_per_acre = float(data["yield_per_acre"])
        if "cost_per_acre" in data:
            estimation.cost_per_acre = float(data["cost_per_acre"])
        if "harvest_cost_per_box" in data:
            estimation.harvest_cost_per_box = float(data["harvest_cost_per_box"])
        if "cost_of_box" in data:
            estimation.cost_of_box = float(data["cost_of_box"])
        if "boxes_bonus_per_yield" in data:
            estimation.boxes_bonus_per_yield = float(data["boxes_bonus_per_yield"])
        if "start_date_range" in data:
            estimation.start_date_range = datetime.strptime(data["start_date_range"], "%Y-%m-%d")
        if "end_date_range" in data:
            estimation.end_date_range = datetime.strptime(data["end_date_range"], "%Y-%m-%d")
        
        # Recalculate derived fields if any inputs changed
        if any(key in data for key in [
            "variety", "city", "start_date", "forecast_date", "yield_per_acre", 
            "cost_per_acre", "harvest_cost_per_box", "cost_of_box", "boxes_bonus_per_yield"
        ]):
            # Recalculate using the same function as the forecast component
            forecasted_price = calculate_forecasted_price(
                estimation.variety, 
                estimation.start_date, 
                estimation.forecast_date, 
                estimation.city
            )
            
            # Update the estimation fields
            estimation.forecasted_price = forecasted_price
            estimation.revenue_per_acre = forecasted_price * estimation.yield_per_acre
            
            # Recalculate total costs
            total_costs = estimation.cost_per_acre + \
                         (estimation.harvest_cost_per_box * estimation.yield_per_acre) + \
                         (estimation.cost_of_box * estimation.yield_per_acre) + \
                         estimation.boxes_bonus_per_yield
            
            # Update revenue after costs
            estimation.revenue_after_costs = estimation.revenue_per_acre - total_costs
            
            # Update season
            estimation.season = determine_season_for_dashboard(estimation.forecast_date)
            
            # Recalculate revenue per box
            if estimation.yield_per_acre > 0:
                estimation.revenue_per_box = estimation.revenue_after_costs / estimation.yield_per_acre
        
        db.session.commit()
        
        return jsonify({
            "message": "Estimation updated successfully",
            "estimation": estimation.to_dict()
        }), 200
    except ValueError as e:
        return jsonify({"error": f"Invalid data: {str(e)}"}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Failed to update estimation: {str(e)}"}), 500


@app.route("/api/break_even_chart_data", methods=["GET"])
def get_break_even_chart_data():
    """
    Generate chart data specifically for break-even analysis using calculate_forecasted_price
    Accepts parameters as sent by the frontend
    """
    try:
        # Extract parameters from the request
        variety = request.args.get("variety")
        city = request.args.get("city")
        start_date_str = request.args.get("start_date")
        forecast_date_str = request.args.get("forecast_date")
        is_all_cities = request.args.get("is_all_cities") == "true"
        
        # Validate required parameters
        if not variety or not start_date_str:
            return jsonify({"error": "Missing required parameters"}), 400
            
        # Parse dates
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        
        # If forecast date is provided, use it; otherwise default to start_date + 1 year
        if forecast_date_str:
            forecast_date = datetime.strptime(forecast_date_str, "%Y-%m-%d")
        else:
            forecast_date = start_date + timedelta(days=365)
        
        # Handle "All Cities" case
        if is_all_cities or city == "All Cities":
            city = "All cities"
        
        # Initialize result structure
        result = {
            "labels": [],
            "datasets": []
        }
        
        # Define seasons and determine start season
        seasons = ["Winter", "Spring", "Summer", "Autumn"]
        
        month = start_date.month
        start_season = (
            "Spring" if 3 <= month <= 5 else
            "Summer" if 6 <= month <= 8 else
            "Autumn" if 9 <= month <= 11 else
            "Winter"
        )
        
        start_season_index = seasons.index(start_season)
        start_year = start_date.year
        
        # Generate labels for 1 year (4 seasons)
        labels = []
        for i in range(4):
            season_index = (start_season_index + i) % 4
            year = start_year + ((start_season_index + i) // 4)
            labels.append(f"{seasons[season_index]} {year}")
        
        result["labels"] = labels
        
        # Calculate prices for each season
        prices = []
        for label in labels:
            season, year = label.split()
            year = int(year)
            
            # Calculate the middle date of this season
            if season == "Spring":
                month = 4  # April
            elif season == "Summer":
                month = 7  # July
            elif season == "Autumn":
                month = 10  # October
            else:  # Winter
                month = 1  # January
                
            # Middle of the month
            day = 15
            
            # Create forecast date for this season
            season_date = datetime(year, month, day)
            
            # Use the calculate_forecasted_price function
            price = calculate_forecasted_price(variety, start_date, season_date, city)
            prices.append(round(price, 2))
        
        # Add dataset for the variety
        result["datasets"] = [{
            "label": f"{variety} Forecast Price",
            "data": prices,
            "borderColor": "#FF6384",
            "backgroundColor": "#FF6384",
            "borderWidth": 2,
            "pointRadius": 3,
            "pointHoverRadius": 5,
            "tension": 0.1,
            "fill": 'false',
            "borderDash": [5, 5]
        }]
        
        return jsonify(result)
        
    except Exception as e:
        app.logger.error(f"Break-even chart data error: {str(e)}")
        return jsonify({"error": str(e)}), 500
    

@app.route('/api/commodities', methods=['GET'])
# @jwt_required()
def get_commodities():
    """Get list of available commodities."""
    try:
        # Get unique commodity names from price data
        commodities = db.session.query(PriceData.commodity).distinct().all()
        commodity_list = [commodity[0] for commodity in commodities]
        
        # Standardize commodity names
        standardized_commodities = []
        for commodity in commodity_list:
            # Handle the specific case of Cubanelle/Cubanelles
            if commodity.lower() in ['cubanelle', 'cubanelles']:
                if 'Cubanelles' not in standardized_commodities:
                    standardized_commodities.append('Cubanelles')
            else:
                standardized_commodities.append(commodity)
        
        return jsonify(sorted(standardized_commodities))
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    



# A function to check whether the current route should be cached
def should_cache_route():
    return request.path not in app.config['CACHE_NO_CACHE_ROUTES']

# Update your caching logic in the relevant routes:




@app.route('/api/alert-entries-fresh', methods=['GET'])
@jwt_required()
def get_alert_settings_fresh():
    """Fetch all alert settings for the current user with guaranteed fresh data."""
    try:
        # Get current user ID from the JWT token
        current_user_id = get_jwt_identity()
        
        # Always bypass cache for this route
        print(f"Fetching fresh alert settings for user {current_user_id}")
        
        # Query the database directly to get the most recent data for this user
        settings = AlertSetting.query.filter_by(user_id=current_user_id).all()
        result = [alert.to_dict() for alert in settings]
        
        # Completely remove any cached data
        cache.delete(f'alert_settings:{current_user_id}')
        
        # Create response with aggressive no-cache headers
        response = jsonify(result)
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0, s-maxage=0, proxy-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        response.headers['X-Accel-Expires'] = '0'
        
        return response
    
    except Exception as e:
        print(f"Error fetching alert settings: {str(e)}")
        
        response = jsonify({
            "error": "Failed to retrieve alert settings",
            "details": str(e)
        })
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0, s-maxage=0, proxy-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        response.headers['X-Accel-Expires'] = '0'
        
        return response, 500


@app.route('/api/alert-settings', methods=['POST'])
@jwt_required()
def create_alert_setting():
    """Create a new alert setting for the current user."""
    # Get current user ID from the JWT token
    current_user_id = get_jwt_identity()
    
    # Log the incoming request details
    app.logger.info(f"Received alert setting creation request for user {current_user_id}")
    
    try:
        # Validate request data
        data = request.json
        if not data:
            response = jsonify({"error": "No data provided"})
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response, 400

        # Validate city (required field)
        city = data.get('city')
        if not city or not isinstance(city, str):
            response = jsonify({"error": "Valid city is required"})
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response, 400
        
        # Validate commodity (required field)
        commodity = data.get('commodity')
        if not commodity or not isinstance(commodity, str):
            response = jsonify({"error": "Valid commodity is required"})
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response, 400
        
        # Validate and sanitize threshold
        try:
            threshold = float(data.get('threshold', 5.0))
            # No bounds check on threshold so it can be negative for price decreases
        except (TypeError, ValueError):
            response = jsonify({"error": "Invalid threshold value"})
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response, 400
        
        # Validate is_active (use default if not provided)
        is_active = bool(data.get('isActive', True))
        
        # Check for duplicate alert setting for this user
        existing_alert = AlertSetting.query.filter_by(
            user_id=current_user_id,
            city=city,
            commodity=commodity,
            threshold=threshold
        ).first()
        
        if existing_alert:
            response = jsonify({
                "error": "You already have an alert for this commodity and city with the same threshold",
                "existing_alert_id": existing_alert.id
            })
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response, 409  # Conflict status code
        
        # Create new alert setting with user_id and city
        new_alert = AlertSetting(
            user_id=current_user_id,
            city=city.strip(),
            commodity=commodity.strip(),
            threshold=threshold,
            is_active=is_active
        )
        
        # Add the new alert to the session and commit
        db.session.add(new_alert)
        db.session.commit()
        
        # Update the user-specific cache
        settings = AlertSetting.query.filter_by(user_id=current_user_id).all()
        result = [alert.to_dict() for alert in settings]
        cache.set(f'alert_settings:{current_user_id}', result, timeout=300)
        
        # Log successful creation
        app.logger.info(f"Alert setting created for user {current_user_id}: City={city}, Commodity={commodity}, Threshold={threshold}")
        
        # Prepare response
        response = jsonify(new_alert.to_dict())
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response, 201
    
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error creating alert setting: {str(e)}")
        
        response = jsonify({
            "error": "An unexpected error occurred",
            "details": str(e)
        })
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response, 500  


@app.route('/api/alert-settings/<int:alert_id>', methods=['PATCH'])
@jwt_required()
def update_alert_setting(alert_id):
    """Update an existing alert setting."""
    current_user_id = get_jwt_identity()
    data = request.json
    
    try:
        # Find the alert setting for this user
        alert = AlertSetting.query.filter_by(
            id=alert_id, 
            user_id=current_user_id
        ).first()
        
        if not alert:
            return jsonify({"error": "Alert setting not found"}), 404
        
        # Update fields
        if 'threshold' in data:
            alert.threshold = float(data['threshold'])
        
        if 'isActive' in data:
            alert.is_active = bool(data['isActive'])
        
        alert.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify(alert.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/delete-alert-by-id', methods=['POST'])
@jwt_required()
def delete_alert_by_id():
    """Delete an alert by ID for the current user."""
    current_user_id = get_jwt_identity()
    try:
        data = request.json
        alert_id = data.get('id')
        
        if not alert_id:
            response = jsonify({"error": "Alert ID is required"})
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response, 400
        
        # Find the alert by ID and user_id
        alert = AlertSetting.query.filter_by(id=alert_id, user_id=current_user_id).first()
        
        if not alert:
            response = jsonify({"error": "No alert found with this ID for your account"})
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response, 404
        
        # Delete the alert
        db.session.delete(alert)
        db.session.commit()

        # Update user's cache
        cache.delete(f'alert_settings:{current_user_id}')
        
        response = jsonify({
            "success": True, 
            "message": f"Alert with ID {alert_id} deleted successfully",
            "deleted_alert": {
                "id": alert.id,
                "commodity": alert.commodity,
                "threshold": alert.threshold
            }
        })
        
        # Add no-cache headers
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
    
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/notifications', methods=['GET'])
@jwt_required()
def get_notifications():
    """Get all notifications for the current user."""
    current_user_id = get_jwt_identity()
    try:
        notifications = Notification.query.filter_by(user_id=current_user_id).order_by(
            Notification.created_at.desc()
        ).all()
        return jsonify([notification.to_dict() for notification in notifications])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/notifications/unread-count', methods=['GET'])
@jwt_required()
def get_unread_count():
    """Get count of unread notifications for the current user."""
    current_user_id = get_jwt_identity()
    try:
        # Count unread notifications for this user
        count = Notification.query.filter_by(
            user_id=current_user_id,
            read=False
        ).count()
        
        return jsonify({"count": count})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    

@app.route('/api/notifications/<int:notification_id>', methods=['PATCH'])
@jwt_required()  # Add this decorator
def update_notification(notification_id):
    """Update a notification (mark as read)."""
    current_user_id = get_jwt_identity()  # Get the current user
    data = request.json
    
    try:
        # Find the notification for this user
        notification = Notification.query.filter_by(
            id=notification_id,
            user_id=current_user_id  # Only allow users to update their own notifications
        ).first()
        
        if not notification:
            return jsonify({"error": "Notification not found"}), 404
        
        # Update read status
        if 'read' in data:
            notification.read = bool(data['read'])
        
        db.session.commit()
        
        return jsonify(notification.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@app.route('/api/notifications/<int:notification_id>', methods=['DELETE'])
@jwt_required()  # Add this decorator
def delete_notification(notification_id):
    """Delete a notification."""
    current_user_id = get_jwt_identity()  # Get the current user
    
    try:
        # Find the notification for this user
        notification = Notification.query.filter_by(
            id=notification_id,
            user_id=current_user_id  # Only allow users to delete their own notifications
        ).first()
        
        if not notification:
            return jsonify({"error": "Notification not found"}), 404
        
        db.session.delete(notification)
        db.session.commit()
        
        return jsonify({"message": "Notification deleted successfully"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@app.route('/api/notifications/mark-all-read', methods=['POST'])
@jwt_required()  # Add this decorator
def mark_all_read():
    """Mark all notifications as read for the current user."""
    current_user_id = get_jwt_identity()  # Get the current user
    
    try:
        # Update all unread notifications for this user only
        notifications = Notification.query.filter_by(
            user_id=current_user_id,
            read=False
        ).all()
        
        for notification in notifications:
            notification.read = True
        
        db.session.commit()
        
        return jsonify({"message": f"{len(notifications)} notifications marked as read"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500



import calendar


@app.route("/api/monthly-average-prices", methods=["GET"])
def get_monthly_average_prices():
    """
    Retrieve detailed monthly average prices for a given commodity
    """
    try:
        # Extract query parameters
        commodity = request.args.get('commodity')
        start_month = int(request.args.get('start_month', 1))  # Default January
        end_month = int(request.args.get('end_month', 12))    # Default December
        start_year = int(request.args.get('start_year'))
        end_year = int(request.args.get('end_year'))
        
        # Validate required parameters
        if not commodity or not start_year or not end_year:
            return jsonify({"error": "Commodity, start year, and end year are required"}), 400
        
        # Prepare months list based on selected range
        months_to_analyze = list(range(start_month, end_month + 1))
        
        # Prepare results
        monthly_analysis = []
        
        # For SQLite: Properly calculate day ranges for each month
        for month in months_to_analyze:
            # Query for this specific month across years
            month_data = []
            
            # For each year in the range
            for year in range(start_year, end_year + 1):
                # Determine if it's a leap year
                is_leap_year = (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)
                
                # Define days in each month for the current year
                days_in_month = [31, 29 if is_leap_year else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
                
                # Calculate cumulative days at the beginning of each month
                cumulative_days = [0]  # Start with 0
                running_sum = 0
                for days in days_in_month:
                    running_sum += days
                    cumulative_days.append(running_sum)
                
                # Calculate day range for the specified month (1-indexed)
                month_start_day = cumulative_days[month-1] + 1
                month_end_day = cumulative_days[month]
                
                # Query for this month/year combination
                year_month_data = (
                    db.session.query(
                        func.avg(PriceData.price).label('avg_price'),
                        func.min(PriceData.price).label('min_price'),
                        func.max(PriceData.price).label('max_price')
                    )
                    .filter(
                        PriceData.commodity == commodity,
                        PriceData.year == year,
                        PriceData.day >= month_start_day,
                        PriceData.day <= month_end_day,
                        PriceData.source == 'ProduceIQ'
                    )
                    .first()
                )
                
                # Only add if we have valid data
                if year_month_data and year_month_data.avg_price is not None:
                    month_data.append({
                        'year': year,
                        'avg_price': round(float(year_month_data.avg_price), 2),
                        'min_price': round(float(year_month_data.min_price), 2),
                        'max_price': round(float(year_month_data.max_price), 2)
                    })
            
            # Process the data for this month
            if month_data:
                # Calculate overall statistics for this month across years
                all_prices = [entry['avg_price'] for entry in month_data]
                all_min_prices = [entry['min_price'] for entry in month_data]
                all_max_prices = [entry['max_price'] for entry in month_data]
                
                monthly_analysis.append({
                    'month': calendar.month_name[month],
                    'avg_price': round(sum(all_prices) / len(all_prices), 2),
                    'min_price': round(min(all_min_prices), 2),
                    'max_price': round(max(all_max_prices), 2),
                    'years_data': month_data
                })
            else:
                # If no data for this month, add a default entry
                monthly_analysis.append({
                    'month': calendar.month_name[month],
                    'avg_price': 0,
                    'min_price': 0,
                    'max_price': 0,
                    'years_data': []
                })
        
        return jsonify({
            "monthly_prices": monthly_analysis,
            "commodity": commodity,
            "start_month": start_month,
            "end_month": end_month,
            "start_year": start_year,
            "end_year": end_year
        }), 200
    
    except Exception as e:
        app.logger.error(f"Error in monthly average prices: {str(e)}")
        return jsonify({"error": str(e)}), 500



# TEST ROUTE
@app.route("/api/trigger_usda_fetch", methods=["GET"])
def trigger_usda_fetch():
    fetch_usda_daily_data()
    fetch_shipping_point_data()
    fetch_daily_data()
    return "USDA Data Fetch Triggered"


# Function to check file extension
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# Route for uploading historical data automatically on route trigger
@app.route("/api/upload_historical", methods=["GET"])
def upload_historical():
    try:
        # Log when the upload process starts
        logging.info("Starting upload of historical data from 'data/' directory")

        # Loop through all CSV files in the 'data/' directory
        for csv_file in os.listdir(CSV_DIRECTORY):
            # Ensure we're only processing CSV files
            if csv_file.endswith(".csv"):
                commodity = os.path.splitext(csv_file)[
                    0
                ]  # Commodity name is the filename without extension
                file_path = os.path.join(CSV_DIRECTORY, csv_file)

                logging.info(f"Processing file: {csv_file}")

                # Open the CSV file and insert data into the PriceData table

                with open(file_path, newline="", encoding="utf-8") as csvfile:
                    reader = csv.DictReader(csvfile)

                    for row in reader:
                        city_name = row["CityName"]
                        year = int(row["Year"])
                        day = int(row["Day"])
                        price_str = row["Price"]

                        # Check if the price is empty or not a valid float
                        if price_str.strip() == "":
                            logging.warning(
                                f"Skipping row with empty price for {city_name} on day {day} in year {year}"
                            )
                            continue

                        try:
                            price = float(price_str)
                        except ValueError:
                            logging.error(
                                f"Invalid price value: {price_str} for {city_name} on day {day} in year {year}"
                            )
                            continue

                        source = "ProduceIQ"

                        # Correct the season logic using proper year-day conversion
                        try:
                            report_date = datetime.strptime(
                                f"{year}-{day}", "%Y-%j"
                            )  # %Y-%j converts year and day of year
                            season = determine_season(
                                report_date.strftime("%m/%d/%Y")
                            )  # Convert date to month/day/year format
                        except Exception as e:
                            logging.error(
                                f"Error converting year-day to date for {year}-{day}: {str(e)}"
                            )
                            continue

                        # Insert the data into the PriceData table
                        new_price_data = PriceData(
                            city_name=city_name,
                            commodity=commodity,
                            year=year,
                            day=day,
                            price=price,
                            source=source,
                            season=season,
                        )
                        db.session.add(new_price_data)

                    # Commit after processing each file
                    db.session.commit()
                    logging.info(
                        f"Data for {commodity} uploaded successfully from {csv_file}"
                    )

        logging.info("Historical data upload process completed")
        return "Historical data upload process completed", 200

    except Exception as e:
        logging.error(f"Error during upload: {str(e)}")
        return f"Error during upload: {str(e)}", 500




from sqlalchemy.sql import text  # Import `text` from SQLAlchemy
        
import plotly.graph_objects as go
from flask import jsonify


# voilin plot for terminal data

@app.route("/api/terminal_price_violin", methods=["GET"])
def terminal_price_violin():
    try:
        app.logger.info("Generating terminal violin plots for USDA and ProduceIQ without downsampling...")

        # Get the time frame from query parameters (default to '7d')
        time_frame = request.args.get("timeFrame", "7d")

        # Map timeFrame to PostgreSQL-compatible intervals
        time_intervals = {
            "3d": "3 days",
            "7d": "7 days",
            "1m": "1 month",
            "3m": "3 months",
            "1y": "1 year",
        }

        # Get the corresponding PostgreSQL interval for the time frame
        postgres_interval = time_intervals.get(time_frame.lower(), "7 days")

        # PostgreSQL-compatible query to fetch data filtered by source and time range
        query = text(f"""
            SELECT commodity, price, source
            FROM price_data
            WHERE source IN ('USDA', 'ProduceIQ')
            AND make_date(year, 1, 1) + (day - 1) * INTERVAL '1 day' >= NOW() - INTERVAL '{postgres_interval}'
            AND price > 2
        """)
        result = db.session.execute(query).fetchall()

        # Group data by source and commodity
        grouped_data = {"USDA": {}, "ProduceIQ": {}}
        for row in result:
            commodity, price, source = row
            if commodity not in grouped_data[source]:
                grouped_data[source][commodity] = []
            grouped_data[source][commodity].append(price)

        # Create violin traces for USDA
        usda_traces = [
            go.Violin(
                y=prices,
                name=commodity,
                box_visible=True,
                meanline_visible=True,
                marker_color='blue'  # Color for USDA
            )
            for commodity, prices in grouped_data["USDA"].items()
        ]

        # Create violin traces for ProduceIQ
        produceiq_traces = [
            go.Violin(
                y=prices,
                name=commodity,
                box_visible=True,
                meanline_visible=True,
                marker_color='green'  # Color for ProduceIQ
            )
            for commodity, prices in grouped_data["ProduceIQ"].items()
        ]

        # Layout for USDA
        usda_layout = {
            "title": {"text": "USDA Terminal Price Distribution by Commodity", "font": {"size": 16, "weight": "bold"}},
            "xaxis": {"title": {"text": "Commodity", "font": {"size": 14, "weight": "bold"}}, "automargin": True},
            "yaxis": {"title": {"text": "Price", "font": {"size": 14, "weight": "bold"}}, "zeroline": True, "automargin": True},
            "height": 500,
            "width": 700,
            "showlegend": False,
            "plot_bgcolor": "#f0f8ff",
            "paper_bgcolor": "white",
        }

        # Layout for ProduceIQ
        produceiq_layout = {
            "title": {"text": "ProduceIQ Terminal Price Distribution by Commodity", "font": {"size": 16, "weight": "bold"}},
            "xaxis": {"title": {"text": "Commodity", "font": {"size": 14, "weight": "bold"}}, "automargin": True},
            "yaxis": {"title": {"text": "Price", "font": {"size": 14, "weight": "bold"}}, "automargin": True},
            "height": 500,
            "width": 700,
            "showlegend": False,
            "plot_bgcolor": "#f0f8ff",
            "paper_bgcolor": "white",
        }

        # Convert traces to JSON serializable format
        usda_traces_json = [trace.to_plotly_json() for trace in usda_traces]
        produceiq_traces_json = [trace.to_plotly_json() for trace in produceiq_traces]

        # Return JSON response containing separate charts for USDA and ProduceIQ
        return jsonify({
            "usda": {"data": usda_traces_json, "layout": usda_layout},
            "produceiq": {"data": produceiq_traces_json, "layout": produceiq_layout}
        }), 200

    except Exception as e:
        app.logger.error(f"Error generating terminal violin plots: {str(e)}")
        return jsonify({"error": "Failed to generate terminal violin plots"}), 500




@app.route("/api/shipping_price_violin", methods=["GET"])
def shipping_price_violin():
    try:
        app.logger.info("Fetching data for shipping violin plot...")

        # Get the time frame from query parameters (default to '7d')
        time_frame = request.args.get("timeFrame", "7d")

        # Map timeFrame to PostgreSQL-compatible interval
        time_intervals = {
            "3d": "'3 days'",
            "7d": "'7 days'",
            "1m": "'1 month'",
            "3m": "'3 months'",
            "1y": "'1 year'",
        }

        # Get the corresponding PostgreSQL interval for the time frame
        postgres_interval = time_intervals.get(time_frame.lower(), "'7 days'")

        # SQL query to fetch data filtered by source and time range
        query = text(f"""
    SELECT commodity, price
    FROM shipping_price_data
    WHERE source = 'ProduceIQ'
      AND TO_DATE(year || '-01-01', 'YYYY-MM-DD') + (day - 1) * interval '1 day' >= NOW() - INTERVAL {postgres_interval}
      AND price > 1  
        """)

        result = db.session.execute(query).fetchall()

        # Group data by commodity
        data = {}
        for row in result:
            commodity = row[0]  # Commodity (varietyName)
            price = row[1]  # Price
            if commodity not in data:
                data[commodity] = []
            data[commodity].append(price)

        # Create violin traces for each commodity
        traces = []
        for commodity, prices in data.items():
            traces.append(
                go.Violin(
                    y=prices,
                    name=commodity,
                    box_visible=True,
                    meanline_visible=True,
                    marker_color='green',  # Custom color

                )
            )

        # Create the layout for the chart
        layout = {
            "title": {"text": "Shipping Price Distribution by Commodity", "font": {"size": 16, "weight": "bold"}},
            "xaxis": {"title": {"text": "Commodity", "font": {"size": 14, "weight": "bold"}}, "automargin": True},
            "yaxis": {"title": {"text": "Shipping Price", "font": {"size": 14, "weight": "bold"}}, "automargin": True},
            "height": 500,
            "width": 700,
            "showlegend": False,
            "plot_bgcolor": "#f0f8ff",
            "paper_bgcolor": "white",
        }

        # Return the chart data and layout as JSON
        return jsonify({"data": [trace.to_plotly_json() for trace in traces], "layout": layout}), 200

    except Exception as e:
        app.logger.error(f"Error generating shipping violin plot: {str(e)}")
        return jsonify({"error": "Failed to generate shipping violin plot"}), 500




# terminal empricial probability chart fetch

from sqlalchemy import func

@app.route('/api/terminal_empricial_probability', methods=['GET'])
def get_terminal_empricial_probability():
    try:
        # Get the time frame from query parameters (default to '7d')
        time_frame = request.args.get('timeFrame', '7d')

        # Map timeFrame to PostgreSQL-compatible intervals
        time_intervals = {
            "3d": "3 days",
            "7d": "7 days",
            "1m": "1 month",
            "3m": "3 months",
            "1y": "1 year",
        }

        # Get the corresponding PostgreSQL interval for the time frame
        postgres_interval = time_intervals.get(time_frame.lower(), '7 days')

        # Query to get mean and standard deviation directly from the database
        query = text(f"""
            SELECT commodity, price, source
            FROM price_data
            WHERE source = 'USDA'
            AND make_date(year, 1, 1) + (day - 1) * INTERVAL '1 day' >= NOW() - INTERVAL '{postgres_interval}'
            AND price > 2
        """)
        result = db.session.execute(query).fetchall()

        if not result:
            return jsonify([])

        # Group data by commodity
        grouped_data = {}
        for row in result:
            commodity, price, source = row
            if commodity not in grouped_data:
                grouped_data[commodity] = []
            grouped_data[commodity].append(price)

        # Create violin traces
        charts = []
        for commodity, prices in grouped_data.items():
            hist, bin_edges = np.histogram(prices, bins=50)
            histogram_trace = go.Bar(
                x=bin_edges[:-1].tolist(), 
                y=hist.tolist(),
                name=f"{commodity} Histogram",
                marker_color="#636EFA"
            )
            mean = np.mean(prices)
            std_dev = np.std(prices)
            mean_trace = go.Scatter(
                x=[mean, mean], 
                y=[0, max(hist)], 
                mode="lines", 
                line=dict(color="red", dash="dash"), 
                name=f"{commodity} Mean"
            )
            std_dev_trace = go.Scatter(
                x=[mean - std_dev, mean + std_dev], 
                y=[0, 0], 
                mode="markers", 
                marker=dict(color="blue", size=8, symbol="cross"), 
                name=f"{commodity} Std Dev"
            )

            layout = {
                "title": f"{commodity} Price Distribution",
                "xaxis": {"title": "Price", "automargin": True},
                "yaxis": {"title": "Frequency", "automargin": True},
                "height": 400,
                "width": 370,
                "showlegend": False,
                "plot_bgcolor": "#f0f8ff",
                "paper_bgcolor": "white",
            }

            charts.append({
                "commodity": commodity,
                "data": [histogram_trace.to_plotly_json(), mean_trace.to_plotly_json(), std_dev_trace.to_plotly_json()],
                "layout": layout,
            })

        return jsonify(charts), 200

    except Exception as e:
        error_message = str(e)
        print("Error Traceback:", error_message)
        return jsonify({"error": error_message}), 500







# shipping empricial probability chart fetch ProduceIQ
@app.route('/api/shipping_empricial_probability', methods=['GET'])
def get_shipping_empricial_probability():
    try:
        # Get the time frame from query parameters (default to '7d')
        time_frame = request.args.get('timeFrame', '7d')

        # Map timeFrame to PostgreSQL-compatible intervals
        time_intervals = {
            "3d": "3 days",
            "7d": "7 days",
            "1m": "1 month",
            "3m": "3 months",
            "1y": "1 year",
        }

        # Get the corresponding PostgreSQL interval for the time frame
        postgres_interval = time_intervals.get(time_frame.lower(), '7 days')

        # Query to get shipping data from ProduceIQ within the given time frame
        query = text(f"""
            SELECT commodity, price
            FROM shipping_price_data
            WHERE source = 'ProduceIQ'
            AND make_date(year, 1, 1) + (day - 1) * INTERVAL '1 day' >= NOW() - INTERVAL '{postgres_interval}'
            AND price > 2
        """)
        result = db.session.execute(query).fetchall()

        if not result:
            return jsonify([])

        # Group data by commodity
        grouped_data = {}
        for row in result:
            commodity, price = row
            if commodity not in grouped_data:
                grouped_data[commodity] = []
            grouped_data[commodity].append(price)

        # Create violin traces
        charts = []
        for commodity, prices in grouped_data.items():
            hist, bin_edges = np.histogram(prices, bins=50)
            histogram_trace = go.Bar(
                x=bin_edges[:-1].tolist(),
                y=hist.tolist(),
                name=f"{commodity} Histogram",
                marker_color="green"
            )
            mean = np.mean(prices)
            std_dev = np.std(prices)
            mean_trace = go.Scatter(
                x=[mean, mean], 
                y=[0, max(hist)], 
                mode="lines", 
                line=dict(color="red", dash="dash"), 
                name=f"{commodity} Mean"
            )
            std_dev_trace = go.Scatter(
                x=[mean - std_dev, mean + std_dev], 
                y=[0, 0], 
                mode="markers", 
                marker=dict(color="green", size=8, symbol="cross"), 
                name=f"{commodity} Std Dev"
            )

            layout = {
                "title": f"{commodity} Price Distribution",
                "xaxis": {"title": "Price", "automargin": True},
                "yaxis": {"title": "Frequency", "automargin": True},
                "height": 400,
                "width": 370,
                "showlegend": False,
                "plot_bgcolor": "#f0f8ff",
                "paper_bgcolor": "white",
            }

            charts.append({
                "commodity": commodity,
                "data": [histogram_trace.to_plotly_json(), mean_trace.to_plotly_json(), std_dev_trace.to_plotly_json()],
                "layout": layout,
            })

        return jsonify(charts), 200

    except Exception as e:
        error_message = str(e)
        print("Error Traceback:", error_message)
        return jsonify({"error": error_message}), 500






@app.route("/api/terminal_correlation", methods=["GET"])
def get_terminal_correlation():
    try:
        # Query data with proper filtering
        source = request.args.get('source', 'ProduceIQ')  # Default to USDA
        query = text("""
            SELECT 
                commodity,
                price,
                DATE(CONCAT(year, '-01-01')::date + (day - 1) * INTERVAL '1 day') as date
            FROM price_data
            WHERE source = :source
            ORDER BY date, commodity
        """)
        
        result = db.session.execute(query, {'source': source}).fetchall()
        print(f"Query result length: {len(result)}")  # Debugging log

        if not result:
            return jsonify({"error": "No data found"}), 404

        # Create DataFrame and process data
        df = pd.DataFrame(result, columns=['commodity', 'price', 'date'])
        print(f"Initial DataFrame shape: {df.shape}")  # Debugging log

        # Ensure 'date' column is treated as datetime
        df['date'] = pd.to_datetime(df['date'])

        # Handle duplicate (date, commodity) entries by aggregating
        df = df.groupby(['date', 'commodity'], as_index=False)['price'].mean()
        print(f"DataFrame shape after grouping: {df.shape}")  # Debugging log

        # Create pivot table
        pivot_df = df.pivot(index='date', columns='commodity', values='price')
        print(f"Pivot table shape: {pivot_df.shape}")  # Debugging log

        del df
        import gc
        gc.collect()

        # Forward fill missing values before calculating percentage change
        pivot_df = pivot_df.ffill(limit=3)  # Forward fill up to 3 missing values
        returns = pivot_df.pct_change()  # Calculate percentage change
        print(f"Returns DataFrame shape before filtering: {returns.shape}")  # Debugging log

        # Remove columns with too many missing values
        min_valid_ratio = 0.3  # At least 30% valid data
        valid_columns = returns.columns[returns.count() > len(returns) * min_valid_ratio]
        returns = returns[valid_columns]
        print(f"Returns DataFrame shape after filtering: {returns.shape}")  # Debugging log

        del pivot_df
        gc.collect()

        # Calculate correlation matrix
        correlation_matrix = returns.corr(method='pearson')
        correlation_matrix.fillna(0, inplace=True)
        print(f"Correlation matrix shape: {correlation_matrix.shape}")  # Debugging log

        # Prepare data for plotting
        labels = correlation_matrix.columns.tolist()
        z_values = correlation_matrix.values.tolist()

        # Reverse the order of y-axis labels and rows in z_values
        reversed_labels = labels[::-1]  # Reverse the y-axis labels
        reversed_z_values = z_values[::-1]  # Reverse the rows of the correlation matrix

        # Create annotations dynamically for reversed labels and z_values
        annotations = [
            {
                "x": labels[col],
                "y": reversed_labels[row],  # Adjust for reversed y-axis
                "text": f"{reversed_z_values[row][col]:.2f}",  # Use reversed data
                "showarrow": False,
                "font": {
                    "color": "white",
                    "size": 10,
                },
            }
            for row in range(len(reversed_labels))
            for col in range(len(labels))
        ]

        # Prepare Plotly chart
        chart = {
            "data": [
                {
                    "z": reversed_z_values,  # Use reversed rows for heatmap
                    "x": labels,
                    "y": reversed_labels,  # Use reversed y-axis labels
                    "type": "heatmap",
                    "colorscale": "CoolWarm",
                    "showscale": True,
                    "text": [[f"{val:.2f}" for val in row] for row in reversed_z_values],
                    "hoverinfo": "text",
                }
            ],
            "layout": {
                "title": "Correlation Matrix of Terminal Market Prices",
                "xaxis": {
                    "title": "Commodities",
                    "tickangle": -45,
                    "automargin": True,
                },
                "yaxis": {
                    "title": "Commodities",
                    "automargin": True,
                },
                "height": 600,
                "width": 600,
                "annotations": annotations,
            },
        }

        # Add summary statistics
        summary_stats = {
            "total_records": len(returns),  # Updated to use 'returns'
            "unique_commodities": len(labels),
            "date_range": {
                "start": returns.index.min().strftime('%Y-%m-%d'),
                "end": returns.index.max().strftime('%Y-%m-%d'),
            },
            "average_correlation": float(correlation_matrix.mean().mean()),
        }

        print(f"Final chart labels: {len(labels)}, annotations: {len(annotations)}")  # Debugging log

        return jsonify({"chart": chart, "stats": summary_stats}), 200

    except Exception as e:
        app.logger.error(f"Error generating terminal correlation chart: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Internal Server Error", "details": str(e)}), 500








@app.route("/api/shipping_correlation", methods=["GET"])
def get_shipping_correlation():
    try:
        # Query data with proper filtering
        source = request.args.get('source', 'ProduceIQ')  # Default to USDA
        query = text("""
            SELECT 
                commodity,
                price,
                DATE(CONCAT(year, '-01-01')::date + (day - 1) * INTERVAL '1 day') as date
            FROM shipping_price_data
            WHERE source = :source
            ORDER BY date, commodity
        """)
        
        result = db.session.execute(query, {'source': source}).fetchall()
        print(f"Query result length: {len(result)}")  # Debugging log

        if not result:
            return jsonify({"error": "No data found"}), 404

        # Create DataFrame and process data
        df = pd.DataFrame(result, columns=['commodity', 'price', 'date'])
        print(f"Initial DataFrame shape: {df.shape}")  # Debugging log

        # Ensure 'date' column is treated as datetime
        df['date'] = pd.to_datetime(df['date'])

        # Handle duplicate (date, commodity) entries by aggregating
        df = df.groupby(['date', 'commodity'], as_index=False)['price'].mean()
        print(f"DataFrame shape after grouping: {df.shape}")  # Debugging log

        # Create pivot table
        pivot_df = df.pivot(index='date', columns='commodity', values='price')
        print(f"Pivot table shape: {pivot_df.shape}")  # Debugging log

        del df
        import gc
        gc.collect()

        # Calculate returns and handle missing values
        pivot_df = pivot_df.ffill(limit=3)  # Forward fill missing values up to 3 days
        returns = pivot_df.pct_change()  # Calculate percentage change
        print(f"Returns DataFrame shape before filtering: {returns.shape}")  # Debugging log

        # Remove columns with too many missing values
        min_valid_ratio = 0.3  # At least 30% valid data
        valid_columns = returns.columns[returns.count() > len(returns) * min_valid_ratio]
        returns = returns[valid_columns]
        print(f"Returns DataFrame shape after filtering: {returns.shape}")  # Debugging log

        del pivot_df
        gc.collect()

        # Calculate correlation matrix
        correlation_matrix = returns.corr(method='pearson')
        correlation_matrix.fillna(0, inplace=True)
        print(f"Correlation matrix shape: {correlation_matrix.shape}")  # Debugging log

        # Prepare data for plotting
        labels = correlation_matrix.columns.tolist()
        z_values = correlation_matrix.values.tolist()

        # Reverse the order of y-axis labels and rows in z_values
        reversed_labels = labels[::-1]  # Reverse the y-axis labels
        reversed_z_values = z_values[::-1]  # Reverse the rows of the correlation matrix

        # Create annotations dynamically for reversed labels and z_values
        annotations = [
            {
                "x": labels[col],
                "y": reversed_labels[row],  # Adjust for reversed y-axis
                "text": f"{reversed_z_values[row][col]:.2f}",  # Use reversed data
                "showarrow": False,
                "font": {
                    "color": "white",
                    "size": 10,
                },
            }
            for row in range(len(reversed_labels))
            for col in range(len(labels))
        ]

        # Prepare Plotly chart
        chart = {
            "data": [
                {
                    "z": reversed_z_values,  # Use reversed rows for heatmap
                    "x": labels,
                    "y": reversed_labels,  # Use reversed y-axis labels
                    "type": "heatmap",
                    "colorscale": "CoolWarm",
                    "showscale": True,
                    "text": [[f"{val:.2f}" for val in row] for row in reversed_z_values],
                    "hoverinfo": "text",
                }
            ],
            "layout": {
                "title": "Correlation Matrix of Shipping Prices",
                "xaxis": {
                    "title": "Commodities",
                    "tickangle": -45,
                    "automargin": True,
                },
                "yaxis": {
                    "title": "Commodities",
                    "automargin": True,
                },
                "height": 600,
                "width": 600,
                "annotations": annotations,
            },
        }

        # Add summary statistics
        summary_stats = {
            "total_records": len(returns),  # Updated to use 'returns'
            "unique_commodities": len(labels),
            "date_range": {
                "start": returns.index.min().strftime('%Y-%m-%d'),
                "end": returns.index.max().strftime('%Y-%m-%d'),
            },
            "average_correlation": float(correlation_matrix.mean().mean()),
        }

        print(f"Final chart labels: {len(labels)}, annotations: {len(annotations)}")  # Debugging log

        return jsonify({"chart": chart, "stats": summary_stats}), 200

    except Exception as e:
        app.logger.error(f"Error generating shipping correlation chart: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Internal Server Error", "details": str(e)}), 500





# terminal scatterplot for usda data
@app.route("/api/terminal_scatterplot_matrix", methods=["POST"])
def get_terminal_scatterplot_matrix():
    try:
        # Parse input from the frontend
        data = request.get_json()
        commodity_x = data.get("commodity_x")
        commodity_y = data.get("commodity_y")
        source = data.get("source", "ProduceIQ")  # Default to 'USDA'

        if not commodity_x or not commodity_y:
            return jsonify({"error": "Both commodities must be provided"}), 400

        # Query data for the selected commodities from PriceData where source is 'USDA'
        result = db.session.query(PriceData.commodity, PriceData.price).filter(
            PriceData.commodity.in_([commodity_x, commodity_y]),
            PriceData.source == source
        ).all()

        if not result:
            return jsonify({"error": "No data found for the selected commodities"}), 404

        # Create a DataFrame
        df = pd.DataFrame(result, columns=["commodity", "price"]).dropna()

        if df.empty:
            return jsonify({"error": "No valid data available"}), 404

        # Separate data for each commodity
        x_data = df[df['commodity'] == commodity_x]['price'].tolist()
        y_data = df[df['commodity'] == commodity_y]['price'].tolist()

        # Get the minimum length to ensure equal pairs
        min_length = min(len(x_data), len(y_data))

        if min_length == 0:
            scatter_plot = {
                "data": [],
                "layout": {
                    "title": "No Data Available",
                    "xaxis": {"title": f"{commodity_x} Prices"},
                    "yaxis": {"title": f"{commodity_y} Prices"},
                    "annotations": [
                        {
                            "text": "No valid data available for the selected commodities.",
                            "xref": "paper",
                            "yref": "paper",
                            "showarrow": False,
                            "font": {"size": 14},
                        }
                    ],
                    "height": 600,
                    "width": 600,
                },
            }
            return jsonify(scatter_plot), 200

        # Generate scatter plot JSON
        scatter_plot = {
            "data": [
                {
                    "x": x_data[:min_length],
                    "y": y_data[:min_length],
                    "mode": "markers",
                    "marker": {"size": 5, "opacity": 0.8, "color": "#33b1a7"},
                    "type": "scatter",
                }
            ],
            "layout": {
                "title": f"Scatter Plot: {commodity_x} vs {commodity_y}",
                "xaxis": {"title": f"{commodity_x} Prices"},
                "yaxis": {"title": f"{commodity_y} Prices"},
                "height": 600,
                "width": 600,
                "font": {"family": "Arial"},
            },
        }

        # Debug logs
        app.logger.info(f"{commodity_x} data (first 5): {x_data[:5]}")
        app.logger.info(f"{commodity_y} data (first 5): {y_data[:5]}")
        app.logger.info(f"Total points being plotted: {min_length}")

        return jsonify(scatter_plot), 200

    except Exception as e:
        app.logger.error(f"Error generating terminal scatterplot: {str(e)}")
        return jsonify({"error": "Internal Server Error", "details": str(e)}), 500






# shipping scatterplot for ProduceIQ data
@app.route("/api/shipping_scatterplot_matrix", methods=["POST"])
def get_shipping_scatterplot_matrix():
    try:
        # Parse input from the frontend
        data = request.get_json()
        commodity_x = data.get("commodity_x")
        commodity_y = data.get("commodity_y")
        source = data.get("source", "ProduceIQ")  # Default to 'USDA'

        if not commodity_x or not commodity_y:
            return jsonify({"error": "Both commodities must be provided"}), 400

        # Query data for the selected commodities from ShippingPriceData where source is 'ProduceIQ'
        result = db.session.query(ShippingPriceData.commodity, ShippingPriceData.price).filter(
            ShippingPriceData.commodity.in_([commodity_x, commodity_y]),
            ShippingPriceData.source == source
        ).all()

        if not result:
            return jsonify({"error": "No data found for the selected commodities"}), 404

        # Create a DataFrame
        df = pd.DataFrame(result, columns=["commodity", "price"]).dropna()

        if df.empty:
            return jsonify({"error": "No valid data available"}), 404

        # Separate data for each commodity
        x_data = df[df['commodity'] == commodity_x]['price'].tolist()
        y_data = df[df['commodity'] == commodity_y]['price'].tolist()

        # Get the minimum length to ensure equal pairs
        min_length = min(len(x_data), len(y_data))

        if min_length == 0:
            scatter_plot = {
                "data": [],
                "layout": {
                    "title": "No Data Available",
                    "xaxis": {"title": f"{commodity_x} Prices"},
                    "yaxis": {"title": f"{commodity_y} Prices"},
                    "annotations": [
                        {
                            "text": "No valid data available for the selected commodities.",
                            "xref": "paper",
                            "yref": "paper",
                            "showarrow": False,
                            "font": {"size": 14},
                        }
                    ],
                    "height": 600,
                    "width": 600,
                },
            }
            return jsonify(scatter_plot), 200

        # Generate scatter plot JSON
        scatter_plot = {
            "data": [
                {
                    "x": x_data[:min_length],
                    "y": y_data[:min_length],
                    "mode": "markers",
                    "marker": {"size": 5, "opacity": 0.8, "color": "#33b1a7"},
                    "type": "scatter",
                }
            ],
            "layout": {
                "title": f"Scatter Plot: {commodity_x} vs {commodity_y}",
                "xaxis": {"title": f"{commodity_x} Prices"},
                "yaxis": {"title": f"{commodity_y} Prices"},
                "height": 600,
                "width": 600,
                "font": {"family": "Arial"},
            },
        }

        # Debug logs
        app.logger.info(f"{commodity_x} data (first 5): {x_data[:5]}")
        app.logger.info(f"{commodity_y} data (first 5): {y_data[:5]}")
        app.logger.info(f"Total points being plotted: {min_length}")

        return jsonify(scatter_plot), 200

    except Exception as e:
        app.logger.error(f"Error generating shipping scatterplot: {str(e)}")
        return jsonify({"error": "Internal Server Error", "details": str(e)}), 500






# Rolling Correlations for Terminal prices

import plotly.express as px

def calculate_rolling_price_correlations(window, source_df, series1, series2):
    """
    Calculate rolling correlations between two price series.
    """
    try:
        # Sort index to ensure proper time-based calculations
        source_df = source_df.sort_index()
        
        # Forward fill missing values (up to 7 days)
        source_df = source_df.fillna(method='ffill', limit=7)
        
        # Calculate rolling correlation
        roll_corr = source_df[series1].rolling(
            window=window,
            min_periods=window // 2  # Allow for some missing data
        ).corr(source_df[series2])
        
        # Create result DataFrame
        result_df = pd.DataFrame({
            'date': roll_corr.index,
            'correlation': roll_corr.values
        })

        # Convert numpy values to Python native types
        result_df['correlation'] = result_df['correlation'].astype(float)
        result_df['date'] = result_df['date'].dt.strftime('%Y-%m-%d')

        # Clean up memory
        del roll_corr
        gc.collect()

        return result_df

    except KeyError as e:
        raise ValueError(f"KeyError: {e}. At least one of the selected varieties has no price data.")
    except Exception as e:
        raise ValueError(f"Unexpected error: {str(e)}")


def plot_rolling_price_correlations(roll_corr_df, series1, series2, window):
    """
    Create a plotly figure for rolling correlations.
    """
    try:
        # Convert date strings back to datetime for plotting
        roll_corr_df['date'] = pd.to_datetime(roll_corr_df['date'])
        
        # Create the plot
        fig_roll_corr = px.line(
            roll_corr_df,
            x='date',
            y='correlation',
            title=f"{window}-day rolling correlation between {series1} and {series2}"
        )

        fig_roll_corr.update_layout(
            title={
                "text": f"{window}-day rolling correlation between {series1} and {series2}",
                "x": 0.5,
                "y": 0.9,
                "xanchor": "center",
                "yanchor": "top"
            },
            xaxis_title="Date",
            yaxis_title="Correlation Coefficient",
            yaxis=dict(
                tickformat='.2f',
                range=[-1, 1]  # Correlation coefficient ranges from -1 to 1
            ),
            height=600,
            width=600,
            showlegend=False
        )

        # Add reference lines
        fig_roll_corr.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.5)
        fig_roll_corr.add_hline(y=1, line_dash="dot", line_color="gray", opacity=0.3)
        fig_roll_corr.add_hline(y=-1, line_dash="dot", line_color="gray", opacity=0.3)

        # Clean up memory
        del roll_corr_df
        gc.collect()

        return fig_roll_corr

    except Exception as e:
        raise ValueError(f"Error creating plot: {str(e)}")



# terminal correlations for usda data

@app.route("/api/terminal_rolling_correlations", methods=["POST"])
def terminal_rolling_correlations():
    try:
        # Parse input from the frontend
        data = request.get_json()
        series1 = data.get("series1")
        series2 = data.get("series2")
        window = int(data.get("window", 30))  # Default to 30 days
        source = data.get("source", "ProduceIQ")  # Default to USDA

        if not series1 or not series2:
            return jsonify({"error": "Both commodities must be provided"}), 400

        if window < 5:
            return jsonify({"error": "Window size must be at least 5 days"}), 400

        # Fetch data from the database, filtering by source 'USDA'
        result = db.session.query(
            PriceData.year,
            PriceData.day,
            PriceData.commodity,
            PriceData.price
        ).filter(
            PriceData.commodity.in_([series1, series2]),
            PriceData.price > 0,  # Ensure we only get valid prices
            PriceData.source == source  # Use selected source
        ).all()

        if not result:
            return jsonify({"error": "No data found for the selected commodities"}), 404

        # Create a DataFrame
        df = pd.DataFrame(result, columns=["year", "day", "commodity", "price"])

        # Generate a date column from year and day
        df["date"] = pd.to_datetime(
            df.apply(lambda row: f"{row.year}-{row.day}", axis=1),
            format="%Y-%j"
        )

        # Pivot the data for rolling correlation
        pivot_data = df.pivot_table(
            values="price",
            index="date",
            columns="commodity",
            aggfunc="mean"
        ).sort_index()

        # Clean up the main DataFrame after pivoting
        del df
        gc.collect()

        # Check for minimum data points
        min_required_points = window * 2
        if len(pivot_data) < min_required_points:
            del pivot_data
            gc.collect()
            return jsonify({
                "error": f"Insufficient data points. Need at least {min_required_points} days of data for {window}-day window"
            }), 400

        # Calculate rolling correlation
        roll_corr_df = calculate_rolling_price_correlations(
            window=window,
            source_df=pivot_data,
            series1=series1,
            series2=series2
        )

        # Clean up the pivot DataFrame after rolling correlation
        del pivot_data
        gc.collect()

        # Create the Plotly figure directly
        fig_roll_corr = px.line(
            roll_corr_df,
            x="date",
            y="correlation",
            title=f"{window}-day rolling correlation between {series1} and {series2}"
        )

        # Customize the layout
        fig_roll_corr.update_layout(
            title={
                "text": f"{window}-day rolling correlation between {series1} and {series2}",
                "x": 0.5,
                "y": 0.9,
                "xanchor": "center",
                "yanchor": "top"
            },
            xaxis_title="Date",
            yaxis_title="Correlation Coefficient",
            yaxis=dict(
                tickformat='.2f',
                range=[-1, 1]  # Correlation coefficient ranges from -1 to 1
            ),
            height=600,
            width=600,
            showlegend=False
        )

        # Add reference lines
        fig_roll_corr.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.5)
        fig_roll_corr.add_hline(y=1, line_dash="dot", line_color="gray", opacity=0.3)
        fig_roll_corr.add_hline(y=-1, line_dash="dot", line_color="gray", opacity=0.3)

        # Convert the Plotly figure to JSON, ensuring all values are serializable
        fig_json = fig_roll_corr.to_json()

        # Return the chart as JSON
        return app.response_class(
            response=fig_json,
            status=200,
            mimetype='application/json'
        )

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        app.logger.error(f"Error generating terminal rolling correlations: {str(e)}")
        return jsonify({"error": "Internal Server Error", "details": str(e)}), 500





# Rolling Correlations for Shipping prices

def calculate_rolling_price_correlations(window, source_df, series1, series2):
    """
    Calculate rolling correlations between two price series.
    """
    try:
        # Sort index to ensure proper time-based calculations
        source_df = source_df.sort_index()
        
        # Forward fill missing values (up to 7 days)
        source_df = source_df.ffill(limit=7)  # Use .ffill() instead of fillna(method='ffill')

        # Calculate rolling correlation
        roll_corr = source_df[series1].rolling(
            window=window,
            min_periods=window // 2  # Allow for some missing data
        ).corr(source_df[series2])
        
        # Create result DataFrame
        result_df = pd.DataFrame({
            'date': roll_corr.index,
            'correlation': roll_corr.values
        })

        # Convert numpy values to Python native types
        result_df['correlation'] = result_df['correlation'].astype(float)
        result_df['date'] = result_df['date'].dt.strftime('%Y-%m-%d')

        # Clean up memory
        del roll_corr
        gc.collect()

        return result_df

    except KeyError as e:
        raise ValueError(f"KeyError: {e}. At least one of the selected varieties has no price data.")
    except Exception as e:
        raise ValueError(f"Unexpected error: {str(e)}")



def plot_rolling_price_correlations(roll_corr_df, series1, series2, window):
    """
    Create a Plotly figure for rolling correlations and convert it to JSON.
    """
    try:
        # Convert date strings back to datetime for plotting
        roll_corr_df['date'] = pd.to_datetime(roll_corr_df['date'])
        
        # Create the Plotly figure
        fig_roll_corr = px.line(
            roll_corr_df,
            x='date',
            y='correlation',
            title=f"{window}-day Rolling Correlation between {series1} and {series2}",
            labels={'correlation': 'Correlation Coefficient', 'date': 'Date'}
        )

        # Update layout for better visualization
        fig_roll_corr.update_layout(
            title={
                "x": 0.5,
                "y": 0.9,
                "xanchor": "center",
                "yanchor": "top"
            },
            xaxis=dict(title="Date"),
            yaxis=dict(
                title="Correlation Coefficient",
                tickformat='.2f',
                range=[-1, 1]  # Correlation coefficient ranges from -1 to 1
            ),
            height=600,
            width=600,
            showlegend=False
        )

        # Add reference lines
        fig_roll_corr.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.5)
        fig_roll_corr.add_hline(y=1, line_dash="dot", line_color="gray", opacity=0.3)
        fig_roll_corr.add_hline(y=-1, line_dash="dot", line_color="gray", opacity=0.3)

        # Return the Plotly figure as JSON
        return fig_roll_corr.to_json()

    except Exception as e:
        raise ValueError(f"Error creating plot: {str(e)}")



# rolling correlations for shipping produceiq
@app.route("/api/shipping_rolling_correlations", methods=["POST"])
def shipping_rolling_correlations():
    """
    Endpoint to calculate and return rolling correlations chart for shipping price data.
    """
    try:
        # Parse input from the frontend
        data = request.get_json()
        series1 = data.get("series1")
        series2 = data.get("series2")
        window = int(data.get("window", 30))  # Default to 30 days
        source = data.get("source", "ProduceIQ")  # Default to USDA

        if not series1 or not series2:
            return jsonify({"error": "Both commodities must be provided"}), 400

        if window < 5:
            return jsonify({"error": "Window size must be at least 5 days"}), 400

        # Fetch data from the database
        result = db.session.query(
            ShippingPriceData.year,
            ShippingPriceData.day,
            ShippingPriceData.commodity,
            ShippingPriceData.price
        ).filter(
            ShippingPriceData.commodity.in_([series1, series2]),
            ShippingPriceData.price > 0,  # Ensure we only get valid prices
            ShippingPriceData.source == source  # Use selected source
        ).all()

        if not result:
            return jsonify({"error": "No data found for the selected commodities"}), 404

        # Create a DataFrame
        df = pd.DataFrame(result, columns=["year", "day", "commodity", "price"])

        # Generate a date column from year and day
        df["date"] = pd.to_datetime(
            df.apply(lambda row: f"{row.year}-{row.day}", axis=1),
            format="%Y-%j"
        )

        # Pivot the data for rolling correlation
        pivot_data = df.pivot_table(
            values="price",
            index="date",
            columns="commodity",
            aggfunc="mean"
        ).sort_index()

        # Check for minimum data points
        min_required_points = window * 2
        if len(pivot_data) < min_required_points:
            return jsonify({
                "error": f"Insufficient data points. Need at least {min_required_points} days of data for {window}-day window"
            }), 400

        # Calculate rolling correlation
        roll_corr_df = calculate_rolling_price_correlations(
            window=window,
            source_df=pivot_data,
            series1=series1,
            series2=series2
        )

        # Plot and serialize the rolling correlation chart
        fig_json = plot_rolling_price_correlations(
            roll_corr_df=roll_corr_df,
            series1=series1,
            series2=series2,
            window=window
        )

        return app.response_class(
            response=fig_json,
            status=200,
            mimetype='application/json'
        )

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        app.logger.error(f"Error generating shipping rolling correlations: {str(e)}")
        return jsonify({"error": "Internal Server Error", "details": str(e)}), 500



# Run the app
if __name__ == "__main__":
    app.run(debug=True)

# Initialize the cache extension
